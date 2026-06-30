"""Raw dataset file deduplication (Note 4) — orchestrator Phase 2, before extraction.

Runs **before** the shared header store is written (Phase 2b). Detects duplicate
candidate groups via filename normalization, applies Tier 1 (perfect column match
+ row count) and Tier 2 (header superset) rules without reading cell values.
When ``header_extraction.json`` is not yet present, resolves headers and row
counts via direct row-1 / count-only reads (``resolve_headers`` /
``resolve_row_count`` fail-soft fallback). Ambiguous groups route to
``audit/human_review/`` with count-only notes.
"""

from __future__ import annotations

import re
import shutil
from dataclasses import dataclass, field
from pathlib import Path

from scripts.audit.review_paths import excel_duplicate_review_path, human_review_root
from scripts.extraction.forms_manifest import check_forms_manifest
from scripts.extraction.header_store import (
    load_header_store,
    resolve_headers,
    resolve_row_count,
)
from scripts.extraction.io.file_discovery import SUPPORTED_TABULAR_EXTENSIONS, discover_files

__all__ = [
    "RawDedupReport",
    "dedup_raw_datasets",
    "normalize_dataset_stem",
]

_LOCK_PREFIXES = ("~$",)
_NUMERIC_SUFFIX_RE = re.compile(r"(\d+)$")


def normalize_dataset_stem(stem: str) -> str:
    """Normalize a dataset filename stem for duplicate grouping (Note 4)."""

    base = stem.strip().lower()
    base = re.sub(r"[_\-\s]+", "", base)
    while True:
        trimmed = _NUMERIC_SUFFIX_RE.sub("", base)
        if trimmed == base:
            break
        base = trimmed
    return base


def count_data_rows_only(path: Path) -> int:
    """Count data rows only — never reads cell values below the row-count boundary."""

    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as fh:
            total = sum(1 for _ in fh)
        return max(total - 1, 0)
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=False)
        try:
            sheet = workbook.active
            rows = sheet.max_row or 0
            return max(rows - 1, 0)
        finally:
            workbook.close()
    raise ValueError(f"Unsupported dataset format: {path.suffix}")


@dataclass
class RawDedupReport:
    study: str
    groups_scanned: int = 0
    auto_resolved: list[str] = field(default_factory=list)
    held_for_review: list[str] = field(default_factory=list)
    removed_paths: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    # Note 4: one structured, value-free record per AUTO-resolved merge decision
    # (what files / what action / why / counts) — written to the audit dir so
    # every dedup decision, not just held ones, leaves an audit trail.
    merge_decisions: list[dict] = field(default_factory=list)


def _file_fingerprint(path: Path, store: dict | None = None) -> tuple[list[str], int]:
    # Note 6: read headers + row count from the shared header-extraction store
    # when available; fall back to a direct row-1 / count-only read otherwise.
    return (
        resolve_headers(store, path.stem, path),
        resolve_row_count(store, path.stem, path),
    )


def _headers_superset(larger: list[str], smaller: list[str]) -> bool:
    if len(larger) < len(smaller):
        return False
    smaller_set = set(smaller)
    return smaller_set.issubset(set(larger)) and len(larger) > len(smaller)


def _write_review_note(path: Path, *, reason: str, files: list[Path], details: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [
        "# Dataset deduplication — human review required",
        "",
        f"**Reason:** {reason}",
        "",
        "## Files",
        *[f"- `{f.name}`" for f in files],
        "",
        "## Details (counts/headers only — no row values)",
        *[f"- {line}" for line in details],
        "",
        "Resolve via manifest/config, then re-run with `--resume-held`.",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _archive_duplicate(path: Path, *, archive_dir: Path) -> Path:
    archive_dir.mkdir(parents=True, exist_ok=True)
    dest = archive_dir / path.name
    if dest.exists():
        dest = archive_dir / f"{path.stem}__dup{path.suffix}"
    shutil.move(str(path), str(dest))
    return dest


def dedup_raw_datasets(
    study: str,
    *,
    datasets_dir: Path,
    audit_dir: Path,
    archive_dir: Path | None = None,
    run_dir: Path | None = None,
) -> RawDedupReport:
    """Deduplicate raw dataset files in *datasets_dir* (Note 4).

    When *run_dir* is given, headers/row-counts are read from the shared
    header-extraction store (Note 6) instead of re-opening each file, with a
    fail-soft fallback to a direct read.
    """

    report = RawDedupReport(study=study)
    if not datasets_dir.is_dir():
        report.errors.append(f"missing datasets dir: {datasets_dir}")
        return report

    store = load_header_store(run_dir)

    rejected = {name.lower() for name in check_forms_manifest(datasets_dir).rejected_files}
    archive_root = archive_dir or (audit_dir.parent / "tmp_dedup_archive")

    try:
        paths = discover_files(
            datasets_dir,
            extensions=SUPPORTED_TABULAR_EXTENSIONS,
            label="Dataset",
        )
    except (FileNotFoundError, ValueError):
        return report

    candidates = [
        p for p in paths if not p.name.startswith(_LOCK_PREFIXES) and p.name.lower() not in rejected
    ]

    groups: dict[str, list[Path]] = {}
    for path in candidates:
        key = normalize_dataset_stem(path.stem)
        groups.setdefault(key, []).append(path)

    report.groups_scanned = sum(1 for members in groups.values() if len(members) > 1)

    for norm_key, members in sorted(groups.items()):
        if len(members) < 2:
            continue

        fingerprints: dict[Path, tuple[list[str], int]] = {}
        for path in members:
            try:
                fingerprints[path] = _file_fingerprint(path, store)
            except (OSError, ValueError, StopIteration) as exc:
                report.errors.append(f"{path.name}: {type(exc).__name__}")
                fingerprints[path] = ([], -1)

        valid = {p: fp for p, fp in fingerprints.items() if fp[1] >= 0}
        if len(valid) < 2:
            continue

        header_sets = {p: fp[0] for p, fp in valid.items()}
        unique_headers = {tuple(h) for h in header_sets.values()}

        # Tier 1 — perfect column match (name, count, order)
        if len(unique_headers) == 1:
            row_counts = {p: valid[p][1] for p in valid}
            if len(set(row_counts.values())) == 1:
                keep = sorted(valid.keys(), key=lambda p: p.name)[0]
                archived_names: list[str] = []
                for path in sorted(valid.keys(), key=lambda p: p.name)[1:]:
                    archived = _archive_duplicate(path, archive_dir=archive_root / norm_key)
                    report.removed_paths.append(str(archived))
                    archived_names.append(path.name)
                report.auto_resolved.append(
                    f"{norm_key}: tier1 perfect duplicate — kept {keep.name}, "
                    f"archived {len(valid) - 1} file(s)"
                )
                report.merge_decisions.append(
                    {
                        "group": norm_key,
                        "tier": "tier1_perfect_duplicate",
                        "action": "dataset_duplicate_file",
                        "kept": keep.name,
                        "archived": archived_names,
                        "reason": "identical headers (name+count+order) and identical data-row count",
                        "headers_count": len(header_sets[keep]),
                        "data_rows": valid[keep][1],
                    }
                )
                continue

            review_path = excel_duplicate_review_path(audit_dir, norm_key)
            _write_review_note(
                review_path,
                reason="tier1_header_match_row_count_mismatch",
                files=members,
                details=[
                    f"{p.name}: headers={len(header_sets[p])}, data_rows={valid[p][1]}"
                    for p in sorted(valid)
                ],
            )
            report.held_for_review.append(norm_key)
            continue

        # Tier 2 — strict superset (one file contains all columns of every other)
        by_size = sorted(valid.keys(), key=lambda p: len(valid[p][0]), reverse=True)
        largest = by_size[0]
        largest_headers = valid[largest][0]
        if all(p is largest or _headers_superset(largest_headers, valid[p][0]) for p in by_size):
            archived_names = []
            for path in by_size[1:]:
                archived = _archive_duplicate(path, archive_dir=archive_root / norm_key)
                report.removed_paths.append(str(archived))
                archived_names.append(path.name)
            report.auto_resolved.append(
                f"{norm_key}: tier2 superset — kept {largest.name}, "
                f"archived {len(by_size) - 1} file(s)"
            )
            report.merge_decisions.append(
                {
                    "group": norm_key,
                    "tier": "tier2_superset",
                    "action": "dataset_duplicate_file",
                    "kept": largest.name,
                    "archived": archived_names,
                    "reason": "kept file's headers are a strict superset of every other file in the group",
                    "headers_count": len(largest_headers),
                    "data_rows": valid[largest][1],
                }
            )
            continue

        review_path = excel_duplicate_review_path(audit_dir, norm_key)
        _write_review_note(
            review_path,
            reason="ambiguous_duplicate_group",
            files=members,
            details=[
                f"{p.name}: headers={len(header_sets[p])}, data_rows={valid[p][1]}"
                for p in sorted(valid)
            ],
        )
        report.held_for_review.append(norm_key)

    human_review_root(audit_dir).mkdir(parents=True, exist_ok=True)
    if report.merge_decisions:
        _write_merge_report(audit_dir, study, report.merge_decisions)
    return report


def _write_merge_report(audit_dir: Path, study: str, decisions: list[dict]) -> Path:
    """Write the value-free auto-resolved dedup audit record (Note 4 + Note 17).

    Every AUTO-resolved merge — not just held cases — leaves an on-disk audit
    trail (what files, what action, why, counts) in the audit zone. Counts and
    file names only; never a row value.
    """
    import json

    out_dir = audit_dir / "dataset_dedup"
    out_dir.mkdir(parents=True, exist_ok=True)
    json_path = out_dir / "merge_report.json"
    json_path.write_text(
        json.dumps({"study": study, "auto_resolved_merges": decisions}, indent=2, sort_keys=True)
        + "\n",
        encoding="utf-8",
    )
    lines = [
        "# Dataset duplicate merge report (auto-resolved)",
        "",
        f"study: {study}",
        f"auto-resolved decisions: {len(decisions)}",
        "",
    ]
    for d in decisions:
        archived = ", ".join(f"`{a}`" for a in d["archived"]) or "(none)"
        lines += [
            f"## {d['group']} — {d['tier']}",
            f"- kept: `{d['kept']}`",
            f"- archived: {archived}",
            f"- reason: {d['reason']}",
            f"- headers: {d['headers_count']} · data_rows: {d['data_rows']}",
            "",
        ]
    (out_dir / "dataset_duplicate_merge_report.md").write_text(
        "\n".join(lines) + "\n", encoding="utf-8"
    )
    return json_path
