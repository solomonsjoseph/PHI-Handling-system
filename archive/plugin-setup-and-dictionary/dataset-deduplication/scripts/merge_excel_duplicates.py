#!/usr/bin/env python3
"""Lossless Excel duplicate merge helper.

This helper is intentionally quiet about cell values. It may process workbook
values internally to create the merged workbook, but reports only metadata,
counts, headers, and provenance coordinates.
"""

from __future__ import annotations

import argparse
import csv
import shutil
import zipfile
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


def _infer_study(path: Path) -> str:
    parts = path.parts
    for idx in range(len(parts) - 3):
        if parts[idx : idx + 2] == ("data", "raw") and parts[idx + 3] == "datasets":
            return parts[idx + 2]
    return "UNKNOWN_STUDY"


def _default_dataset_name(path: Path) -> str:
    return path.stem.removeprefix("~$")


@dataclass
class SheetPlan:
    source_file: Path
    sheet_name: str
    headers: list[str]
    rows_including_header: int
    data_rows: int
    role: str


@dataclass
class InvalidSource:
    source_file: Path
    role: str
    reason: str


@dataclass
class MergeStats:
    source_sheets: list[SheetPlan] = field(default_factory=list)
    invalid_sources: list[InvalidSource] = field(default_factory=list)
    union_headers: list[str] = field(default_factory=list)
    backup_path: Path | None = None
    removed_active_branch_files: list[str] = field(default_factory=list)
    output_rows: int = 0
    appended_rows: int = 0
    collapsed_exact_duplicate_rows: int = 0
    preserved_main_rows: int = 0
    provenance_rows: list[list[Any]] = field(default_factory=list)


@dataclass
class BatchItem:
    dataset: str
    main: Path | None
    branches: list[Path]
    status: str
    report: Path | None = None
    provenance: Path | None = None
    merged_workbook: Path | None = None
    reason: str | None = None
    stats: MergeStats | None = None


class MergeNotSafeError(Exception):
    """Raised when a candidate pair must be routed to human review."""

    def __init__(self, reason: str, review_path: Path, stats: MergeStats) -> None:
        super().__init__(reason)
        self.reason = reason
        self.review_path = review_path
        self.stats = stats


def _load_sheet_plans(path: Path, role: str) -> tuple[list[SheetPlan], InvalidSource | None]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except (InvalidFileException, OSError, ValueError, KeyError, zipfile.BadZipFile) as exc:
        return [], InvalidSource(path, role, f"{type(exc).__name__}: {exc}")

    plans: list[SheetPlan] = []
    try:
        for worksheet in workbook.worksheets:
            rows = worksheet.max_row or 0
            cols = worksheet.max_column or 0
            if rows < 1 or cols < 1:
                continue
            header_row = next(
                worksheet.iter_rows(min_row=1, max_row=1, values_only=True),
                (),
            )
            headers = [str(value).strip() for value in header_row if value not in (None, "")]
            if not headers:
                continue
            if len(set(headers)) != len(headers):
                return [], InvalidSource(
                    path, role, f"duplicate headers in sheet {worksheet.title}"
                )
            plans.append(
                SheetPlan(
                    source_file=path,
                    sheet_name=worksheet.title,
                    headers=headers,
                    rows_including_header=rows,
                    data_rows=max(rows - 1, 0),
                    role=role,
                )
            )
    finally:
        workbook.close()
    return plans, None


def _build_union_headers(plans: list[SheetPlan]) -> list[str]:
    headers: list[str] = []
    seen: set[str] = set()
    for plan in plans:
        for header in plan.headers:
            if header not in seen:
                headers.append(header)
                seen.add(header)
    return headers


def _aligned_row(
    values: tuple[Any, ...], source_headers: list[str], union_headers: list[str]
) -> list[Any]:
    source_by_header = dict(zip(source_headers, values, strict=False))
    return [source_by_header.get(header) for header in union_headers]


def _copy_cell(source: Any, target: Any) -> None:
    target.value = source.value
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.font:
        target.font = copy(source.font)
    if source.fill:
        target.fill = copy(source.fill)
    if source.border:
        target.border = copy(source.border)
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.protection:
        target.protection = copy(source.protection)
    if source.hyperlink:
        target._hyperlink = copy(source.hyperlink)
    if source.comment:
        target.comment = copy(source.comment)


def _cells_by_header(cells: tuple[Any, ...], headers: list[str]) -> dict[str, Any]:
    return dict(zip(headers, cells, strict=False))


def _row_key_from_cells(
    cells: tuple[Any, ...], headers: list[str], union_headers: list[str]
) -> tuple[Any, ...]:
    cells_by_header = _cells_by_header(cells, headers)
    return tuple(
        cells_by_header[header].value if header in cells_by_header else None
        for header in union_headers
    )


def _write_provenance_csv(path: Path, stats: MergeStats) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            ["output_row", "source_file", "source_sheet", "source_row", "role", "action"]
        )
        writer.writerows(stats.provenance_rows)


def _unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    for index in range(1, 10_000):
        candidate = path.with_name(f"{path.stem}__backup_{index}{path.suffix}")
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"Unable to find an unused backup path for {path}")


def _unique_dir(path: Path) -> Path:
    if not path.exists():
        return path
    for index in range(1, 10_000):
        candidate = path.with_name(f"{path.name}__backup_{index}")
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"Unable to find an unused backup directory for {path}")


def _same_path(left: Path, right: Path) -> bool:
    try:
        return left.resolve() == right.resolve()
    except OSError:
        return left.absolute() == right.absolute()


def _temporary_merge_path(out_path: Path) -> Path:
    for index in range(1, 10_000):
        candidate = out_path.with_name(f".{out_path.stem}.merge_tmp_{index}{out_path.suffix}")
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"Unable to find an unused temporary merge path for {out_path}")


def _project_dataset_path(artifact_root: Path, study: str, filename: str) -> Path:
    return artifact_root / "data" / "raw" / study / "datasets" / filename


def _project_dataset_dir(artifact_root: Path, study: str) -> Path:
    return artifact_root / "data" / "raw" / study / "datasets"


def _project_backup_dir(artifact_root: Path, study: str) -> Path:
    return artifact_root / "data" / "raw" / study / "_dataset"


def _prepare_project_dataset_dirs(
    source_dataset_dir: Path, artifact_root: Path, study: str
) -> tuple[Path, Path]:
    """Snapshot the full source datasets folder and return active dataset dir.

    In production, active_dir is the existing source `datasets/` directory.
    In scratch mode, active_dir is a full working copy under artifact_root.
    """
    backup_dir = _unique_dir(_project_backup_dir(artifact_root, study))
    backup_dir.parent.mkdir(parents=True, exist_ok=True)
    shutil.copytree(source_dataset_dir, backup_dir)

    active_dir = _project_dataset_dir(artifact_root, study)
    if _same_path(source_dataset_dir, active_dir):
        return backup_dir, source_dataset_dir
    if active_dir.exists():
        raise SystemExit(
            f"Active scratch dataset directory already exists: {active_dir}. "
            "Use a fresh --artifact-root for a clean duplicate-handling run."
        )
    active_dir.parent.mkdir(parents=True, exist_ok=True)
    shutil.copytree(source_dataset_dir, active_dir)
    return backup_dir, active_dir


def _human_review_path(artifact_root: Path, study: str, dataset: str) -> Path:
    from scripts.audit.review_paths import excel_duplicate_review_path

    audit_dir = artifact_root / "output" / study / "audit"
    return excel_duplicate_review_path(audit_dir, dataset)


def _header_relationship(
    left: SheetPlan, right: SheetPlan
) -> tuple[str, int, int, int, float, float]:
    left_headers = set(left.headers)
    right_headers = set(right.headers)
    common = len(left_headers & right_headers)
    left_only = len(left_headers - right_headers)
    right_only = len(right_headers - left_headers)
    smaller = min(len(left_headers), len(right_headers))
    union = len(left_headers | right_headers)
    containment = common / smaller if smaller else 0.0
    jaccard = common / union if union else 0.0
    if left.headers == right.headers:
        relationship = "exact_ordered_headers"
    elif left_headers == right_headers:
        relationship = "same_header_set_different_order"
    elif left_headers < right_headers:
        relationship = f"{right.source_file.name} header_superset"
    elif right_headers < left_headers:
        relationship = f"{left.source_file.name} header_superset"
    else:
        relationship = "partial_overlap"
    return relationship, common, left_only, right_only, containment, jaccard


def _merge_safety_reason(stats: MergeStats) -> str | None:
    main_plans = [plan for plan in stats.source_sheets if plan.role == "main"]
    branch_plans = [plan for plan in stats.source_sheets if plan.role == "branch"]
    for invalid in stats.invalid_sources:
        if invalid.role == "branch" and not invalid.source_file.name.startswith("~$"):
            return "branch workbook could not be opened; route to human review before merging"
    if not branch_plans:
        return None

    for plan in main_plans:
        if plan.headers != stats.union_headers:
            return (
                "main workbook is not the full merge schema; choose a clear header superset "
                "as main or route the candidate to human review"
            )

    union_header_set = set(stats.union_headers)
    for plan in branch_plans:
        if not set(plan.headers).issubset(union_header_set):
            return "branch headers cannot be aligned to the current merge schema"
    return None


def _write_human_review_report(path: Path, stats: MergeStats, reason: str) -> None:
    main_plans = [plan for plan in stats.source_sheets if plan.role == "main"]
    branch_plans = [plan for plan in stats.source_sheets if plan.role == "branch"]
    lines = [
        "# Human Review: Excel Duplicate Candidate",
        "",
        "Boundary: this report uses filenames, sheet names, row counts, header counts, "
        "and header-overlap metadata only. No dataset row values were written to this report.",
        "",
        "## Decision",
        "",
        "- status: `not_handled_by_auto_merge`",
        "- action_taken: no merged workbook was created",
        f"- reason: {reason}",
        "- required_next_step: human review with study/form context before any merge decision",
        "",
        "## Candidate Sheets",
        "",
        "| file | sheet | role | data rows | headers |",
        "| --- | --- | --- | ---: | ---: |",
    ]
    lines.extend(
        f"| `{plan.source_file.name}` | `{plan.sheet_name}` | {plan.role} | "
        f"{plan.data_rows} | {len(plan.headers)} |"
        for plan in stats.source_sheets
    )
    if stats.invalid_sources:
        lines.extend(["", "## Invalid Or Skipped Sources", ""])
        lines.extend(
            f"- `{invalid.source_file.name}` ({invalid.role}): {invalid.reason}"
            for invalid in stats.invalid_sources
        )
    if main_plans and branch_plans:
        lines.extend(
            [
                "",
                "## Header Relationships",
                "",
                "| main file | branch file | relationship | common headers | main-only headers | "
                "branch-only headers | containment | jaccard |",
                "| --- | --- | --- | ---: | ---: | ---: | ---: | ---: |",
            ]
        )
        for main_plan in main_plans:
            for branch_plan in branch_plans:
                relationship, common, main_only, branch_only, containment, jaccard = (
                    _header_relationship(main_plan, branch_plan)
                )
                lines.append(
                    f"| `{main_plan.source_file.name}` | `{branch_plan.source_file.name}` | "
                    f"`{relationship}` | {common} | {main_only} | {branch_only} | "
                    f"{containment:.2f} | {jaccard:.2f} |"
                )
    lines.extend(
        [
            "",
            "## Integrity Decision",
            "",
            "The skill did not create a backup or replace a dataset workbook because the "
            "schema evidence was not strong enough for an automatic lossless merge.",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")


def _write_markdown_report(
    report_path: Path, out_path: Path, provenance_path: Path, stats: MergeStats
) -> None:
    lines = [
        "# Excel Duplicate Merge Report",
        "",
        "Boundary: workbook values were copied only into the merged workbook. This audit report "
        "contains counts, headers, and provenance metadata only.",
        "",
        f"- dataset_workbook: `{out_path}`",
        f"- raw_dataset_snapshot: `{stats.backup_path}`",
        f"- provenance_csv: `{provenance_path}`",
        f"- source_sheet_count: {len(stats.source_sheets)}",
        f"- invalid_source_count: {len(stats.invalid_sources)}",
        f"- output_data_rows: {stats.output_rows}",
        f"- preserved_main_rows: {stats.preserved_main_rows}",
        f"- appended_rows: {stats.appended_rows}",
        f"- collapsed_exact_duplicate_rows: {stats.collapsed_exact_duplicate_rows}",
        f"- removed_active_branch_file_count: {len(stats.removed_active_branch_files)}",
        f"- header_count: {len(stats.union_headers)}",
        "- integrity_mode: full raw datasets folder snapshotted under `_dataset`; "
        "safe merged workbook written to the active `datasets` path; "
        "branch rows appended after existing main rows",
        "",
        "## Headers",
        "",
    ]
    lines.extend(f"- `{header}`" for header in stats.union_headers)
    lines.extend(["", "## Source Sheets", ""])
    for plan in stats.source_sheets:
        lines.extend(
            [
                f"### {plan.source_file.name} / {plan.sheet_name}",
                "",
                f"- role: {plan.role}",
                f"- rows_including_header: {plan.rows_including_header}",
                f"- data_rows: {plan.data_rows}",
                f"- header_count: {len(plan.headers)}",
                "",
            ]
        )
    if stats.invalid_sources:
        lines.extend(["## Invalid Or Skipped Sources", ""])
        for invalid in stats.invalid_sources:
            lines.extend(
                [
                    f"- `{invalid.source_file.name}` ({invalid.role}): {invalid.reason}",
                ]
            )
        lines.append("")
    if stats.removed_active_branch_files:
        lines.extend(["## Removed Active Branch Files", ""])
        lines.extend(f"- `{name}`" for name in stats.removed_active_branch_files)
        lines.append("")
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text("\n".join(lines), encoding="utf-8")


def merge_workbooks(
    main: Path,
    branches: list[Path],
    out_path: Path,
    report_path: Path,
    provenance_path: Path,
    raw_snapshot_path: Path | None = None,
    review_path: Path | None = None,
) -> MergeStats:
    stats = MergeStats()
    inputs = [(main, "main"), *((branch, "branch") for branch in branches)]

    for path, role in inputs:
        plans, invalid = _load_sheet_plans(path, role)
        if invalid is not None:
            stats.invalid_sources.append(invalid)
        stats.source_sheets.extend(plans)

    if not any(plan.role == "main" for plan in stats.source_sheets):
        raise SystemExit("No valid main workbook/sheet was available to merge.")

    stats.union_headers = _build_union_headers(stats.source_sheets)

    safety_reason = _merge_safety_reason(stats)
    if safety_reason is not None:
        if review_path is None:
            from scripts.audit.review_paths import excel_duplicate_review_path

            review_path = excel_duplicate_review_path(
                report_path.parent.parent.parent,
                report_path.parent.name,
            )
        _write_human_review_report(review_path, stats, safety_reason)
        raise MergeNotSafeError(safety_reason, review_path, stats)

    # Preserve the main workbook exactly as the merge base, including dates,
    # number formats, formulas, widths, styles, workbook metadata, and sheets.
    out_path.parent.mkdir(parents=True, exist_ok=True)
    work_path = _temporary_merge_path(out_path)
    shutil.copy2(main, work_path)
    stats.backup_path = raw_snapshot_path

    main_plans = [plan for plan in stats.source_sheets if plan.role == "main"]
    branch_plans = [plan for plan in stats.source_sheets if plan.role == "branch"]
    try:
        if not branch_plans:
            for plan in main_plans:
                for source_row_idx in range(2, plan.rows_including_header + 1):
                    stats.output_rows += 1
                    stats.preserved_main_rows += 1
                    stats.provenance_rows.append(
                        [
                            source_row_idx,
                            plan.source_file.name,
                            plan.sheet_name,
                            source_row_idx,
                            plan.role,
                            "preserved_main",
                        ]
                    )
        else:
            workbook_out = load_workbook(work_path)
            try:
                seen_rows: dict[tuple[Any, ...], int] = {}

                for plan in main_plans:
                    worksheet = workbook_out[plan.sheet_name]
                    for source_row_idx, cells in enumerate(
                        worksheet.iter_rows(min_row=2, max_col=len(plan.headers)),
                        start=2,
                    ):
                        row_key = _row_key_from_cells(cells, plan.headers, stats.union_headers)
                        seen_rows.setdefault(row_key, source_row_idx)
                        stats.output_rows += 1
                        stats.preserved_main_rows += 1
                        stats.provenance_rows.append(
                            [
                                source_row_idx,
                                plan.source_file.name,
                                plan.sheet_name,
                                source_row_idx,
                                plan.role,
                                "preserved_main",
                            ]
                        )

                for plan in branch_plans:
                    target_sheet = (
                        workbook_out[plan.sheet_name]  # noqa: SIM401
                        if plan.sheet_name in workbook_out
                        else workbook_out.active
                    )
                    branch_workbook = load_workbook(plan.source_file)
                    try:
                        branch_sheet = branch_workbook[plan.sheet_name]
                        for source_row_idx, cells in enumerate(
                            branch_sheet.iter_rows(min_row=2, max_col=len(plan.headers)),
                            start=2,
                        ):
                            row_key = _row_key_from_cells(cells, plan.headers, stats.union_headers)
                            if row_key in seen_rows:
                                stats.collapsed_exact_duplicate_rows += 1
                                stats.provenance_rows.append(
                                    [
                                        seen_rows[row_key],
                                        plan.source_file.name,
                                        plan.sheet_name,
                                        source_row_idx,
                                        plan.role,
                                        "collapsed_exact_duplicate",
                                    ]
                                )
                                continue
                            target_row = target_sheet.max_row + 1
                            target_sheet.row_dimensions[
                                target_row
                            ].height = branch_sheet.row_dimensions[source_row_idx].height
                            source_cells = _cells_by_header(cells, plan.headers)
                            for column_idx, header in enumerate(stats.union_headers, start=1):
                                source_cell = source_cells.get(header)
                                if source_cell is not None:
                                    _copy_cell(
                                        source_cell,
                                        target_sheet.cell(row=target_row, column=column_idx),
                                    )
                            seen_rows[row_key] = target_row
                            stats.output_rows += 1
                            stats.appended_rows += 1
                            stats.provenance_rows.append(
                                [
                                    target_row,
                                    plan.source_file.name,
                                    plan.sheet_name,
                                    source_row_idx,
                                    plan.role,
                                    "appended_branch",
                                ]
                            )
                    finally:
                        branch_workbook.close()

                workbook_out.save(work_path)
            finally:
                workbook_out.close()

        shutil.move(str(work_path), out_path)
        if raw_snapshot_path is not None:
            for branch in branches:
                if (
                    branch.exists()
                    and not _same_path(branch, out_path)
                    and _same_path(branch.parent, out_path.parent)
                ):
                    branch.unlink()
                    stats.removed_active_branch_files.append(branch.name)
    except BaseException:
        if work_path.exists():
            work_path.unlink()
        raise

    _write_provenance_csv(provenance_path, stats)
    _write_markdown_report(report_path, out_path, provenance_path, stats)
    return stats


def discover_lock_temp_groups(dataset_dir: Path) -> list[tuple[str, Path | None, list[Path]]]:
    """Find valid-workbook + Excel lock/temp sibling groups."""
    groups: list[tuple[str, Path | None, list[Path]]] = []
    for branch in sorted(dataset_dir.glob("~$*.xlsx")):
        dataset = _default_dataset_name(branch)
        main = dataset_dir / f"{dataset}.xlsx"
        groups.append((dataset, main if main.exists() else None, [branch]))
    return groups


def _write_batch_report(report_path: Path, items: list[BatchItem]) -> None:
    lines = [
        "# Excel Duplicate Batch Merge Report",
        "",
        "Boundary: this report contains filenames, statuses, counts, and audit paths only.",
        "",
        "| dataset | status | dataset workbook | audit report | reason |",
        "| --- | --- | --- | --- | --- |",
    ]
    lines.extend(
        "| "
        + " | ".join(
            [
                f"`{item.dataset}`",
                item.status,
                f"`{item.merged_workbook}`" if item.merged_workbook else "",
                f"`{item.report}`" if item.report else "",
                item.reason or "",
            ]
        )
        + " |"
        for item in items
    )
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text("\n".join(lines), encoding="utf-8")


def merge_lock_temp_groups(
    *,
    dataset_dir: Path,
    study: str,
    artifact_root: Path,
    batch_report: Path | None,
) -> list[BatchItem]:
    items: list[BatchItem] = []
    raw_snapshot_path, active_dataset_dir = _prepare_project_dataset_dirs(
        dataset_dir, artifact_root, study
    )
    for dataset, main, branches in discover_lock_temp_groups(active_dataset_dir):
        if main is None:
            items.append(
                BatchItem(
                    dataset=dataset,
                    main=None,
                    branches=branches,
                    status="skipped",
                    reason="no matching main workbook",
                )
            )
            continue

        out_path = _project_dataset_path(artifact_root, study, main.name)
        audit_dir = artifact_root / "output" / study / "audit" / "datasets" / dataset
        report_path = audit_dir / "merge_report.md"
        provenance_path = audit_dir / "merge_provenance.csv"
        review_path = _human_review_path(artifact_root, study, dataset)
        try:
            stats = merge_workbooks(
                main,
                branches,
                out_path,
                report_path,
                provenance_path,
                raw_snapshot_path,
                review_path,
            )
        except MergeNotSafeError as exc:
            items.append(
                BatchItem(
                    dataset=dataset,
                    main=main,
                    branches=branches,
                    status="human_review_required",
                    report=exc.review_path,
                    reason=exc.reason,
                    stats=exc.stats,
                )
            )
            continue
        items.append(
            BatchItem(
                dataset=dataset,
                main=main,
                branches=branches,
                status="merged",
                report=report_path,
                provenance=provenance_path,
                merged_workbook=out_path,
                stats=stats,
            )
        )

    if batch_report is not None:
        _write_batch_report(batch_report, items)
    return items


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--main", type=Path)
    parser.add_argument("--branch", action="append", default=[], type=Path)
    parser.add_argument(
        "--dataset-dir",
        type=Path,
        help="Scan a datasets directory for Excel lock/temp siblings such as ~$10_TST.xlsx.",
    )
    parser.add_argument("--study")
    parser.add_argument("--dataset")
    parser.add_argument(
        "--artifact-root",
        type=Path,
        default=Path("."),
        help="Root for project-shaped artifacts. Defaults to repo root.",
    )
    parser.add_argument("--out", type=Path)
    parser.add_argument("--audit-dir", type=Path)
    parser.add_argument("--report", type=Path)
    parser.add_argument("--provenance", type=Path)
    parser.add_argument("--review-report", type=Path)
    parser.add_argument("--batch-report", type=Path)
    args = parser.parse_args()

    if args.dataset_dir is not None:
        study = args.study or _infer_study(args.dataset_dir)
        batch_report = args.batch_report or (
            args.artifact_root / "output" / study / "audit" / "dataset_duplicate_merge_report.md"
        )
        items = merge_lock_temp_groups(
            dataset_dir=args.dataset_dir,
            study=study,
            artifact_root=args.artifact_root,
            batch_report=batch_report,
        )
        print(f"batch_report={batch_report}")
        print(f"groups={len(items)}")
        print(f"merged_groups={sum(1 for item in items if item.status == 'merged')}")
        print(f"skipped_groups={sum(1 for item in items if item.status == 'skipped')}")
        print(
            "human_review_groups="
            f"{sum(1 for item in items if item.status == 'human_review_required')}"
        )
        return

    if args.main is None:
        raise SystemExit("--main is required unless --dataset-dir is used.")

    study = args.study or _infer_study(args.main)
    dataset = args.dataset or _default_dataset_name(args.main)
    raw_snapshot_path, active_dataset_dir = _prepare_project_dataset_dirs(
        args.main.parent, args.artifact_root, study
    )
    main_path = active_dataset_dir / args.main.name
    branch_paths = [active_dataset_dir / branch.name for branch in args.branch]
    out_path = args.out or _project_dataset_path(args.artifact_root, study, args.main.name)
    audit_dir = (
        args.audit_dir or args.artifact_root / "output" / study / "audit" / "datasets" / dataset
    )
    report_path = args.report or audit_dir / "merge_report.md"
    provenance_path = args.provenance or audit_dir / "merge_provenance.csv"
    review_path = args.review_report or _human_review_path(args.artifact_root, study, dataset)

    try:
        stats = merge_workbooks(
            main_path,
            branch_paths,
            out_path,
            report_path,
            provenance_path,
            raw_snapshot_path,
            review_path,
        )
    except MergeNotSafeError as exc:
        print("status=human_review_required")
        print(f"review_report={exc.review_path}")
        print(f"reason={exc.reason}")
        return
    print(f"dataset_workbook={out_path}")
    print(f"raw_dataset_snapshot={stats.backup_path}")
    print(f"report={report_path}")
    print(f"provenance={provenance_path}")
    print(f"output_data_rows={stats.output_rows}")
    print(f"preserved_main_rows={stats.preserved_main_rows}")
    print(f"appended_rows={stats.appended_rows}")
    print(f"collapsed_exact_duplicate_rows={stats.collapsed_exact_duplicate_rows}")
    print(f"invalid_source_count={len(stats.invalid_sources)}")


if __name__ == "__main__":
    main()
