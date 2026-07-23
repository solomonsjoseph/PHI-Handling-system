"""Build a deterministic, intake-manifest/v3-ready source tree for the
standalone-pipeline stress test. Writes synthetic PHI values only (never
real data).

    python -m harness.make_stress_fixtures --out tmp/stress-source [--seed 42]

Writes ``<out>/`` -- a complete mandatory-component package
(``datasets/``, ``forms/``, ``data_dictionary/``, ``mappings/``) built
entirely from accepted formats (``.csv``, single-sheet ``.xlsx``, ``.pdf``,
plus a legacy ``.xls`` dataset) so ``intake_add`` reaches ``status ==
"ready"`` and the whole tree organizes/runs -- and
``<out>.manifest/stress_manifest.json``: a complete per-entry filesystem
snapshot (type, sha256, size, mode, mtime_ns, uid, gid, symlink target;
``atime`` deliberately excluded) of every regular file, directory, and
symlink under ``<out>/`` at build time, consumed by
``harness/spec_check.py``'s ``source_immutability`` check. The manifest
deliberately lives OUTSIDE ``<out>/`` so it is never itself picked up by
``intake_add``'s walk.

``build_review_required_fixtures`` builds a SEPARATE, smaller source tree
whose deliberately unsupported/malformed/symlinked files each trip one
fixed intake-preflight review reason. Under intake-manifest/v3 a review
item anywhere blocks the WHOLE study (``status == "review_required"``,
``organize``/``run`` refuse to proceed) -- unlike the pre-v3 pipeline,
there is no "mostly good, a few files reviewed" partial outcome, so this
tree is never organized/run and is kept structurally separate from the
canonical ready package above.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import shutil
import stat
from pathlib import Path
from typing import Any

from faker import Faker
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

MANIFEST_FILENAME = "stress_manifest.json"

_IMMUTABILITY_FIELDS = ("type", "mode", "size", "mtime_ns", "uid", "gid", "sha256", "symlink_target")


def _fake(seed: int) -> Faker:
    fk = Faker()
    Faker.seed(seed)
    return fk


# --- accepted-format writers (mandatory-component package) ---------------------------------


def _write_csv(path: Path, headers: list[str], rows: list[list[Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(headers)
        writer.writerows(rows)


def _write_single_sheet_xlsx(
    path: Path, headers: list[str], rows: list[list[Any]], *, sheet_title: str = "Sheet1"
) -> None:
    from openpyxl import Workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title
    ws.append(headers)
    for row in rows:
        ws.append(row)
    wb.save(str(path))


def _write_xls(path: Path, fk: Faker) -> bool:
    """Genuine legacy .xls via xlwt when installable; returns False (caller
    ships a mislabeled fallback file instead) when xlwt is unavailable."""
    try:
        import xlwt
    except ImportError:
        return False
    wb = xlwt.Workbook()
    ws = wb.add_sheet("Legacy")
    for col, header in enumerate(["SUBJID", "SITE_CODE"]):
        ws.write(0, col, header)
    for i in range(5):
        ws.write(i + 1, 0, f"XLS-{i:03d}")
        ws.write(i + 1, 1, f"SITE-{i % 2}")
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return True


def _write_mislabeled_xls(path: Path) -> None:
    """A file with a .xls extension that is not real BIFF -- exercises the
    organizer's fail-closed 'unreadable/mislabeled .xls' review-bucket
    routing. .xls carries no intake-time content check (only the closed
    suffix matrix), so this is accepted at intake and only fails, per-file
    and non-blocking, when the organizer tries to actually parse it."""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(b"this is not a real xls workbook, just text bytes\n")


def _write_pdf_with_table(path: Path, fk: Faker) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    doc = SimpleDocTemplate(str(path), pagesize=letter)
    data = [["SUBJID", "RESULT"]] + [
        [f"PDF-{i:03d}", fk.random_element(["Negative", "Positive"])] for i in range(5)
    ]
    table = Table(data)
    table.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 1, colors.black)]))
    doc.build([table])


def _write_plain_pdf(path: Path, lines: list[str]) -> None:
    """A PDF with no extractable table -- exercises the organizer's
    non-blocking 'pdf-no-extractable-table' pdf_roles branch. Forms/ content
    gets zero intake-time structural validation, so this is accepted at
    intake regardless."""
    path.parent.mkdir(parents=True, exist_ok=True)
    c = canvas.Canvas(str(path), pagesize=letter)
    y = 720
    for line in lines:
        c.drawString(72, y, line)
        y -= 20
    c.save()


def _write_phi_in_unexpected_columns_csv(path: Path, fk: Faker, n: int = 6) -> list[dict[str, str]]:
    """SSN-shaped values under an innocuous 'NOTES' header, phone-shaped
    values under 'COMMENT' -- the value-profiler ESCALATION stress case.
    Returns the planted rows so the test suite can assert none of these
    exact values leak into published output."""
    rows: list[dict[str, str]] = []
    for i in range(n):
        rows.append({"SUBJID": f"PH-{i:03d}", "NOTES": fk.ssn(), "COMMENT": fk.numerify("###-###-####")})
    _write_csv(path, ["SUBJID", "NOTES", "COMMENT"], [[r["SUBJID"], r["NOTES"], r["COMMENT"]] for r in rows])
    return rows


# --- deliberately-invalid writers (review-required package only) ---------------------------


def _write_malformed_xlsx(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(b"PK\x03\x04not-a-real-zip-central-directory")


def _write_multi_sheet_xlsx(path: Path, fk: Faker) -> None:
    from openpyxl import Workbook

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["SUBJID", "AGE"])
    ws1.append([f"MS-{i:03d}" for i in range(1)] + [fk.random_int(18, 85)])
    ws2 = wb.create_sheet("Sheet2")
    ws2.append(["SUBJID", "AGE"])
    wb.save(str(path))


def _write_unknown_dat(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("not a recognized structured format\n", encoding="utf-8")


def _write_json_list(path: Path, rows: list[dict[str, Any]]) -> None:
    """Not an accepted dataset format under intake-manifest/v3 (only
    ``.csv``/``.xls``/``.xlsx`` are) -- deliberately kept ONLY as an
    explicit unsupported-format/``_unclassified`` review case."""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(rows), encoding="utf-8")


def _write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    """Not an accepted dataset format under intake-manifest/v3 -- see
    :func:`_write_json_list`."""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(json.dumps(r) for r in rows) + "\n", encoding="utf-8")


# --- complete, atime-excluding per-entry filesystem snapshot --------------------------------


def _sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def _snapshot_entry(path: Path) -> dict[str, Any]:
    info = path.lstat()
    mode = info.st_mode
    entry: dict[str, Any] = {
        "mode": stat.S_IMODE(mode),
        "size": info.st_size,
        "mtime_ns": info.st_mtime_ns,
        "uid": info.st_uid,
        "gid": info.st_gid,
        "sha256": None,
        "symlink_target": None,
    }
    if stat.S_ISLNK(mode):
        entry["type"] = "symlink"
        entry["symlink_target"] = os.readlink(path)
    elif stat.S_ISREG(mode):
        entry["type"] = "file"
        entry["sha256"] = _sha256_file(path)
    elif stat.S_ISDIR(mode):
        entry["type"] = "dir"
    else:
        entry["type"] = "other"
    return entry


def _snapshot_tree(root: Path) -> dict[str, dict[str, Any]]:
    """Complete entry set (files, directories, symlinks) under *root* at
    build time, keyed by POSIX-relative path -- everything
    ``harness.spec_check``'s ``source_immutability`` check compares."""
    entries: dict[str, dict[str, Any]] = {}
    for path in sorted(root.rglob("*")):
        entries[str(path.relative_to(root))] = _snapshot_entry(path)
    return entries


# --- canonical, mandatory-component, ready-producing package --------------------------------


def build_stress_fixtures(out_dir: Path, *, seed: int = 42) -> dict[str, Any]:
    """Build the deterministic mandatory-component source tree at *out_dir*
    (``datasets/`` + ``forms/`` + ``data_dictionary/`` + ``mappings/``, all
    accepted formats) and its complete-entry-set immutability manifest.
    Idempotent: wipes and rebuilds *out_dir* every call. Reaches
    intake-manifest/v3 ``status == "ready"`` and organizes/runs end to end."""
    fk = _fake(seed)

    if out_dir.exists():
        shutil.rmtree(out_dir)
    out_dir.mkdir(parents=True)

    # -- datasets/ --------------------------------------------------------------------------
    labs_headers = ["SUBJID", "VISITDAT", "WEIGHT_KG"]
    labs_rows = [
        [f"CSV-{i:03d}", fk.date_between(start_date="-2y").isoformat(), fk.random_int(45, 95)]
        for i in range(6)
    ]
    _write_csv(out_dir / "datasets" / "labs.csv", labs_headers, labs_rows)
    # Same bytes, nested subdirectory, different name -- duplicate content
    # and nested directories must both survive intake as distinct entries.
    _write_csv(out_dir / "datasets" / "batch_2026" / "site_04" / "labs_dup.csv", labs_headers, labs_rows)

    _write_single_sheet_xlsx(
        out_dir / "datasets" / "screening.xlsx",
        ["SUBJID", "AGE", "SEX", "SITE_CODE"],
        [[f"CRF-{i:03d}", fk.random_int(18, 85), fk.random_element(["M", "F"]), f"SITE-{i % 3}"] for i in range(8)],
        sheet_title="Screening",
    )

    xls_path = out_dir / "datasets" / "legacy_site.xls"
    if not _write_xls(xls_path, fk):
        _write_mislabeled_xls(xls_path)

    planted_unexpected_phi_rows = _write_phi_in_unexpected_columns_csv(out_dir / "datasets" / "site_notes.csv", fk)

    # -- forms/ -------------------------------------------------------------------------------
    _write_pdf_with_table(out_dir / "forms" / "consent_table.pdf", fk)
    _write_plain_pdf(out_dir / "forms" / "screening_form.pdf", ["Screening Form", "Subject ID: ____________"])

    # -- data_dictionary/ ---------------------------------------------------------------------
    dict_rows = [
        ["SUBJID", "Subject identifier"],
        ["VISITDAT", "Visit date"],
        ["WEIGHT_KG", "Weight in kilograms"],
        ["AGE", "Age in years"],
        ["SEX", "Sex at enrollment"],
        ["SITE_CODE", "Enrolling site code"],
        ["NOTES", "Free-text notes"],
        ["COMMENT", "Free-text comment"],
    ]
    _write_csv(out_dir / "data_dictionary" / "dict.csv", ["variable", "label"], dict_rows)
    # Cross-component duplicate: identical bytes to datasets/labs.csv, kept
    # as its own independent entry (never merged/deduplicated).
    _write_csv(out_dir / "data_dictionary" / "labs_dup.csv", labs_headers, labs_rows)

    # -- mappings/ ------------------------------------------------------------------------------
    _write_csv(
        out_dir / "mappings" / "site_map.csv",
        ["code", "label"],
        [[f"SITE-{i}", f"Study Site {i}"] for i in range(3)],
    )

    manifest = {
        "source_root": str(out_dir.resolve()),
        "seed": seed,
        "entries": _snapshot_tree(out_dir),
        "planted_unexpected_phi_rows": planted_unexpected_phi_rows,
        "planted_unexpected_phi_file": "datasets/site_notes.csv",
    }
    return manifest


# --- deliberately-invalid, review-required-only package ---------------------------------------


def build_review_required_fixtures(out_dir: Path, *, seed: int = 43) -> dict[str, Any]:
    """Build a source tree that reaches intake-manifest/v3
    ``status == "review_required"``: mostly-valid ``datasets/``, ``forms/``,
    and ``data_dictionary/`` content plus one file per fixed preflight
    review reason (unsupported format -- including the JSON/JSONL cases
    that v3 explicitly demotes from an accepted dataset format to an
    ``_unclassified`` review item -- invalid xlsx workbook, multi-sheet
    dataset xlsx, and a source symlink). Deliberately never organized/run --
    a v3 review item blocks the whole study, so this tree exists only to
    prove intake itself fails closed with the right reasons. Idempotent:
    wipes and rebuilds *out_dir* every call."""
    fk = _fake(seed)

    if out_dir.exists():
        shutil.rmtree(out_dir)
    out_dir.mkdir(parents=True)

    _write_csv(out_dir / "datasets" / "good.csv", ["SUBJID", "AGE"], [[f"G-{i:03d}", fk.random_int(18, 85)] for i in range(4)])
    _write_unknown_dat(out_dir / "datasets" / "mystery_export.dat")
    _write_malformed_xlsx(out_dir / "datasets" / "corrupted_workbook.xlsx")
    _write_multi_sheet_xlsx(out_dir / "datasets" / "multi_sheet.xlsx", fk)
    broken_link = out_dir / "datasets" / "broken_link.csv"
    os.symlink(out_dir / "datasets" / "does_not_exist.csv", broken_link)
    _write_json_list(out_dir / "datasets" / "extra_group.json", [{"SUBJID": "JS-001", "GROUP": "A"}])
    _write_jsonl(out_dir / "datasets" / "demographics.jsonl", [{"SUBJID": "JL-001", "AGE": 40}])

    _write_pdf_with_table(out_dir / "forms" / "consent.pdf", fk)
    _write_csv(out_dir / "data_dictionary" / "dict.csv", ["variable", "label"], [["SUBJID", "Subject identifier"], ["AGE", "Age in years"]])

    return {
        "source_root": str(out_dir.resolve()),
        "seed": seed,
        "expected_review_reasons": {
            "datasets/mystery_export.dat": "unsupported-format",
            "datasets/corrupted_workbook.xlsx": "xlsx-workbook-invalid",
            "datasets/multi_sheet.xlsx": "dataset-xlsx-multiple-sheets",
            "datasets/broken_link.csv": "source-symlink-not-allowed",
            "datasets/extra_group.json": "unsupported-format",
            "datasets/demographics.jsonl": "unsupported-format",
        },
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--out", required=True, help="Output source directory")
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument(
        "--manifest-out",
        default=None,
        help="Manifest path (default: <out>.manifest/stress_manifest.json, outside --out)",
    )
    args = parser.parse_args(argv)

    out_dir = Path(args.out).resolve()
    manifest = build_stress_fixtures(out_dir, seed=args.seed)

    manifest_path = (
        Path(args.manifest_out)
        if args.manifest_out
        else out_dir.parent / f"{out_dir.name}.manifest" / MANIFEST_FILENAME
    )
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(manifest, indent=2, sort_keys=True), encoding="utf-8")

    print(f"stress source tree: {out_dir} ({len(manifest['entries'])} entries)")
    print(f"manifest: {manifest_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
