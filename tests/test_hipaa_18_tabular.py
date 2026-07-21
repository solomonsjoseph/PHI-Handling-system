from __future__ import annotations

import csv
import hashlib

import json
import subprocess
import sys

import pytest
from openpyxl import load_workbook

from generators.hipaa_18_tabular import (
    HIPAA18_IDENTIFIER_SPECS,
    USHIPAA18TabularCorpusGenerator,
    generate_edge_cases,
    validate_corpus,
)
from harness.capability_registry import load_capabilities


def test_baseline_has_every_hipaa_category_once():
    corpus = USHIPAA18TabularCorpusGenerator(seed=42).generate(n_subjects=18)

    assert [row["hipaa_category"] for row in corpus.dictionary_rows] == list(
        "ABCDEFGHIJKLMNOPQR"
    )
    assert len(corpus.dictionary_rows) == 18


def test_every_subject_has_the_ordered_eighteen_identifier_columns():
    corpus = USHIPAA18TabularCorpusGenerator(seed=42).generate(n_subjects=18)
    expected_columns = [spec.source_column for spec in HIPAA18_IDENTIFIER_SPECS]

    assert len(corpus.dataset_rows) == 18
    assert all(list(row) == expected_columns for row in corpus.dataset_rows)
    assert all(all(row.values()) for row in corpus.dataset_rows)


def test_baseline_user_output_drops_direct_identifiers_and_retains_only_dob_year():
    corpus = USHIPAA18TabularCorpusGenerator(seed=42).generate(n_subjects=2)
    assert len(corpus.dataset_rows) == len(corpus.expected_user_rows)

    for source, output in zip(corpus.dataset_rows, corpus.expected_user_rows):
        assert output["ROW_TOKEN"].startswith("ROW_")
        for spec in HIPAA18_IDENTIFIER_SPECS:
            if spec.hipaa_category == "C":
                assert output[spec.source_column] == source[spec.source_column][:4]
            else:
                assert output[spec.source_column] == ""


def test_in_memory_generation_is_seed_deterministic():
    first = USHIPAA18TabularCorpusGenerator(seed=42).generate(n_subjects=18)
    second = USHIPAA18TabularCorpusGenerator(seed=42).generate(n_subjects=18)

    assert second == first


def test_generate_requires_at_least_one_subject():
    with pytest.raises(ValueError, match="n_subjects must be >= 1"):
        USHIPAA18TabularCorpusGenerator(seed=42).generate(n_subjects=0)


def test_audit_events_do_not_embed_plaintext_identifiers():
    corpus = USHIPAA18TabularCorpusGenerator(seed=42).generate(n_subjects=2)
    audit_text = json.dumps(corpus.audit_events, sort_keys=True)

    for row in corpus.dataset_rows:
        for value in row.values():
            assert value not in audit_text


def test_baseline_validates_and_cell_evidence_is_complete():
    corpus = USHIPAA18TabularCorpusGenerator(seed=42).generate(n_subjects=18)

    assert validate_corpus(corpus) == []
    assert len(corpus.gold_entries) == 18 * 18
    assert len(corpus.audit_events) == 18 * 18


def test_gold_entries_reference_exact_dataset_cells():
    corpus = USHIPAA18TabularCorpusGenerator(seed=42).generate(n_subjects=3)

    for entry in corpus.gold_entries:
        assert (
            corpus.dataset_rows[entry["row_index"]][entry["column"]]
            == entry["original_value"]
        )


def test_audit_events_have_what_why_how_and_hmac_evidence():
    corpus = USHIPAA18TabularCorpusGenerator(seed=42).generate(n_subjects=1)

    for event in corpus.audit_events:
        assert set(event["what"]) == {"action", "outcome"}
        assert set(event["why"]) == {"rule_id", "authority", "reason"}
        assert set(event["how"]) == {"method"}
        assert set(event["evidence"]) == {
            "input_hmac_sha256",
            "output_hmac_sha256",
        }
        assert "original_value" not in json.dumps(event, sort_keys=True)


def _read_csv(path):
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def _read_xlsx(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    values = list(worksheet.iter_rows(values_only=True))
    workbook.close()
    headers = [str(value) for value in values[0]]
    return [
        {
            header: "" if value is None else str(value)
            for header, value in zip(headers, row)
        }
        for row in values[1:]
    ]


def _tree_hashes(root):
    return {
        path.relative_to(root).as_posix(): hashlib.sha256(path.read_bytes()).hexdigest()
        for path in sorted(root.rglob("*"))
        if path.is_file()
    }


def test_write_emits_equivalent_csv_and_xlsx_with_valid_manifest(tmp_path):
    report = USHIPAA18TabularCorpusGenerator(seed=42).write(
        tmp_path, n_subjects=18, include_edge_cases=False
    )

    baseline = tmp_path / "baseline"
    assert report["validation_status"] == "PASS"
    assert _read_csv(baseline / "input/datasets/hipaa18.csv") == _read_xlsx(
        baseline / "input/datasets/hipaa18.xlsx"
    )
    assert _read_csv(
        baseline / "input/data_dictionary/hipaa18_dictionary.csv"
    ) == _read_xlsx(
        baseline / "input/data_dictionary/hipaa18_dictionary.xlsx"
    )

    manifest = json.loads((baseline / "MANIFEST.json").read_text(encoding="utf-8"))
    for relative_path, evidence in manifest["files"].items():
        payload = (baseline / relative_path).read_bytes()
        assert evidence == {
            "bytes": len(payload),
            "sha256": hashlib.sha256(payload).hexdigest(),
        }


def test_write_is_byte_deterministic(tmp_path):
    first = tmp_path / "first"
    second = tmp_path / "second"
    generator = USHIPAA18TabularCorpusGenerator(seed=42)

    generator.write(first, n_subjects=18, include_edge_cases=False)
    generator.write(second, n_subjects=18, include_edge_cases=False)

    assert _tree_hashes(first) == _tree_hashes(second)


@pytest.mark.parametrize(
    ("case_name", "expected_codes"),
    [
        (
            "missing_dictionary_entry",
            {"MISSING_HIPAA_CATEGORY", "UNMAPPED_DATASET_COLUMN"},
        ),
        ("orphan_dictionary_entry", {"ORPHAN_DICTIONARY_VARIABLE"}),
        (
            "duplicate_mapping",
            {"DUPLICATE_HIPAA_CATEGORY", "DUPLICATE_MAPPING"},
        ),
        (
            "conflicting_category",
            {
                "CONFLICTING_HIPAA_CATEGORY",
                "DUPLICATE_HIPAA_CATEGORY",
                "MISSING_HIPAA_CATEGORY",
            },
        ),
        ("explicit_alias_mapping", set()),
    ],
)
def test_edge_case_declares_observed_validation(case_name, expected_codes):
    case = generate_edge_cases(seed=42)[case_name]

    assert {issue.code for issue in validate_corpus(case)} == expected_codes


def test_write_emits_all_edge_cases_with_exact_expected_validation(tmp_path):
    report = USHIPAA18TabularCorpusGenerator(seed=42).write(
        tmp_path, n_subjects=3, include_edge_cases=True
    )

    assert report["edge_cases"] == 5
    for case_name, case in generate_edge_cases(seed=42).items():
        case_dir = tmp_path / "edge_cases" / case_name
        expected = json.loads(
            (case_dir / "expected_validation.json").read_text(encoding="utf-8")
        )
        assert expected["issues"] == [
            {
                "code": issue.code,
                "column": issue.column,
                "detail": issue.detail,
            }
            for issue in validate_corpus(case)
        ]
        assert (case_dir / "input/datasets/hipaa18.csv").is_file()
        assert (
            case_dir / "input/data_dictionary/hipaa18_dictionary.csv"
        ).is_file()


def test_cli_writes_passing_baseline_and_edges(tmp_path):
    result = subprocess.run(
        [
            sys.executable,
            "-m",
            "harness.generate_hipaa18_tabular",
            "--seed",
            "42",
            "--n-subjects",
            "18",
            "--out-dir",
            str(tmp_path),
        ],
        text=True,
        capture_output=True,
        check=False,
    )

    assert result.returncode == 0, result.stderr
    assert "HIPAA categories: 18/18" in result.stdout
    assert "Baseline validation: PASS" in result.stdout
    assert "Edge cases: 5" in result.stdout
    coverage = json.loads(
        (tmp_path / "baseline" / "coverage_report.json").read_text(
            encoding="utf-8"
        )
    )
    assert coverage["validation_status"] == "PASS"


def test_capability_registry_tracks_hipaa18_tabular_corpus():
    capability = next(
        item
        for item in load_capabilities()
        if item.id == "format_hipaa18_tabular"
    )

    assert capability.status == "tested"
    assert (
        capability.generator
        == "generators.hipaa_18_tabular:USHIPAA18TabularCorpusGenerator.write"
    )
    assert capability.tests == ("tests/test_hipaa_18_tabular.py",)
