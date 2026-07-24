from __future__ import annotations

import hashlib
import io
import inspect
import os
import shutil
import struct
import tempfile
import warnings
import zipfile
from pathlib import Path

import openpyxl
import pytest
import xlwt

import phi_engine.pipeline.intake_preflight as intake_preflight_module
from phi_engine.pipeline import xls_isolation
from phi_engine.pipeline.intake_preflight import (
    count_xlsx_sheets,
    inspect_intake_source,
)
from phi_engine.pipeline.support_files import DEFAULT_LIMITS


def _make_canonical(root: Path) -> None:
    (root / "datasets").mkdir(parents=True)
    (root / "forms").mkdir(parents=True)
    (root / "dictionary_mapping").mkdir(parents=True)
    (root / "datasets" / "labs.csv").write_text("SUBJID,AGE\n1,40\n", encoding="utf-8")
    (root / "forms" / "consent.pdf").write_bytes(b"%PDF-1.4\n%%EOF")
    _write_workbook(root / "dictionary_mapping" / "dict.xlsx", sheet_names=["Sheet1"])


def _write_workbook(path: Path, *, sheet_names: list[str]) -> None:
    wb = openpyxl.Workbook()
    wb.active.title = sheet_names[0]
    for name in sheet_names[1:]:
        wb.create_sheet(name)
    wb.save(path)


def _write_xls_workbook(path: Path, *, sheet_names: list[str]) -> None:
    """Genuine ``xlwt`` BIFF (.xls) bytes -- matches
    ``tests/test_xls_isolation.py``'s own fixture convention, exercising
    the real ``xlrd``-backed worker end to end rather than a stub."""
    wb = xlwt.Workbook()
    for name in sheet_names:
        ws = wb.add_sheet(name)
        ws.write(0, 0, "value")
    wb.save(str(path))


def _raw_workbook_zip_bytes(sheets_inner_xml: str, *, namespace: str | None = None) -> bytes:
    ns = namespace or "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<workbook xmlns="{ns}" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        f"<sheets>{sheets_inner_xml}</sheets></workbook>"
    )
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("xl/workbook.xml", workbook_xml)
    return buf.getvalue()


def _reasons_for_path(result, path: str) -> set[str]:
    return {item["reason"] for item in result.review_items if item["path"] == path} | {
        item["reason"] for item in result.errors if item["path"] == path
    }


def _by_path(candidates) -> dict[str, object]:
    return {c.relative_path: c for c in candidates}


def _reasons(items) -> set[str]:
    return {item["reason"] for item in items}


def _review_map(result) -> dict[str, str]:
    return {item["path"]: item["reason"] for item in result.review_items}


def _error_map(result) -> dict[str, str]:
    return {item["path"]: item["reason"] for item in result.errors}


# --- public contract -------------------------------------------------------------------


def test_public_names_match_the_approved_contract() -> None:
    assert set(intake_preflight_module.__all__) == {
        "Component",
        "IntakeCandidate",
        "IntakePreflight",
        "inspect_intake_source",
        "count_xlsx_sheets",
    }
    sig = inspect.signature(count_xlsx_sheets)
    assert list(sig.parameters) == ["fileobj"]
    sig2 = inspect.signature(inspect_intake_source)
    assert list(sig2.parameters) == ["source"]


# --- candidates for a canonical package -------------------------------------------------


def test_canonical_package_produces_ready_candidates_no_review_no_errors(tmp_path: Path) -> None:
    _make_canonical(tmp_path)

    result = inspect_intake_source(tmp_path)

    assert result.review_items == ()
    assert result.errors == ()
    by_path = _by_path(result.candidates)
    assert by_path["datasets/labs.csv"].component == "datasets"
    assert by_path["forms/consent.pdf"].component == "forms"
    assert by_path["dictionary_mapping/dict.xlsx"].component == "dictionary_mapping"
    assert by_path["dictionary_mapping/dict.xlsx"].sheet_count == 1
    assert by_path["datasets/labs.csv"].sheet_count is None
    for candidate in result.candidates:
        assert candidate.sha256 and len(candidate.sha256) == 64
        assert candidate.identity.size >= 0


def test_forms_alone_satisfies_the_support_requirement(tmp_path: Path) -> None:
    (tmp_path / "datasets").mkdir(parents=True)
    (tmp_path / "forms").mkdir(parents=True)
    (tmp_path / "datasets" / "labs.csv").write_text("SUBJID,AGE\n1,40\n", encoding="utf-8")
    (tmp_path / "forms" / "consent.pdf").write_bytes(b"%PDF-1.4\n%%EOF")

    result = inspect_intake_source(tmp_path)

    assert result.review_items == ()
    assert result.errors == ()
    assert _by_path(result.candidates)["forms/consent.pdf"].component == "forms"


def test_nested_directories_and_duplicate_bytes_remain_distinct_candidates(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    nested = tmp_path / "datasets" / "site_a" / "visit_1"
    nested.mkdir(parents=True)
    (nested / "dup.csv").write_text("SUBJID,AGE\n1,40\n", encoding="utf-8")  # same bytes as labs.csv

    result = inspect_intake_source(tmp_path)

    by_path = _by_path(result.candidates)
    assert "datasets/site_a/visit_1/dup.csv" in by_path
    assert by_path["datasets/site_a/visit_1/dup.csv"].sha256 == by_path["datasets/labs.csv"].sha256
    assert by_path["datasets/site_a/visit_1/dup.csv"].relative_path != by_path["datasets/labs.csv"].relative_path


def test_candidates_and_review_items_are_deterministically_ordered(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    (tmp_path / "datasets" / "z.txt").write_text("x", encoding="utf-8")
    (tmp_path / "datasets" / "a.txt").write_text("x", encoding="utf-8")

    first = inspect_intake_source(tmp_path)
    second = inspect_intake_source(tmp_path)

    assert [c.relative_path for c in first.candidates] == sorted(c.relative_path for c in first.candidates)
    assert first.candidates == second.candidates
    assert first.review_items == second.review_items
    assert first.errors == second.errors


# --- source root itself ------------------------------------------------------------------


def test_symlinked_source_root_is_rejected_root_never_followed(tmp_path: Path) -> None:
    real = tmp_path / "real_package"
    _make_canonical(real)
    link = tmp_path / "source_link"
    link.symlink_to(real)

    result = inspect_intake_source(link)

    assert result.candidates == ()
    assert result.review_items == ()
    assert result.errors == ({"path": "", "reason": "source-symlink-not-allowed"},)


def test_symlink_loop_source_root_fails_closed_without_raw_exception(tmp_path: Path) -> None:
    loop = tmp_path / "loop"
    loop.symlink_to(loop)

    result = inspect_intake_source(loop)

    assert result.errors == ({"path": "", "reason": "source-symlink-not-allowed"},)


def test_unreadable_source_root_yields_single_error_no_candidates(tmp_path: Path) -> None:
    missing = tmp_path / "does_not_exist"

    result = inspect_intake_source(missing)

    assert result.candidates == ()
    assert result.review_items == ()
    assert result.errors == ({"path": "", "reason": "source-unreadable"},)


def test_source_root_that_is_a_regular_file_is_rejected(tmp_path: Path) -> None:
    not_a_dir = tmp_path / "file.txt"
    not_a_dir.write_text("x", encoding="utf-8")

    result = inspect_intake_source(not_a_dir)

    assert result.errors == ({"path": "", "reason": "source-unreadable"},)


def test_symlinked_ancestor_of_source_root_is_rejected_not_only_the_final_component(tmp_path: Path) -> None:
    # The root pin walks EVERY ancestor path segment, not only the final
    # "source" component -- a symlink anywhere in the supplied ancestry must
    # be rejected, even though the final segment itself is a real directory.
    real = tmp_path / "real_package"
    _make_canonical(real)
    linked_parent = tmp_path / "linked_parent"
    linked_parent.symlink_to(tmp_path)
    source_through_symlinked_ancestor = linked_parent / "real_package"

    result = inspect_intake_source(source_through_symlinked_ancestor)

    assert result.candidates == ()
    assert result.errors == ({"path": "", "reason": "source-symlink-not-allowed"},)


def test_root_path_renamed_mid_traversal_never_produces_a_mixed_package_snapshot(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    # Synchronized swap: after the first candidate has been fully processed,
    # atomically rename the exact filesystem-namespace path preflight was
    # given away, then rename a replacement package into that same path --
    # simulating a concurrent directory replacement of the whole root.
    # Because the root descriptor is pinned by inode (not by path) and every
    # subsequent listing/open is descriptor-relative to it, this rename can
    # never be observed by the in-flight traversal at all: the result must
    # reflect only the originally-pinned tree, never the replacement and
    # never a mixture of the two.
    original = tmp_path / "original"
    replacement = tmp_path / "replacement"
    _make_canonical(original)
    _make_canonical(replacement)
    (replacement / "datasets" / "labs.csv").write_text("SUBJID,AGE\n9,99\n", encoding="utf-8")  # distinct content
    source = tmp_path / "source"
    shutil.copytree(original, source)

    real_hash = intake_preflight_module._hash_from_fd
    call_count = {"n": 0}

    def swapping_hash(fd: int) -> str:
        call_count["n"] += 1
        result = real_hash(fd)
        if call_count["n"] == 1:
            # Atomic (each step is its own atomic rename) whole-root swap,
            # performed only after the first candidate's own hash finished.
            aside = tmp_path / "source_old"
            os.rename(source, aside)
            os.rename(replacement, source)
        return result

    monkeypatch.setattr(intake_preflight_module, "_hash_from_fd", swapping_hash)

    result = inspect_intake_source(source)

    assert call_count["n"] >= 1, "the swap hook never fired -- test setup is broken"
    original_sha = hashlib.sha256(b"SUBJID,AGE\n1,40\n").hexdigest()
    replacement_sha = hashlib.sha256(b"SUBJID,AGE\n9,99\n").hexdigest()
    by_path = _by_path(result.candidates)
    assert "datasets/labs.csv" in by_path
    assert by_path["datasets/labs.csv"].sha256 == original_sha
    assert by_path["datasets/labs.csv"].sha256 != replacement_sha
    assert result.review_items == ()
    assert result.errors == ()


def test_final_file_replaced_between_discovery_and_open_is_rejected(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    # Synchronized swap at the narrowest possible window: mutate the file's
    # content in place (same directory entry, so the parent directory's own
    # mtime is untouched) exactly between the moment its identity is
    # captured from discovery and the moment it is actually opened for
    # verified reading.
    _make_canonical(tmp_path)
    target = tmp_path / "datasets" / "labs.csv"

    real_dirent_identity = intake_preflight_module._dirent_identity

    def swapping_dirent_identity(entry):
        result = real_dirent_identity(entry)
        if entry.name == "labs.csv":
            with open(target, "w", encoding="utf-8") as fh:
                fh.write("SUBJID,AGE\n9,99\nEXTRA,ROW\n")
        return result

    monkeypatch.setattr(intake_preflight_module, "_dirent_identity", swapping_dirent_identity)

    result = inspect_intake_source(tmp_path)

    assert "datasets/labs.csv" not in _by_path(result.candidates)
    assert _error_map(result).get("datasets/labs.csv") == "source-unreadable"


def test_nested_directory_replaced_during_child_processing_never_mixes_trees(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    # Synchronized swap: replace a NESTED directory (not the root) with a
    # different physical tree while its own children are still being
    # processed. Data integrity comes from descriptor pinning: the "nested"
    # frame's own held fd is inode-pinned and unaffected by the external
    # rename, so its still-unread children are guaranteed to come from the
    # ORIGINAL tree regardless of the swap. The post-children identity
    # recheck (on "nested" itself, or its parent "datasets" whose entries
    # the rename also touched) is an ADDITIONAL best-effort layer that can
    # occasionally miss a rename happening within the same filesystem mtime
    # tick -- so the guarantee actually under test is "never a mixture",
    # not "detection always fires": either the whole result is invalidated
    # (detection fired), or every candidate is provably from the original
    # tree only (detection missed but pinning still held).
    _make_canonical(tmp_path)
    nested = tmp_path / "datasets" / "nested"
    nested.mkdir()
    (nested / "a.csv").write_text("a,b\n1,2\n", encoding="utf-8")
    (nested / "b.csv").write_text("a,b\n3,4\n", encoding="utf-8")

    replacement_base = Path(tempfile.mkdtemp())
    replacement_nested = replacement_base / "nested_replacement_source"
    replacement_nested.mkdir()
    (replacement_nested / "a.csv").write_text("a,b\n1,2\n", encoding="utf-8")
    (replacement_nested / "b.csv").write_text("a,b\nZZZ,ZZZ\n", encoding="utf-8")

    real_hash = intake_preflight_module._hash_from_fd
    a_csv_size = (nested / "a.csv").stat().st_size
    swapped = {"done": False}

    def swapping_hash(fd: int) -> str:
        result = real_hash(fd)
        if not swapped["done"] and os.fstat(fd).st_size == a_csv_size:
            swapped["done"] = True
            aside = replacement_base / "nested_old"
            os.rename(nested, aside)
            os.rename(replacement_nested, nested)
        return result

    monkeypatch.setattr(intake_preflight_module, "_hash_from_fd", swapping_hash)

    result = inspect_intake_source(tmp_path)

    assert swapped["done"], "the swap hook never fired -- test setup is broken"
    if result.candidates:
        # Best-effort mtime-based detection missed this particular rename
        # (clock-tick granularity); the data itself must still be provably
        # from the original tree only -- never the replacement, never mixed.
        by_path = _by_path(result.candidates)
        assert by_path["datasets/nested/a.csv"].sha256 == hashlib.sha256(b"a,b\n1,2\n").hexdigest()
        assert by_path["datasets/nested/b.csv"].sha256 == hashlib.sha256(b"a,b\n3,4\n").hexdigest()
        replacement_b_sha = hashlib.sha256(b"a,b\nZZZ,ZZZ\n").hexdigest()
        assert by_path["datasets/nested/b.csv"].sha256 != replacement_b_sha
    else:
        assert result.errors == ({"path": "", "reason": "source-unreadable"},)
        assert result.review_items == ()


# --- blocking structural layouts ---------------------------------------------------------


def test_flat_source_layout_uses_fixed_root_path_never_none(tmp_path: Path) -> None:
    (tmp_path / "labs.csv").write_text("SUBJID,AGE\n1,40\n", encoding="utf-8")

    result = inspect_intake_source(tmp_path)

    assert _review_map(result) == {"": "flat-source-layout"}
    # Retained, but never AI-eligible.
    assert _by_path(result.candidates)["labs.csv"].component == "_unclassified"


def test_stray_root_file_alongside_known_directories_is_flat_source_layout_not_unknown_directory(
    tmp_path: Path,
) -> None:
    _make_canonical(tmp_path)
    (tmp_path / "mystery.csv").write_text("a,b\n1,2\n", encoding="utf-8")

    result = inspect_intake_source(tmp_path)

    review = _review_map(result)
    assert review[""] == "flat-source-layout"
    assert "mystery.csv" not in review
    assert "unknown-top-level-directory" not in review.values()
    assert _by_path(result.candidates)["mystery.csv"].component == "_unclassified"


def test_unknown_only_root_reports_both_flat_layout_and_unknown_directory(tmp_path: Path) -> None:
    (tmp_path / "extra").mkdir()
    (tmp_path / "extra" / "note.csv").write_text("a,b\n1,2\n", encoding="utf-8")

    result = inspect_intake_source(tmp_path)

    review = _review_map(result)
    assert review[""] == "flat-source-layout"
    assert review["extra"] == "unknown-top-level-directory"
    assert _by_path(result.candidates)["extra/note.csv"].component == "_unclassified"


def test_missing_required_directory_is_blocking_review(tmp_path: Path) -> None:
    (tmp_path / "forms").mkdir()
    (tmp_path / "forms" / "consent.pdf").write_bytes(b"%PDF-1.4\n%%EOF")

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["datasets"] == "missing-component-directory"


def test_component_directory_present_but_empty_of_accepted_files_is_blocking_review(tmp_path: Path) -> None:
    (tmp_path / "datasets").mkdir()
    (tmp_path / "datasets" / "readme.txt").write_text("no dataset here", encoding="utf-8")
    (tmp_path / "forms").mkdir()
    (tmp_path / "forms" / "consent.pdf").write_bytes(b"%PDF-1.4\n%%EOF")

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["datasets"] == "missing-component-content"
    assert _by_path(result.candidates)["datasets/readme.txt"].component == "_unclassified"


def test_malformed_xlsx_only_dataset_content_still_reports_missing_component_content(tmp_path: Path) -> None:
    # A directory that superficially "has a file" but whose only file fails
    # workbook validation must not count as satisfying the requirement.
    (tmp_path / "datasets").mkdir()
    (tmp_path / "forms").mkdir()
    (tmp_path / "dictionary_mapping").mkdir()
    (tmp_path / "datasets" / "bad.xlsx").write_text("not a zip", encoding="utf-8")
    (tmp_path / "forms" / "consent.pdf").write_bytes(b"%PDF-1.4")
    (tmp_path / "dictionary_mapping" / "map.csv").write_text("a,b\n1,2\n", encoding="utf-8")

    result = inspect_intake_source(tmp_path)

    review = _review_map(result)
    assert review["datasets"] == "missing-component-content"
    assert review["datasets/bad.xlsx"] == "xlsx-workbook-invalid"


def test_symlink_only_dataset_content_still_reports_missing_component_content(tmp_path: Path) -> None:
    (tmp_path / "datasets").mkdir()
    (tmp_path / "forms").mkdir()
    (tmp_path / "dictionary_mapping").mkdir()
    other = tmp_path / "elsewhere.csv"
    other.write_text("a,b\n1,2\n", encoding="utf-8")
    (tmp_path / "datasets" / "linked.csv").symlink_to(other)
    (tmp_path / "forms" / "consent.pdf").write_bytes(b"%PDF-1.4")
    (tmp_path / "dictionary_mapping" / "map.csv").write_text("a,b\n1,2\n", encoding="utf-8")

    result = inspect_intake_source(tmp_path)

    review = _review_map(result)
    assert review["datasets"] == "missing-component-content"
    assert review["datasets/linked.csv"] == "source-symlink-not-allowed"


def test_hardlink_only_support_content_still_blocks_the_whole_study(tmp_path: Path) -> None:
    # No forms/ at all, and the only dictionary_mapping content is a
    # cross-component hardlink alias -- neither support component has a
    # final accepted candidate, so the single alternative-group review item
    # fires even though the directory itself "has a file" on disk.
    (tmp_path / "datasets").mkdir()
    (tmp_path / "dictionary_mapping").mkdir()
    (tmp_path / "datasets" / "labs.csv").write_text("a,b\n1,2\n", encoding="utf-8")
    os.link(tmp_path / "datasets" / "labs.csv", tmp_path / "dictionary_mapping" / "aliased.csv")

    result = inspect_intake_source(tmp_path)

    review = _review_map(result)
    assert review[""] == "missing-support-component"
    assert "cross-component-hardlink" in _reasons_for_path(result, "dictionary_mapping/aliased.csv")


def test_neither_dictionary_mapping_nor_forms_present_blocks_the_whole_study(tmp_path: Path) -> None:
    (tmp_path / "datasets").mkdir()
    (tmp_path / "datasets" / "labs.csv").write_text("SUBJID,AGE\n1,40\n", encoding="utf-8")

    result = inspect_intake_source(tmp_path)

    # Exactly one root-level review item -- the alternative group is a
    # single requirement, never two independent per-component findings.
    assert result.review_items == ({"path": "", "reason": "missing-support-component", "blocking": True},)


def test_dictionary_mapping_present_with_content_means_forms_absence_is_not_flagged(tmp_path: Path) -> None:
    _make_canonical(tmp_path)

    result = inspect_intake_source(tmp_path)

    assert "missing-support-component" not in _reasons(result.review_items)


def test_unknown_top_level_directory_is_blocking_review_and_files_retained_unclassified(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    (tmp_path / "extra_stuff").mkdir()
    (tmp_path / "extra_stuff" / "notes.csv").write_text("a,b\n1,2\n", encoding="utf-8")

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["extra_stuff"] == "unknown-top-level-directory"
    candidate = _by_path(result.candidates)["extra_stuff/notes.csv"]
    assert candidate.component == "_unclassified"
    assert candidate.source_component == "extra_stuff"


# --- closed format matrix -----------------------------------------------------------------


@pytest.mark.parametrize(
    "component,filename,content",
    [
        ("datasets", "notes.txt", b"plain text"),
        ("datasets", "consent.pdf", b"%PDF-1.4"),
        ("forms", "consent.docx", b"not a pdf"),
        ("forms", "data.csv", b"a,b\n1,2\n"),
        ("dictionary_mapping", "dict.txt", b"plain text"),
        ("dictionary_mapping", "map.json", b"{}"),
    ],
)
def test_unsupported_format_per_component_is_blocking_review(
    tmp_path: Path, component: str, filename: str, content: bytes
) -> None:
    _make_canonical(tmp_path)
    (tmp_path / component).mkdir(exist_ok=True)
    (tmp_path / component / filename).write_bytes(content)

    result = inspect_intake_source(tmp_path)

    rel = f"{component}/{filename}"
    assert _review_map(result)[rel] == "unsupported-format"
    assert _by_path(result.candidates)[rel].component == "_unclassified"


# --- .xls inspection (Approach 1's real inspect_xls wired in) ------------------------------


def test_xls_dataset_with_exactly_one_sheet_is_accepted(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    _write_xls_workbook(tmp_path / "datasets" / "legacy.xls", sheet_names=["Sheet1"])

    result = inspect_intake_source(tmp_path)

    candidate = _by_path(result.candidates)["datasets/legacy.xls"]
    assert candidate.component == "datasets"
    assert candidate.sheet_count == 1
    assert result.review_items == ()


def test_xls_dataset_with_multiple_sheets_is_blocking_review(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    _write_xls_workbook(tmp_path / "datasets" / "legacy.xls", sheet_names=["S1", "S2"])

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["datasets/legacy.xls"] == "dataset-xls-multiple-sheets"
    candidate = _by_path(result.candidates)["datasets/legacy.xls"]
    assert candidate.component == "_unclassified"
    assert candidate.sheet_count == 2


def test_xls_support_within_max_sheets_is_accepted(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    _write_xls_workbook(tmp_path / "dictionary_mapping" / "legacy.xls", sheet_names=[f"S{i}" for i in range(5)])

    result = inspect_intake_source(tmp_path)

    candidate = _by_path(result.candidates)["dictionary_mapping/legacy.xls"]
    assert candidate.component == "dictionary_mapping"
    assert candidate.sheet_count == 5
    assert result.review_items == ()


def test_xls_support_exceeding_max_sheets_is_blocking_review(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    max_sheets = DEFAULT_LIMITS["max_sheets"]
    names = [f"S{i}" for i in range(max_sheets + 1)]
    _write_xls_workbook(tmp_path / "dictionary_mapping" / "legacy.xls", sheet_names=names)

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["dictionary_mapping/legacy.xls"] == "support-xls-sheet-limit"
    assert _by_path(result.candidates)["dictionary_mapping/legacy.xls"].component == "_unclassified"


def test_malformed_xls_is_workbook_invalid_and_unclassified(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    (tmp_path / "datasets" / "corrupt.xls").write_bytes(b"not actually a BIFF file at all")

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["datasets/corrupt.xls"] == "xls-workbook-invalid"
    assert _by_path(result.candidates)["datasets/corrupt.xls"].component == "_unclassified"


def test_xls_reader_unavailable_maps_to_fixed_reason(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    _make_canonical(tmp_path)
    _write_xls_workbook(tmp_path / "datasets" / "legacy.xls", sheet_names=["Sheet1"])

    def fake_inspect_xls(data, expected_sha256, *, max_sheets, deadline=None):
        raise xls_isolation.XlsIsolationError("isolation-unavailable")

    monkeypatch.setattr(xls_isolation, "inspect_xls", fake_inspect_xls)

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["datasets/legacy.xls"] == "xls-reader-unavailable"


def test_xls_worker_parse_failure_maps_to_workbook_invalid(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    _make_canonical(tmp_path)
    _write_xls_workbook(tmp_path / "datasets" / "legacy.xls", sheet_names=["Sheet1"])

    def fake_inspect_xls(data, expected_sha256, *, max_sheets, deadline=None):
        raise xls_isolation.XlsWorkerError("parse-error")

    monkeypatch.setattr(xls_isolation, "inspect_xls", fake_inspect_xls)

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["datasets/legacy.xls"] == "xls-workbook-invalid"


def test_xls_generic_isolation_resource_limit_maps_to_workbook_invalid(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    # An ORDINARY (per-file, not package-level) isolation resource-limit
    # falls into the generic bucket -- never the package-abort path.
    _make_canonical(tmp_path)
    _write_xls_workbook(tmp_path / "datasets" / "legacy.xls", sheet_names=["Sheet1"])

    def fake_inspect_xls(data, expected_sha256, *, max_sheets, deadline=None):
        raise xls_isolation.XlsIsolationError("resource-limit")

    monkeypatch.setattr(xls_isolation, "inspect_xls", fake_inspect_xls)

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["datasets/legacy.xls"] == "xls-workbook-invalid"


def test_xls_package_deadline_precedence_aborts_the_whole_package(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    # The external package-deadline code takes precedence over every other
    # per-file finding and aborts the WHOLE package -- discarding every
    # candidate (including the otherwise-valid forms/dictionary_mapping
    # content), never a per-file review item.
    _make_canonical(tmp_path)
    _write_xls_workbook(tmp_path / "datasets" / "legacy.xls", sheet_names=["Sheet1"])

    def fake_inspect_xls(data, expected_sha256, *, max_sheets, deadline=None):
        raise xls_isolation.XlsIsolationError("package-resource-limit")

    monkeypatch.setattr(xls_isolation, "inspect_xls", fake_inspect_xls)

    result = inspect_intake_source(tmp_path)

    assert result.candidates == ()
    assert result.errors == ()
    assert result.review_items == ({"path": "", "reason": "intake-resource-limit", "blocking": True},)


def test_xls_package_deadline_precedence_survives_a_concurrent_identity_change(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    # Regression: the package-abort signal (_PackageLimitError) is raised
    # from INSIDE the `with _open_from_parent_fd(...) as fd:` block. That
    # context manager's own finally block re-checks this file's identity
    # after the block body finishes and raises VerifiedSourceError
    # ("source-unreadable") if it changed -- which, in CPython, SILENTLY
    # REPLACES an exception already propagating out of the `with` body.
    # A genuine concurrent write to this same file racing with the
    # package-deadline abort must never let that per-file identity-mismatch
    # finding downgrade the whole-package abort into a single per-file
    # error that lets the walk continue.
    _make_canonical(tmp_path)
    xls_path = tmp_path / "datasets" / "legacy.xls"
    _write_xls_workbook(xls_path, sheet_names=["Sheet1"])

    def fake_inspect_xls(data, expected_sha256, *, max_sheets, deadline=None):
        # Simulate the concurrent write racing with this file's own
        # inspection: its identity (size/mtime) changes before this
        # function's `with` block exits and its finally re-checks.
        os.utime(xls_path, (2000000000, 2000000000))
        raise xls_isolation.XlsIsolationError("package-resource-limit")

    monkeypatch.setattr(xls_isolation, "inspect_xls", fake_inspect_xls)

    result = inspect_intake_source(tmp_path)

    assert result.candidates == ()
    assert result.errors == ()
    assert result.review_items == ({"path": "", "reason": "intake-resource-limit", "blocking": True},)


def test_xls_inspection_receives_the_package_deadline(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    _make_canonical(tmp_path)
    _write_xls_workbook(tmp_path / "datasets" / "legacy.xls", sheet_names=["Sheet1"])

    seen: dict[str, float] = {}
    real_inspect = xls_isolation.inspect_xls

    def spying_inspect_xls(data, expected_sha256, *, max_sheets, deadline=None):
        seen["deadline"] = deadline
        return real_inspect(data, expected_sha256, max_sheets=max_sheets, deadline=deadline)

    monkeypatch.setattr(xls_isolation, "inspect_xls", spying_inspect_xls)

    inspect_intake_source(tmp_path)

    assert isinstance(seen.get("deadline"), float)


def test_xls_over_max_source_bytes_is_workbook_invalid_without_calling_inspect_xls(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    # Mirrors test_count_xlsx_sheets_rejects_source_over_max_source_bytes's
    # monkeypatch technique for the .xls (xls_isolation.INSPECT_LIMITS)
    # analog of the same per-file size cap: a file between this cap and
    # the 4-GiB package-wide byte ceiling must fail closed as
    # xls-workbook-invalid, and the size check must short-circuit BEFORE
    # ever calling xls_isolation.inspect_xls (proven via a failing spy,
    # not merely implied).
    _make_canonical(tmp_path)
    _write_xls_workbook(tmp_path / "datasets" / "legacy.xls", sheet_names=["Sheet1"])
    monkeypatch.setitem(xls_isolation.INSPECT_LIMITS, "max_source_bytes", 4)

    def failing_inspect_xls(*_args: object, **_kwargs: object) -> int:
        raise AssertionError("inspect_xls must never be called once the per-file size cap is exceeded")

    monkeypatch.setattr(xls_isolation, "inspect_xls", failing_inspect_xls)

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["datasets/legacy.xls"] == "xls-workbook-invalid"
    assert _by_path(result.candidates)["datasets/legacy.xls"].component == "_unclassified"


# --- xlsx sheet-count policy ---------------------------------------------------------------


def test_dataset_xlsx_with_exactly_one_sheet_is_accepted(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    _write_workbook(tmp_path / "datasets" / "single.xlsx", sheet_names=["Only"])

    result = inspect_intake_source(tmp_path)

    candidate = _by_path(result.candidates)["datasets/single.xlsx"]
    assert candidate.component == "datasets"
    assert candidate.sheet_count == 1
    assert result.review_items == ()


def test_dataset_xlsx_with_multiple_sheets_becomes_unclassified_review(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    _write_workbook(tmp_path / "datasets" / "multi.xlsx", sheet_names=["S1", "S2", "S3"])

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["datasets/multi.xlsx"] == "dataset-xlsx-multiple-sheets"
    candidate = _by_path(result.candidates)["datasets/multi.xlsx"]
    assert candidate.component == "_unclassified"
    assert candidate.sheet_count == 3


def test_support_xlsx_within_max_sheets_is_accepted(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    _write_workbook(tmp_path / "dictionary_mapping" / "map.xlsx", sheet_names=[f"S{i}" for i in range(5)])

    result = inspect_intake_source(tmp_path)

    candidate = _by_path(result.candidates)["dictionary_mapping/map.xlsx"]
    assert candidate.component == "dictionary_mapping"
    assert candidate.sheet_count == 5


def test_support_xlsx_exceeding_max_sheets_fails_closed(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    max_sheets = DEFAULT_LIMITS["max_sheets"]
    names = [f"S{i}" for i in range(max_sheets + 2)]
    _write_workbook(tmp_path / "dictionary_mapping" / "huge.xlsx", sheet_names=names)

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["dictionary_mapping/huge.xlsx"] == "support-xlsx-sheet-limit"
    assert _by_path(result.candidates)["dictionary_mapping/huge.xlsx"].component == "_unclassified"


def test_malformed_xlsx_is_workbook_invalid_and_unclassified(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    (tmp_path / "datasets" / "corrupt.xlsx").write_bytes(b"not actually a zip file at all")

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["datasets/corrupt.xlsx"] == "xlsx-workbook-invalid"
    assert _by_path(result.candidates)["datasets/corrupt.xlsx"].component == "_unclassified"


def test_xlsx_missing_workbook_member_is_workbook_invalid(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    bogus = tmp_path / "datasets" / "no_workbook.xlsx"
    with zipfile.ZipFile(bogus, "w") as zf:
        zf.writestr("hello.txt", "not a workbook")

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["datasets/no_workbook.xlsx"] == "xlsx-workbook-invalid"


def test_xlsx_duplicate_workbook_member_is_workbook_invalid(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    bogus = tmp_path / "datasets" / "dup_member.xlsx"
    buf = io.BytesIO()
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", UserWarning)
        with zipfile.ZipFile(buf, "w") as zf:
            zf.writestr("xl/workbook.xml", "content1")
            zf.writestr("xl/workbook.xml", "content2")
    bogus.write_bytes(buf.getvalue())

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["datasets/dup_member.xlsx"] == "xlsx-workbook-invalid"


def test_xlsx_wrong_namespace_sheet_tag_never_counted_as_a_real_sheet(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    bogus_bytes = _raw_workbook_zip_bytes('<sheet name="S1" sheetId="1"/>', namespace="urn:evil")
    (tmp_path / "datasets" / "wrongns.xlsx").write_bytes(bogus_bytes)

    result = inspect_intake_source(tmp_path)

    # Zero real (correctly-namespaced) sheets found -> invalid, not silently
    # accepted as a 1-sheet dataset via a foreign-namespace tag match.
    assert _review_map(result)["datasets/wrongns.xlsx"] == "xlsx-workbook-invalid"
    assert _by_path(result.candidates)["datasets/wrongns.xlsx"].component == "_unclassified"


def test_xlsx_encrypted_member_normalizes_to_workbook_invalid_no_raw_text(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("xl/workbook.xml", "<workbook/>")
    data = bytearray(buf.getvalue())
    local_sig = struct.pack("<I", 0x04034B50)
    central_sig = struct.pack("<I", 0x02014B50)
    idx = data.find(local_sig)
    flags = struct.unpack_from("<H", data, idx + 6)[0]
    struct.pack_into("<H", data, idx + 6, flags | 0x1)
    idx2 = data.find(central_sig)
    flags2 = struct.unpack_from("<H", data, idx2 + 8)[0]
    struct.pack_into("<H", data, idx2 + 8, flags2 | 0x1)
    (tmp_path / "datasets" / "encrypted.xlsx").write_bytes(bytes(data))

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["datasets/encrypted.xlsx"] == "xlsx-workbook-invalid"
    serialized = repr(result.review_items)
    assert "password" not in serialized.lower()
    assert "RuntimeError" not in serialized


def test_xlsx_zip_with_excessive_member_count_fails_closed(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr(
            "xl/workbook.xml",
            '<?xml version="1.0"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            '<sheets><sheet name="S1" sheetId="1"/></sheets></workbook>',
        )
        for i in range(DEFAULT_LIMITS["max_zip_members"] + 5):
            zf.writestr(f"junk/{i}.txt", "")
    (tmp_path / "datasets" / "manymembers.xlsx").write_bytes(buf.getvalue())

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["datasets/manymembers.xlsx"] == "xlsx-workbook-invalid"


def test_xlsx_non_workbook_member_over_per_member_cap_fails_closed(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr(
            "xl/workbook.xml",
            '<?xml version="1.0"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            '<sheets><sheet name="S1" sheetId="1"/></sheets></workbook>',
        )
        zf.writestr("xl/huge_other_member.xml", "x" * (DEFAULT_LIMITS["max_zip_member_bytes"] + 1024))
    (tmp_path / "datasets" / "hugemember.xlsx").write_bytes(buf.getvalue())

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["datasets/hugemember.xlsx"] == "xlsx-workbook-invalid"


# --- source symlink / hardlink / identity attacks ------------------------------------------


def test_source_file_symlink_in_component_dir_is_blocking_review_never_a_candidate_content(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    link = tmp_path / "datasets" / "linked.csv"
    link.symlink_to(tmp_path / "datasets" / "labs.csv")

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["datasets/linked.csv"] == "source-symlink-not-allowed"
    assert "datasets/linked.csv" not in _by_path(result.candidates)


def test_nested_file_symlink_deep_inside_component_dir_is_blocking_review(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    nested = tmp_path / "datasets" / "site_a" / "visit_1"
    nested.mkdir(parents=True)
    (nested / "linked.csv").symlink_to(tmp_path / "datasets" / "labs.csv")

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["datasets/site_a/visit_1/linked.csv"] == "source-symlink-not-allowed"
    assert "datasets/site_a/visit_1/linked.csv" not in _by_path(result.candidates)


def test_nested_directory_symlink_inside_component_dir_is_blocking_review_and_not_walked(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    real_nested = tmp_path / "backing_nested"
    real_nested.mkdir()
    (real_nested / "hidden.csv").write_text("a,b\n1,2\n", encoding="utf-8")
    (tmp_path / "datasets" / "linked_dir").symlink_to(real_nested)

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["datasets/linked_dir"] == "source-symlink-not-allowed"
    assert not any(c.relative_path.startswith("datasets/linked_dir/") for c in result.candidates)


def test_hidden_named_directory_symlink_is_still_reported_not_silently_dropped(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    real_nested = tmp_path / "backing_hidden"
    real_nested.mkdir()
    (real_nested / "x.csv").write_text("a,b\n1,2\n", encoding="utf-8")
    (tmp_path / "datasets" / ".hidden_link").symlink_to(real_nested)

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["datasets/.hidden_link"] == "source-symlink-not-allowed"


def test_hidden_named_file_symlink_is_still_reported_not_silently_dropped(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    (tmp_path / "datasets" / ".hidden_link.csv").symlink_to(tmp_path / "datasets" / "labs.csv")

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["datasets/.hidden_link.csv"] == "source-symlink-not-allowed"


def test_source_directory_symlink_at_top_level_is_blocking_review_and_not_walked(tmp_path: Path) -> None:
    real = tmp_path / "backing_store"
    real.mkdir()
    (real / "labs.csv").write_text("SUBJID,AGE\n1,40\n", encoding="utf-8")
    (tmp_path / "forms").mkdir()
    (tmp_path / "forms" / "consent.pdf").write_bytes(b"%PDF-1.4\n%%EOF")
    (tmp_path / "dictionary_mapping").mkdir()
    (tmp_path / "dictionary_mapping" / "map.csv").write_text("a,b\n1,2\n", encoding="utf-8")
    (tmp_path / "datasets").symlink_to(real)

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["datasets"] == "source-symlink-not-allowed"
    assert not any(c.relative_path.startswith("datasets/") for c in result.candidates)
    assert not any(c.source_component == "datasets" for c in result.candidates)
    by_path = _by_path(result.candidates)
    assert by_path["backing_store/labs.csv"].component == "_unclassified"


def test_hardlink_from_forms_to_dataset_file_is_quarantined(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    os.link(tmp_path / "datasets" / "labs.csv", tmp_path / "forms" / "aliased.pdf")

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["forms/aliased.pdf"] == "cross-component-hardlink"
    candidate = _by_path(result.candidates)["forms/aliased.pdf"]
    assert candidate.component == "_unclassified"
    assert _by_path(result.candidates)["datasets/labs.csv"].component == "datasets"
    # A hardlink trivially also matches by SHA-256 (same bytes, same
    # identity) -- the hardlink finding must be the ONLY reason reported
    # for this path, never also/instead "cross-component-dataset-copy".
    assert _reasons_for_path(result, "forms/aliased.pdf") == {"cross-component-hardlink"}
    assert len([item for item in result.review_items if item["path"] == "forms/aliased.pdf"]) == 1


def test_hardlink_from_dictionary_mapping_to_dataset_file_is_quarantined(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    os.link(tmp_path / "datasets" / "labs.csv", tmp_path / "dictionary_mapping" / "aliased.csv")

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["dictionary_mapping/aliased.csv"] == "cross-component-hardlink"
    assert _by_path(result.candidates)["dictionary_mapping/aliased.csv"].component == "_unclassified"
    assert _reasons_for_path(result, "dictionary_mapping/aliased.csv") == {"cross-component-hardlink"}
    assert len([item for item in result.review_items if item["path"] == "dictionary_mapping/aliased.csv"]) == 1


def test_hardlink_quarantine_applies_even_when_the_support_file_would_otherwise_be_unclassified(tmp_path: Path) -> None:
    # An unsupported-suffix support file that is ALSO a dataset hardlink must
    # still be reported as the hardlink, in addition to its unsupported-format
    # finding -- the aliasing is a distinct, independently-reported concern.
    _make_canonical(tmp_path)
    real = tmp_path / "datasets" / "real_data"
    real.write_bytes(b"binary-ish content")
    os.link(real, tmp_path / "forms" / "aliased.bin")

    result = inspect_intake_source(tmp_path)

    assert "cross-component-hardlink" in _reasons_for_path(result, "forms/aliased.bin")
    assert _by_path(result.candidates)["forms/aliased.bin"].component == "_unclassified"


def test_hardlink_within_datasets_is_not_quarantined(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    os.link(tmp_path / "datasets" / "labs.csv", tmp_path / "datasets" / "labs_alias.csv")

    result = inspect_intake_source(tmp_path)

    assert "cross-component-hardlink" not in _reasons(result.review_items)
    assert _by_path(result.candidates)["datasets/labs_alias.csv"].component == "datasets"


# --- cross-component dataset-copy quarantine (byte-identical, distinct identity) -----------


def test_dataset_byte_identical_copy_in_dictionary_mapping_is_quarantined_as_dataset_copy(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    dataset_bytes = (tmp_path / "datasets" / "labs.csv").read_bytes()
    (tmp_path / "dictionary_mapping" / "copy.csv").write_bytes(dataset_bytes)  # distinct inode, identical bytes

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["dictionary_mapping/copy.csv"] == "cross-component-dataset-copy"
    candidate = _by_path(result.candidates)["dictionary_mapping/copy.csv"]
    assert candidate.component == "_unclassified"
    assert _by_path(result.candidates)["datasets/labs.csv"].component == "datasets"


def test_dataset_byte_identical_copy_in_forms_is_quarantined_as_dataset_copy(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    dataset_bytes = (tmp_path / "datasets" / "labs.csv").read_bytes()
    (tmp_path / "forms" / "copy.pdf").write_bytes(dataset_bytes)

    result = inspect_intake_source(tmp_path)

    assert _review_map(result)["forms/copy.pdf"] == "cross-component-dataset-copy"
    assert _by_path(result.candidates)["forms/copy.pdf"].component == "_unclassified"


def test_dataset_copy_alone_never_satisfies_the_support_requirement(tmp_path: Path) -> None:
    # dictionary_mapping's ONLY content is a byte-identical dataset copy,
    # and no forms/ exists at all -- the support requirement must still be
    # reported missing, exactly like the hardlink-only case.
    (tmp_path / "datasets").mkdir()
    (tmp_path / "dictionary_mapping").mkdir()
    dataset_path = tmp_path / "datasets" / "labs.csv"
    dataset_path.write_text("a,b\n1,2\n", encoding="utf-8")
    (tmp_path / "dictionary_mapping" / "copy.csv").write_bytes(dataset_path.read_bytes())

    result = inspect_intake_source(tmp_path)

    review = _review_map(result)
    assert review[""] == "missing-support-component"
    assert "cross-component-dataset-copy" in _reasons_for_path(result, "dictionary_mapping/copy.csv")


def test_same_component_duplicate_bytes_across_datasets_are_unaffected(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    (tmp_path / "datasets" / "labs_dup.csv").write_bytes((tmp_path / "datasets" / "labs.csv").read_bytes())

    result = inspect_intake_source(tmp_path)

    assert "cross-component-hardlink" not in _reasons(result.review_items)
    assert "cross-component-dataset-copy" not in _reasons(result.review_items)
    assert _by_path(result.candidates)["datasets/labs_dup.csv"].component == "datasets"


def test_same_component_duplicate_bytes_across_dictionary_mapping_are_unaffected(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    (tmp_path / "dictionary_mapping" / "dict_dup.xlsx").write_bytes(
        (tmp_path / "dictionary_mapping" / "dict.xlsx").read_bytes()
    )

    result = inspect_intake_source(tmp_path)

    assert "cross-component-dataset-copy" not in _reasons(result.review_items)
    assert _by_path(result.candidates)["dictionary_mapping/dict_dup.xlsx"].component == "dictionary_mapping"


# --- package matrix (Approach 2.6 acceptance) ------------------------------------------------


def test_dataset_plus_dictionary_mapping_only_is_ready(tmp_path: Path) -> None:
    (tmp_path / "datasets").mkdir(parents=True)
    (tmp_path / "dictionary_mapping").mkdir(parents=True)
    (tmp_path / "datasets" / "labs.csv").write_text("SUBJID,AGE\n1,40\n", encoding="utf-8")
    (tmp_path / "dictionary_mapping" / "dict.csv").write_text("code,label\n1,a\n", encoding="utf-8")

    result = inspect_intake_source(tmp_path)

    assert result.review_items == ()
    assert result.errors == ()


def test_support_present_without_datasets_keeps_the_dataset_missing_reason(tmp_path: Path) -> None:
    (tmp_path / "forms").mkdir(parents=True)
    (tmp_path / "dictionary_mapping").mkdir(parents=True)
    (tmp_path / "forms" / "consent.pdf").write_bytes(b"%PDF-1.4\n%%EOF")
    (tmp_path / "dictionary_mapping" / "dict.csv").write_text("code,label\n1,a\n", encoding="utf-8")

    result = inspect_intake_source(tmp_path)

    # datasets is unconditionally required regardless of how much support
    # content is present -- support presence is irrelevant to this finding.
    assert result.review_items == ({"path": "datasets", "reason": "missing-component-directory", "blocking": True},)


def test_malformed_xls_as_the_only_dictionary_mapping_content_never_satisfies_support(tmp_path: Path) -> None:
    (tmp_path / "datasets").mkdir(parents=True)
    (tmp_path / "dictionary_mapping").mkdir(parents=True)
    (tmp_path / "datasets" / "labs.csv").write_text("SUBJID,AGE\n1,40\n", encoding="utf-8")
    (tmp_path / "dictionary_mapping" / "bad.xls").write_bytes(b"not actually a BIFF file")

    result = inspect_intake_source(tmp_path)

    review = _review_map(result)
    assert review["dictionary_mapping/bad.xls"] == "xls-workbook-invalid"
    assert review[""] == "missing-support-component"


# --- package resource-limit pre-scan (Approach 2.3) -------------------------------------------


def _canonical_entry_count() -> int:
    # 3 top-level component directories + 3 files under them.
    return 6


def _canonical_checked_bytes(root: Path) -> int:
    return sum(
        (root / rel).stat().st_size
        for rel in ("datasets/labs.csv", "forms/consent.pdf", "dictionary_mapping/dict.xlsx")
    )


def test_package_prescan_accepts_exact_entry_count_boundary(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    _make_canonical(tmp_path)
    monkeypatch.setattr(intake_preflight_module, "_MAX_PACKAGE_ENTRIES", _canonical_entry_count())

    result = inspect_intake_source(tmp_path)

    assert "intake-resource-limit" not in _reasons(result.review_items)


def test_package_prescan_rejects_entry_count_plus_one(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    _make_canonical(tmp_path)
    monkeypatch.setattr(intake_preflight_module, "_MAX_PACKAGE_ENTRIES", _canonical_entry_count() - 1)

    result = inspect_intake_source(tmp_path)

    assert result.candidates == ()
    assert result.errors == ()
    assert result.review_items == ({"path": "", "reason": "intake-resource-limit", "blocking": True},)


def test_package_prescan_accepts_exact_byte_boundary(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    _make_canonical(tmp_path)
    monkeypatch.setattr(intake_preflight_module, "_MAX_PACKAGE_CHECKED_BYTES", _canonical_checked_bytes(tmp_path))

    result = inspect_intake_source(tmp_path)

    assert "intake-resource-limit" not in _reasons(result.review_items)


def test_package_prescan_rejects_byte_boundary_plus_one(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    _make_canonical(tmp_path)
    monkeypatch.setattr(intake_preflight_module, "_MAX_PACKAGE_CHECKED_BYTES", _canonical_checked_bytes(tmp_path) - 1)

    result = inspect_intake_source(tmp_path)

    assert result.candidates == ()
    assert result.review_items == ({"path": "", "reason": "intake-resource-limit", "blocking": True},)


def test_package_prescan_accepts_exact_xls_count_boundary(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    _make_canonical(tmp_path)
    for i in range(4):
        (tmp_path / "datasets" / f"d{i}.xls").write_bytes(b"dummy")
    monkeypatch.setattr(intake_preflight_module, "_MAX_PACKAGE_XLS_FILES", 4)

    result = inspect_intake_source(tmp_path)

    assert "intake-resource-limit" not in _reasons(result.review_items)


def test_package_prescan_rejects_xls_count_plus_one(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    _make_canonical(tmp_path)
    for i in range(4):
        (tmp_path / "datasets" / f"d{i}.xls").write_bytes(b"dummy")
    monkeypatch.setattr(intake_preflight_module, "_MAX_PACKAGE_XLS_FILES", 3)

    result = inspect_intake_source(tmp_path)

    assert result.candidates == ()
    assert result.review_items == ({"path": "", "reason": "intake-resource-limit", "blocking": True},)


def test_package_prescan_deadline_expiry_discards_the_whole_package(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    _make_canonical(tmp_path)
    monkeypatch.setattr(intake_preflight_module, "_PACKAGE_WALL_SECONDS", -1)

    result = inspect_intake_source(tmp_path)

    assert result.candidates == ()
    assert result.errors == ()
    assert result.review_items == ({"path": "", "reason": "intake-resource-limit", "blocking": True},)


def test_prescan_package_accepts_the_deadline_instant_itself(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    # Direct unit-level proof of the exact-instant boundary: freeze
    # time.monotonic() so every check inside the call observes EXACTLY the
    # supplied deadline -- never strictly past it -- and confirm the
    # (private) pre-scan itself raises nothing.
    _make_canonical(tmp_path)
    frozen = 1_000_000.0
    monkeypatch.setattr(intake_preflight_module.time, "monotonic", lambda: frozen)

    root_fd = intake_preflight_module._open_pinned_root(tmp_path)
    try:
        intake_preflight_module._prescan_package(root_fd, frozen)  # no exception raised
    finally:
        os.close(root_fd)


def test_package_resource_limit_never_hashes_or_spawns_before_prescan_succeeds(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _make_canonical(tmp_path)
    monkeypatch.setattr(intake_preflight_module, "_MAX_PACKAGE_ENTRIES", 1)  # guaranteed breach

    hash_calls: list[int] = []
    real_hash = intake_preflight_module._hash_from_fd

    def spying_hash(fd: int) -> str:
        hash_calls.append(fd)
        return real_hash(fd)

    monkeypatch.setattr(intake_preflight_module, "_hash_from_fd", spying_hash)

    def failing_inspect_xls(*_args: object, **_kwargs: object) -> int:
        raise AssertionError("inspect_xls must never be called before a successful pre-scan")

    monkeypatch.setattr(xls_isolation, "inspect_xls", failing_inspect_xls)

    result = inspect_intake_source(tmp_path)

    assert hash_calls == []
    assert result.review_items == ({"path": "", "reason": "intake-resource-limit", "blocking": True},)


def test_source_unreadable_file_is_error_not_candidate_or_review(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    unreadable = tmp_path / "datasets" / "noperm.csv"
    unreadable.write_text("SUBJID,AGE\n2,50\n", encoding="utf-8")
    os.chmod(unreadable, 0o000)
    try:
        result = inspect_intake_source(tmp_path)
    finally:
        os.chmod(unreadable, 0o600)

    assert _error_map(result).get("datasets/noperm.csv") == "source-unreadable"
    assert "datasets/noperm.csv" not in _by_path(result.candidates)


def test_unreadable_nested_subtree_blocks_rather_than_silently_omits(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    blocked = tmp_path / "datasets" / "blocked"
    blocked.mkdir()
    (blocked / "unseen.csv").write_text("a,b\n1,2\n", encoding="utf-8")
    os.chmod(blocked, 0o000)
    try:
        result = inspect_intake_source(tmp_path)
    finally:
        os.chmod(blocked, 0o700)

    assert _error_map(result).get("datasets/blocked") == "source-unreadable"
    assert not any("unseen.csv" in c.relative_path for c in result.candidates)


def test_deeply_nested_source_fails_closed_without_recursion_error(tmp_path: Path) -> None:
    (tmp_path / "datasets").mkdir()
    (tmp_path / "forms").mkdir()
    (tmp_path / "forms" / "consent.pdf").write_bytes(b"%PDF-1.4")
    (tmp_path / "dictionary_mapping").mkdir()
    (tmp_path / "dictionary_mapping" / "map.csv").write_text("a,b\n1,2\n", encoding="utf-8")

    deep = tmp_path / "datasets"
    for i in range(200):
        deep = deep / f"s{i}"
        deep.mkdir()
    (deep / "unreachable.csv").write_text("a,b\n1,2\n", encoding="utf-8")

    result = inspect_intake_source(tmp_path)  # must not raise RecursionError

    assert not any("unreachable.csv" in c.relative_path for c in result.candidates)
    assert "source-unreadable" in _reasons(result.errors)


# --- real TOCTOU during preflight's own read ------------------------------------------------


def test_mutation_during_hash_is_rejected_not_a_stale_candidate(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    _make_canonical(tmp_path)
    target = tmp_path / "datasets" / "labs.csv"
    original_size = target.stat().st_size
    real_hash = intake_preflight_module._hash_from_fd
    mutated = {"done": False}

    def mutating_hash(fd: int) -> str:
        result = real_hash(fd)
        # _hash_from_fd is called once per candidate file; only mutate the
        # specific target's own hash call (identified by its distinct
        # original size, since all three canonical files differ in size),
        # and only once -- otherwise a naive "mutate on every call" would
        # rewrite the target's already-mutated content again during a LATER
        # file's unrelated hash call, converging to identical bytes/mtime
        # and making the mutation invisible to any stat-based check.
        if not mutated["done"] and os.fstat(fd).st_size == original_size:
            mutated["done"] = True
            target.write_text("SUBJID,AGE\n1,40\nMUTATED,ROW\n", encoding="utf-8")
        return result

    monkeypatch.setattr(intake_preflight_module, "_hash_from_fd", mutating_hash)

    result = inspect_intake_source(tmp_path)

    assert mutated["done"], "mutation hook never fired -- test setup is broken"
    assert "datasets/labs.csv" not in _by_path(result.candidates)
    assert _error_map(result).get("datasets/labs.csv") == "source-unreadable"
    # No stale record of the pre-mutation identity/hash survives anywhere.
    for candidate in result.candidates:
        assert candidate.relative_path != "datasets/labs.csv"


def test_mutation_during_xlsx_sheet_count_is_rejected_not_a_stale_candidate(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _make_canonical(tmp_path)
    target = tmp_path / "dictionary_mapping" / "dict.xlsx"
    real_count = intake_preflight_module.count_xlsx_sheets

    def mutating_count(stream):
        result = real_count(stream)
        target.write_bytes(target.read_bytes() + b"\x00")
        return result

    monkeypatch.setattr(intake_preflight_module, "count_xlsx_sheets", mutating_count)

    result = inspect_intake_source(tmp_path)

    assert "dictionary_mapping/dict.xlsx" not in _by_path(result.candidates)
    assert _error_map(result).get("dictionary_mapping/dict.xlsx") == "source-unreadable"


# --- reason records stay value-free ---------------------------------------------------------


def test_review_and_error_records_never_include_content_or_raw_exception_text(tmp_path: Path) -> None:
    _make_canonical(tmp_path)
    secret_marker = "TOTALLY_SECRET_PHI_VALUE"
    (tmp_path / "datasets" / "bad.xlsx").write_text(f"corrupt-{secret_marker}", encoding="utf-8")
    link = tmp_path / "datasets" / "linked.csv"
    link.symlink_to(tmp_path / "datasets" / "labs.csv")

    result = inspect_intake_source(tmp_path)

    serialized = repr(result.review_items) + repr(result.errors)
    assert secret_marker not in serialized
    fixed_codes = {
        "flat-source-layout",
        "unknown-top-level-directory",
        "missing-component-directory",
        "missing-component-content",
        "unsupported-format",
        "dataset-xlsx-multiple-sheets",
        "support-xlsx-sheet-limit",
        "source-unreadable",
        "source-target-outside-root",
        "source-symlink-not-allowed",
        "cross-component-hardlink",
        "xlsx-workbook-invalid",
    }
    for item in list(result.review_items) + list(result.errors):
        assert item["reason"] in fixed_codes
        assert isinstance(item["path"], str)  # never None, always a fixed string


def test_no_llm_dispatch_boundary_is_ever_called_during_preflight(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    # Behavioral dispatch spy: hook the actual configured model-client
    # boundary this repository uses (phi_engine.config.config.get_llm_client
    # / LLMClient.complete) so any call raises immediately, then run a full
    # preflight (including the xlsx path) and prove zero calls occurred.
    import phi_engine.config.config as config_module

    def _forbidden_get_llm_client(*args: Any, **kwargs: Any) -> Any:
        raise AssertionError("preflight must never call get_llm_client()")

    def _forbidden_complete(*args: Any, **kwargs: Any) -> Any:
        raise AssertionError("preflight must never call LLMClient.complete()")

    monkeypatch.setattr(config_module, "get_llm_client", _forbidden_get_llm_client)
    monkeypatch.setattr(config_module.LLMClient, "complete", _forbidden_complete)

    _make_canonical(tmp_path)
    result = inspect_intake_source(tmp_path)
    assert result.review_items == ()
    assert result.errors == ()


# --- count_xlsx_sheets direct contract ------------------------------------------------------


def test_count_xlsx_sheets_counts_sheet_elements(tmp_path: Path) -> None:
    path = tmp_path / "book.xlsx"
    _write_workbook(path, sheet_names=["A", "B", "C"])
    with path.open("rb") as fh:
        assert count_xlsx_sheets(fh) == 3


def test_count_xlsx_sheets_rejects_bad_zip() -> None:
    stream = io.BytesIO(b"definitely not a zip archive")
    with pytest.raises(Exception) as excinfo:
        count_xlsx_sheets(stream)
    assert excinfo.value.reason == "xlsx-workbook-invalid"


def test_count_xlsx_sheets_normalizes_a_failing_reset_seek_no_raw_text() -> None:
    # Regression: the fileobj.seek(0, os.SEEK_SET) reset that precedes the
    # bounded ZipFile construction must be inside the hostile-input
    # normalization boundary, not before it -- a filesystem-level failure
    # on that specific call must never escape as a raw OSError.
    class _FailingSeekStream:
        # Let _stream_size's own bounded seek/tell probing succeed
        # (tell, seek(0, SEEK_END), tell, seek(current, SEEK_SET) -- four
        # calls), then fail on the NEXT seek: the count_xlsx_sheets reset
        # at `fileobj.seek(0, os.SEEK_SET)` this regression targets.
        def __init__(self) -> None:
            self._seek_calls = 0

        def tell(self) -> int:
            return 0

        def seek(self, offset: int, whence: int = 0) -> int:
            self._seek_calls += 1
            if self._seek_calls > 2:
                raise OSError(5, "SENTINEL_RAW_SEEK_EIO")
            return 0

        def read(self, size: int = -1) -> bytes:
            return b""

    with pytest.raises(Exception) as excinfo:
        count_xlsx_sheets(_FailingSeekStream())
    assert excinfo.value.reason == "xlsx-workbook-invalid"
    assert "SENTINEL_RAW_SEEK_EIO" not in str(excinfo.value)


def test_count_xlsx_sheets_rejects_oversized_workbook_member(monkeypatch: pytest.MonkeyPatch) -> None:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("xl/workbook.xml", "<workbook>" + ("x" * 200) + "</workbook>")
    buf.seek(0)
    monkeypatch.setitem(DEFAULT_LIMITS, "max_zip_member_bytes", 16)
    with pytest.raises(Exception) as excinfo:
        count_xlsx_sheets(buf)
    assert excinfo.value.reason == "xlsx-workbook-invalid"


def test_count_xlsx_sheets_rejects_over_decompression_ratio_zip_bomb(monkeypatch: pytest.MonkeyPatch) -> None:
    buf = io.BytesIO()
    payload = b"0" * (5 * 1024 * 1024)
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("xl/workbook.xml", payload)
    buf.seek(0)
    monkeypatch.setitem(DEFAULT_LIMITS, "max_decompression_ratio", 2)
    monkeypatch.setitem(DEFAULT_LIMITS, "max_zip_member_bytes", 64 * 1024 * 1024)
    with pytest.raises(Exception) as excinfo:
        count_xlsx_sheets(buf)
    assert excinfo.value.reason == "xlsx-workbook-invalid"


def test_count_xlsx_sheets_rejects_source_over_max_source_bytes(monkeypatch: pytest.MonkeyPatch) -> None:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr(
            "xl/workbook.xml",
            '<?xml version="1.0"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            '<sheets><sheet name="S1" sheetId="1"/></sheets></workbook>',
        )
    buf.seek(0)
    monkeypatch.setitem(DEFAULT_LIMITS, "max_source_bytes", 4)
    with pytest.raises(Exception) as excinfo:
        count_xlsx_sheets(buf)
    assert excinfo.value.reason == "xlsx-workbook-invalid"


def test_count_xlsx_sheets_rejects_expanded_total_over_limit(monkeypatch: pytest.MonkeyPatch) -> None:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr(
            "xl/workbook.xml",
            '<?xml version="1.0"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            '<sheets><sheet name="S1" sheetId="1"/></sheets></workbook>',
        )
        zf.writestr("xl/other.xml", "y" * 5000)
    buf.seek(0)
    monkeypatch.setitem(DEFAULT_LIMITS, "max_zip_member_bytes", 1024 * 1024)
    monkeypatch.setitem(DEFAULT_LIMITS, "max_expanded_workbook_bytes", 1000)
    with pytest.raises(Exception) as excinfo:
        count_xlsx_sheets(buf)
    assert excinfo.value.reason == "xlsx-workbook-invalid"


def test_count_xlsx_sheets_rejects_central_directory_over_max_zip_directory_bytes(monkeypatch: pytest.MonkeyPatch) -> None:
    # Public-API proof that the archive's central directory itself is
    # bounded: an archive far smaller than max_source_bytes, with far fewer
    # entries than max_zip_members, is still rejected once its central
    # directory metadata alone would require more than max_zip_directory_bytes
    # to read -- the cap fires while zipfile is still parsing, not only
    # after full central-directory materialization succeeds.
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr(
            "xl/workbook.xml",
            '<?xml version="1.0"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            '<sheets><sheet name="S1" sheetId="1"/></sheets></workbook>',
        )
        for i in range(50):
            zf.writestr(f"padding/{'x' * 200}_{i}.txt", "")
    buf.seek(0)
    monkeypatch.setitem(DEFAULT_LIMITS, "max_zip_directory_bytes", 512)
    with pytest.raises(Exception) as excinfo:
        count_xlsx_sheets(buf)
    assert excinfo.value.reason == "xlsx-workbook-invalid"


def test_count_xlsx_sheets_accepts_a_valid_member_larger_than_the_directory_cap() -> None:
    # Positive boundary proof for the rearm fix: the construction-only
    # max_zip_directory_bytes budget bounds ZipFile's own central-directory
    # parse, but must NOT still be in effect once the workbook member is
    # actually streamed -- a workbook whose central directory is tiny (one
    # member) but whose compressed member content exceeds the directory cap
    # must be accepted when every approved source/member/expanded/ratio
    # limit independently passes.
    import base64
    import random

    # Deterministic, low-compressibility filler: a fixed-seed PRNG gives
    # the same bytes on every run (unlike os.urandom) while remaining
    # high-entropy enough that DEFLATE cannot shrink it below the member
    # cap this test depends on.
    padding = base64.b64encode(random.Random(20240115).randbytes(4_500_000)).decode("ascii")
    xml = (
        '<?xml version="1.0"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f"<!--{padding}--><sheets><sheet name=\"S1\" sheetId=\"1\"/></sheets></workbook>"
    )
    encoded = xml.encode("utf-8")
    assert len(encoded) > DEFAULT_LIMITS["max_zip_directory_bytes"], "test payload must exceed the directory cap"
    assert len(encoded) <= DEFAULT_LIMITS["max_zip_member_bytes"]

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("xl/workbook.xml", xml)
    buf.seek(0)

    assert count_xlsx_sheets(buf) == 1


def test_count_xlsx_sheets_stops_early_past_max_sheets(monkeypatch: pytest.MonkeyPatch) -> None:
    buf = io.BytesIO()
    sheets_xml = "".join(f'<sheet name="S{i}" sheetId="{i}" r:id="rId{i}"/>' for i in range(200))
    buf.write(_raw_workbook_zip_bytes(sheets_xml))
    buf.seek(0)
    monkeypatch.setitem(DEFAULT_LIMITS, "max_sheets", 10)
    count = count_xlsx_sheets(buf)
    assert count > 10


def test_count_xlsx_sheets_rejects_excessive_member_count(monkeypatch: pytest.MonkeyPatch) -> None:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr(
            "xl/workbook.xml",
            '<?xml version="1.0"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            '<sheets><sheet name="S1" sheetId="1"/></sheets></workbook>',
        )
        for i in range(50):
            zf.writestr(f"junk/{i}.txt", "")
    buf.seek(0)
    monkeypatch.setitem(DEFAULT_LIMITS, "max_zip_members", 10)
    with pytest.raises(Exception) as excinfo:
        count_xlsx_sheets(buf)
    assert excinfo.value.reason == "xlsx-workbook-invalid"


def test_count_xlsx_sheets_normalizes_corrupt_deflate_stream(tmp_path: Path) -> None:
    # Flip a byte inside the compressed data region of a validly-structured
    # DEFLATE member so decompression itself fails with zlib.error, proving
    # that failure is normalized rather than escaping raw.
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(
            "xl/workbook.xml",
            '<?xml version="1.0"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            '<sheets><sheet name="S1" sheetId="1"/></sheets></workbook>' + ("padding" * 200),
        )
    data = bytearray(buf.getvalue())
    # Corrupt a byte well past the local file header, inside the compressed
    # payload, without touching any ZIP structural fields.
    data[80] ^= 0xFF
    corrupt = io.BytesIO(bytes(data))
    with pytest.raises(Exception) as excinfo:
        count_xlsx_sheets(corrupt)
    assert excinfo.value.reason == "xlsx-workbook-invalid"
    assert "zlib" not in str(excinfo.value).lower()


# --- moved walker convention (single source, no duplicate walker) ---------------------------


def test_iter_source_files_is_deterministic_and_filters_junk(tmp_path: Path) -> None:
    (tmp_path / "sub").mkdir()
    (tmp_path / "sub" / "b.csv").write_text("x", encoding="utf-8")
    (tmp_path / "a.csv").write_text("x", encoding="utf-8")
    (tmp_path / ".hidden.csv").write_text("x", encoding="utf-8")
    (tmp_path / "Thumbs.db").write_text("x", encoding="utf-8")

    files = intake_preflight_module._iter_source_files(tmp_path)

    relative = [f.relative_to(tmp_path).as_posix() for f in files]
    assert relative == ["a.csv", "sub/b.csv"]


def test_iter_source_files_includes_nonhidden_symlinked_files_preserving_legacy_intake_behavior(
    tmp_path: Path,
) -> None:
    real = tmp_path / "real.csv"
    real.write_text("x", encoding="utf-8")
    link = tmp_path / "linked.csv"
    link.symlink_to(real)

    files = intake_preflight_module._iter_source_files(tmp_path)

    relative = {f.relative_to(tmp_path).as_posix() for f in files}
    assert relative == {"real.csv", "linked.csv"}


def test_iter_source_files_filters_the_complete_junk_catalog_including_nested_directories(tmp_path: Path) -> None:
    from phi_engine.utils._extraction_io.file_discovery import DEFAULT_JUNK_FILENAMES

    (tmp_path / "keep.csv").write_text("x", encoding="utf-8")
    for junk_name in sorted(DEFAULT_JUNK_FILENAMES):
        (tmp_path / junk_name).write_text("x", encoding="utf-8")
    junk_dir_name = sorted(DEFAULT_JUNK_FILENAMES)[0]
    nested_junk_dir = tmp_path / "sub" / junk_dir_name
    nested_junk_dir.mkdir(parents=True)
    (nested_junk_dir / "unreachable.csv").write_text("x", encoding="utf-8")
    (tmp_path / "sub" / "keep_nested.csv").write_text("x", encoding="utf-8")

    files = intake_preflight_module._iter_source_files(tmp_path)

    relative = {f.relative_to(tmp_path).as_posix() for f in files}
    assert relative == {"keep.csv", "sub/keep_nested.csv"}


def test_inspect_intake_source_filters_the_complete_junk_catalog_including_nested_directories(tmp_path: Path) -> None:
    from phi_engine.utils._extraction_io.file_discovery import DEFAULT_JUNK_FILENAMES

    _make_canonical(tmp_path)
    for junk_name in sorted(DEFAULT_JUNK_FILENAMES):
        (tmp_path / "datasets" / junk_name).write_text("x", encoding="utf-8")
    junk_dir_name = sorted(DEFAULT_JUNK_FILENAMES)[0]
    nested_junk_dir = tmp_path / "datasets" / "sub" / junk_dir_name
    nested_junk_dir.mkdir(parents=True)
    (nested_junk_dir / "unreachable.csv").write_text("x", encoding="utf-8")

    result = inspect_intake_source(tmp_path)

    assert result.review_items == ()
    assert result.errors == ()
    relative = {c.relative_path for c in result.candidates}
    assert relative == {"datasets/labs.csv", "forms/consent.pdf", "dictionary_mapping/dict.xlsx"}
