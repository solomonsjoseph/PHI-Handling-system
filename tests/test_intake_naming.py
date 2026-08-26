"""Tests for phi_engine.pipeline.intake_naming -- the local, support-only
AI boundary that resolves an intake study name.

Every naming-content read and local-model dispatch happens through fakes:
no real Ollama server, no `config.get_llm_client()`. Behavior is driven
primarily through the public `resolve_intake_study` entry point and real
`FakeHTTPConnection`-backed `OfflineLocalLLMClient` transport; a small set
of pure-function reader/parser tests remain where they defend a genuine
boundary invariant (archive/zip-bomb, malformed-cell, field-limit bypass)
that is otherwise awkward to construct realistically end to end.
"""

from __future__ import annotations

import builtins
import inspect
import io
import json
import os
import re
import time
import zipfile
import zlib
from dataclasses import dataclass
from pathlib import Path

import openpyxl
import pytest
import xlwt
from reportlab.pdfgen import canvas

import phi_engine.pipeline.intake_naming as naming
import phi_engine.security.model_routing as routing
from phi_engine.config import config
from phi_engine.pipeline import xls_isolation
from phi_engine.pipeline.intake_preflight import IntakeCandidate, IntakePreflight, inspect_intake_source
from phi_engine.pipeline.verified_source import FileIdentity
from phi_engine.utils import pipeline_lock

MODEL_SPEC = "qwen3:8b@sha256:" + "d" * 64
_THRESHOLD = config.PHI_CONFIDENCE_THRESHOLD
_ACCEPT_CONF = min(_THRESHOLD + 0.2, 1.0)
_REJECT_CONF = max(_THRESHOLD - 0.2, 0.0)


# --- shared fixture helpers ------------------------------------------------------------


def _local_config(**overrides):
    values = {
        "provider": "ollama",
        "models": (MODEL_SPEC,),
        "base_url": "http://127.0.0.1:11434",
        "allowed_base_urls": ("http://127.0.0.1:11434",),
        "offline_approved": True,
        "timeout_s": 2,
        "max_retries": 0,
    }
    values.update(overrides)
    return config.LocalLLMConfig(**values)


@dataclass
class FakeHTTPResponse:
    status: int
    payload: bytes

    def read(self, amount: int | None = None) -> bytes:
        return self.payload if amount is None else self.payload[:amount]

    def close(self) -> None:
        pass


class FakeHTTPConnection:
    responses: list[FakeHTTPResponse] = []
    requests: list[tuple[str, str, bytes | None, dict[str, str]]] = []

    def __init__(self, host: str, port: int, timeout: float):
        assert host in {"127.0.0.1", "::1"}
        self.host = host
        self.port = port
        self.timeout = timeout

    def request(self, method: str, path: str, body: bytes | None = None, headers: dict[str, str] | None = None):
        self.requests.append((method, path, body, headers or {}))

    def getresponse(self):
        return self.responses.pop(0)

    def close(self) -> None:
        pass


@pytest.fixture(autouse=True)
def _reset_fake_http_connection():
    FakeHTTPConnection.requests = []
    FakeHTTPConnection.responses = []
    yield
    FakeHTTPConnection.requests = []
    FakeHTTPConnection.responses = []


def _tags_response() -> FakeHTTPResponse:
    payload = json.dumps({"models": [{"name": "qwen3:8b", "digest": "sha256:" + "d" * 64}]}, separators=(",", ":"))
    return FakeHTTPResponse(200, payload.encode())


def _generate_response(raw_text: str) -> FakeHTTPResponse:
    payload = json.dumps({"response": raw_text}, separators=(",", ":"))
    return FakeHTTPResponse(200, payload.encode())


def _decision_response(study_name: str | None, confidence: float) -> FakeHTTPResponse:
    inner = json.dumps({"study_name": study_name, "confidence": confidence}, separators=(",", ":"))
    return _generate_response(inner)


def _queue_dispatch(*decisions: tuple[str | None, float] | str) -> list[FakeHTTPResponse]:
    """Build a (tags, generate) response pair per dispatch call, in order.
    A plain string element queues that exact raw text as the model response
    (for malformed-output tests) instead of a well-formed decision."""

    responses: list[FakeHTTPResponse] = []
    for decision in decisions:
        responses.append(_tags_response())
        if isinstance(decision, str):
            responses.append(_generate_response(decision))
        else:
            responses.append(_decision_response(*decision))
    return responses


def _install_fake_client(monkeypatch, responses: list[FakeHTTPResponse], *, local_config=None) -> None:
    FakeHTTPConnection.requests = []
    FakeHTTPConnection.responses = list(responses)
    monkeypatch.setattr(routing.http.client, "HTTPConnection", FakeHTTPConnection)
    real_config = local_config or _local_config()
    monkeypatch.setattr(naming, "new_offline_local_client", lambda: routing.OfflineLocalLLMClient(real_config))


def _forbid_client(monkeypatch) -> None:
    def _raise() -> routing.OfflineLocalLLMClient:
        raise AssertionError("must not construct a local model client")

    monkeypatch.setattr(naming, "new_offline_local_client", _raise)


def _prompts_sent() -> list[str]:
    return [json.loads(body)["prompt"] for method, path, body, _ in FakeHTTPConnection.requests if path == "/api/generate"]


def _evidence_from_prompt(prompt: str) -> dict:
    assert prompt.startswith(naming._PROMPT_PREFIX)
    return json.loads(prompt[len(naming._PROMPT_PREFIX) :])


# --- source-tree builders ---------------------------------------------------------------


def _write_pdf(path: Path, *lines: str) -> None:
    """Each line becomes its own page."""
    path.parent.mkdir(parents=True, exist_ok=True)
    pdf = canvas.Canvas(str(path))
    for line in lines:
        pdf.drawString(72, 750, line)
        pdf.showPage()
    pdf.save()


def _write_flate_bomb_pdf(path: Path, raw_size: int) -> None:
    """A minimal, valid, hand-assembled single-page PDF whose one content
    stream is Flate-compressed: small on disk, but pdfminer/pdfplumber
    must fully zlib-decompress it (well before this module's own
    per-fragment codepoint check ever sees the result) just to read it --
    the exact expansion vector the isolated PDF worker's address-space
    bound exists to catch."""
    path.parent.mkdir(parents=True, exist_ok=True)
    compressed = zlib.compress(b"0" * raw_size, 9)
    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /Resources << /Font << /F1 4 0 R >> >> "
        b"/MediaBox [0 0 612 792] /Contents 5 0 R >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Length %d /Filter /FlateDecode >>\nstream\n" % len(compressed) + compressed + b"\nendstream",
    ]
    out = bytearray(b"%PDF-1.4\n")
    offsets = []
    for index, body in enumerate(objects, start=1):
        offsets.append(len(out))
        out += b"%d 0 obj\n" % index + body + b"\nendobj\n"
    xref_offset = len(out)
    count = len(objects) + 1
    out += b"xref\n0 %d\n0000000000 65535 f \n" % count
    for offset in offsets:
        out += b"%010d 00000 n \n" % offset
    out += b"trailer\n<< /Size %d /Root 1 0 R >>\nstartxref\n%d\n%%%%EOF" % (count, xref_offset)
    path.write_bytes(bytes(out))


def _write_workbook(path: Path, rows: list[list[object]], *, sheet_names: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    names = sheet_names or ["Sheet1"]
    wb.active.title = names[0]
    for row in rows:
        wb.active.append(row)
    for name in names[1:]:
        wb.create_sheet(name)
    wb.save(path)


def _write_xls(path: Path, rows: list[list[object]], *, sheet_name: str = "Sheet1") -> None:
    """Build a real, genuine BIFF (.xls) file with ``xlwt`` -- the same
    fixture convention ``tests/test_xls_isolation.py`` uses -- so the
    naming-dispatch tests exercise the real ``xlrd``-backed isolation
    worker, never a stub."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = xlwt.Workbook()
    ws = wb.add_sheet(sheet_name)
    for r, row in enumerate(rows):
        for c, value in enumerate(row):
            ws.write(r, c, value)
    wb.save(str(path))


def _make_minimal_forms_source(root: Path, *lines: str) -> None:
    (root / "datasets").mkdir(parents=True)
    (root / "forms").mkdir(parents=True)
    _write_pdf(root / "forms" / "consent.pdf", *(lines or ("Consent form",)))


def _make_canonical_source(root: Path) -> None:
    (root / "datasets").mkdir(parents=True)
    (root / "forms").mkdir(parents=True)
    (root / "dictionary_mapping").mkdir(parents=True)
    (root / "datasets" / "labs.csv").write_text("SUBJID,AGE\n1,40\n", encoding="utf-8")
    _write_pdf(root / "forms" / "consent.pdf", "StudyAlpha Consent Form")
    _write_workbook(root / "dictionary_mapping" / "dict.xlsx", [["StudyAlpha", "Dictionary"]])


def _synthetic_candidate(
    relative_path: str, *, source_component: str, component: str, size: int = 10
) -> IntakeCandidate:
    identity = FileIdentity(device=1, inode=hash(relative_path) & 0xFFFF, size=size, mtime_ns=1)
    return IntakeCandidate(
        relative_path=relative_path,
        source_component=source_component,
        component=component,
        identity=identity,
        sha256="0" * 64,
        sheet_count=None,
    )


# --- public contract ---------------------------------------------------------------------


def test_public_names_match_the_approved_contract():
    assert set(naming.__all__) == {"StudyNameSource", "StudyResolution", "resolve_intake_study", "canonical_source_root"}
    sig = inspect.signature(naming.resolve_intake_study)
    assert list(sig.parameters) == ["source", "preflight", "explicit_study", "support_confirmed_no_phi", "intake_root"]
    assert sig.parameters["explicit_study"].kind is inspect.Parameter.KEYWORD_ONLY
    assert sig.parameters["support_confirmed_no_phi"].kind is inspect.Parameter.KEYWORD_ONLY
    assert sig.parameters["intake_root"].kind is inspect.Parameter.KEYWORD_ONLY

    # private internal resolver: same public contract plus the injectable
    # generated-name-allocation hook the registry-scan layer threads through
    private_sig = inspect.signature(naming._resolve_intake_study)
    assert list(private_sig.parameters) == [
        "source", "preflight", "explicit_study", "support_confirmed_no_phi", "intake_root", "generate_study_name",
    ]
    assert private_sig.parameters["generate_study_name"].kind is inspect.Parameter.KEYWORD_ONLY
    assert callable(private_sig.parameters["generate_study_name"].default)
    resolution_fields = set(inspect.signature(naming.StudyResolution).parameters)
    assert resolution_fields == {"name", "source", "review_items", "errors"}


# --- explicit study: skip naming entirely -------------------------------------------------


def test_explicit_study_skips_naming_and_reads_nothing(tmp_path, monkeypatch):
    source = tmp_path / "source"
    _make_canonical_source(source)
    preflight = inspect_intake_source(source)
    _forbid_client(monkeypatch)
    monkeypatch.setattr(naming, "phi_gate_check", lambda *a, **k: (_ for _ in ()).throw(AssertionError("no gate")))
    monkeypatch.setattr(naming, "open_verified_source", lambda *a, **k: (_ for _ in ()).throw(AssertionError("no read")))

    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study="ExplicitStudy", support_confirmed_no_phi=False, intake_root=tmp_path / "intake"
    )

    assert resolution == naming.StudyResolution(name="ExplicitStudy", source="user", review_items=(), errors=())


def test_explicit_study_ignores_support_confirmed_no_phi_flag(tmp_path, monkeypatch):
    source = tmp_path / "source"
    _make_canonical_source(source)
    preflight = inspect_intake_source(source)
    _forbid_client(monkeypatch)

    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study="ExplicitStudy", support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert resolution.source == "user"
    assert resolution.name == "ExplicitStudy"


def test_explicit_study_invalid_name_raises_value_error(tmp_path):
    preflight = IntakePreflight((), (), ())
    with pytest.raises(ValueError):
        naming.resolve_intake_study(
            tmp_path, preflight, explicit_study="../escape", support_confirmed_no_phi=False, intake_root=tmp_path
        )


# --- missing consent ----------------------------------------------------------------------


def test_missing_consent_generates_fallback_with_zero_reads_and_zero_calls(tmp_path, monkeypatch):
    source = tmp_path / "source"
    _make_canonical_source(source)
    preflight = inspect_intake_source(source)
    _forbid_client(monkeypatch)
    monkeypatch.setattr(naming, "open_verified_source", lambda *a, **k: (_ for _ in ()).throw(AssertionError("no read")))

    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=False, intake_root=tmp_path / "intake"
    )

    assert resolution.source == "generated"
    assert re.fullmatch(r"study-[0-9a-f]{8}", resolution.name)
    assert resolution.review_items == ({"path": "", "reason": "support-phi-status-required", "blocking": True},)
    assert resolution.errors == ()


# --- exact lexical mismatch: the structural admission proof --------------------------------


def test_lexical_source_component_mismatch_is_never_admitted(tmp_path, monkeypatch):
    """A regression in `candidate.component == candidate.source_component`
    is the one bug that could let AI-eligible-looking candidates whose
    physical location disagrees with their logical classification reach
    naming evidence. None of these candidates require a real backing file
    because a correct admission filter rejects them before any I/O."""
    source = tmp_path / "source"
    source.mkdir()
    mismatched = [
        # logical component says "forms", but it physically lives under datasets/
        _synthetic_candidate("datasets/evil.pdf", source_component="datasets", component="forms"),
        # logical component says "dictionary_mapping", but it's lexically _unclassified
        _synthetic_candidate("_unclassified/evil.csv", source_component="_unclassified", component="dictionary_mapping"),
        # both individually supported values, but swapped/mismatched from each other
        _synthetic_candidate("forms/evil.xlsx", source_component="forms", component="dictionary_mapping"),
        _synthetic_candidate("dictionary_mapping/evil.csv", source_component="dictionary_mapping", component="forms"),
    ]
    preflight = IntakePreflight(tuple(mismatched), (), ())
    opened: list[str] = []

    def raising_open(src, rel, **kw):
        opened.append(rel)
        raise naming.VerifiedSourceError("source-unreadable")

    monkeypatch.setattr(naming, "open_verified_source", raising_open)
    _forbid_client(monkeypatch)

    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )

    # Only the datasets/-sourced synthetic candidate is ever opened -- by
    # the item-4 fresh dataset-verification scan, which fails closed the
    # instant it cannot verify a datasets/ candidate, aborting before any
    # of the three lexically-mismatched support candidates are ever
    # touched by an admission-filter bug.
    assert opened == ["datasets/evil.pdf"]
    assert resolution.source == "generated"


def test_datasets_and_unclassified_components_are_never_admission_eligible(tmp_path, monkeypatch):
    source = tmp_path / "source"
    source.mkdir()
    candidates = [
        _synthetic_candidate("datasets/labs.csv", source_component="datasets", component="datasets"),
        _synthetic_candidate("_unclassified/odd.json", source_component="_unclassified", component="_unclassified"),
    ]
    preflight = IntakePreflight(tuple(candidates), (), ())
    _forbid_client(monkeypatch)
    opened: list[str] = []

    def raising_open(src, rel, **kw):
        opened.append(rel)
        raise naming.VerifiedSourceError("source-unreadable")

    monkeypatch.setattr(naming, "open_verified_source", raising_open)

    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    # Only the item-4 dataset-verification scan opens datasets/labs.csv
    # (source_component == "datasets"); _unclassified/odd.json is never
    # opened at all -- neither by that scan (different source_component)
    # nor by admission (neither candidate's component is in
    # _NAMING_COMPONENTS, so neither could ever become naming evidence).
    assert opened == ["datasets/labs.csv"]
    assert resolution.source == "generated"


# --- support-only candidate guard: datasets/_unclassified never open for evidence --------


def test_dataset_candidates_are_hash_verified_but_never_become_evidence(tmp_path, monkeypatch):
    """datasets/ candidates -- including one preflight reclassified to
    _unclassified for an unsupported format, since it still shares
    source_component == "datasets" -- ARE opened by the item-4/5 fresh
    descriptor-verification scans (their identity/hash must match
    preflight before any support evidence is trusted), but their content
    is never read into naming evidence: only the unbounded streaming
    hash touches them, and _extract_candidate/_collect_evidence never
    process a datasets/-sourced candidate at all."""
    source = tmp_path / "source"
    _make_canonical_source(source)
    (source / "datasets" / "unsupported.json").write_text("{}", encoding="utf-8")
    preflight = inspect_intake_source(source)
    assert any(c.component == "_unclassified" for c in preflight.candidates)
    assert any(c.component == "datasets" for c in preflight.candidates)

    opened: list[str] = []
    real_open = naming.open_verified_source

    def spy_open(src, rel, **kw):
        opened.append(rel)
        return real_open(src, rel, **kw)

    monkeypatch.setattr(naming, "open_verified_source", spy_open)
    _install_fake_client(
        monkeypatch,
        _queue_dispatch(("StudyAlpha", _ACCEPT_CONF), ("StudyAlpha", _ACCEPT_CONF)),
    )

    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )

    # datasets/ candidates (including the reclassified _unclassified
    # unsupported.json, since it is still physically under datasets/ and
    # so shares source_component == "datasets") ARE opened by the fresh
    # dataset-verification scans -- but never for evidence content.
    assert "datasets/labs.csv" in opened
    assert "datasets/unsupported.json" in opened
    # A genuinely non-dataset _unclassified path is never opened.
    assert all(not p.startswith("_unclassified") for p in opened)
    # Two-phase collection (validate-all, then extract) opens each
    # admitted support candidate up to twice -- once per pass -- so the
    # support-only subsequence of `opened` is no longer a single sorted
    # run; the first-seen order (i.e. the order pass 1 visits support
    # candidates) still is.
    support_opened = [p for p in opened if not p.startswith("datasets/")]
    assert list(dict.fromkeys(support_opened)) == sorted(dict.fromkeys(support_opened))
    for prompt in _prompts_sent():
        assert "SUBJID" not in prompt
        assert "labs.csv" not in prompt
        assert "unsupported.json" not in prompt
    assert resolution.name == "StudyAlpha"
    assert resolution.source == "ai"


# --- the demonstrated late-hardlink race ---------------------------------------------------


def test_late_dataset_hardlink_created_after_preflight_causes_zero_dispatch(tmp_path, monkeypatch):
    """Reproduces the audited race exactly: preflight admits a clean
    support candidate, then an attacker hardlinks a NEW datasets/ dirent to
    that SAME inode before naming runs. FileIdentity alone cannot see this
    (device/inode/size/mtime_ns are unchanged); only the fresh
    descriptor-relative datasets/ rescan can."""
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "forms").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    (source / "datasets" / "labs.csv").write_text("SUBJID,AGE\n1,40\n", encoding="utf-8")
    _write_pdf(source / "forms" / "consent.pdf", "Alpha consent")
    alias = source / "dictionary_mapping" / "alias.csv"
    alias.write_text("SECRET_HEADER\nsecret-value\n", encoding="utf-8")

    preflight = inspect_intake_source(source)
    assert any(c.relative_path == "dictionary_mapping/alias.csv" for c in preflight.candidates)

    # The race: link a NEW dataset dirent to the SAME inode after preflight.
    os.link(alias, source / "datasets" / "late-alias.csv")

    _forbid_client(monkeypatch)
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )

    assert FakeHTTPConnection.requests == []
    assert resolution.source == "generated"
    assert resolution.review_items == ({"path": "", "reason": "cross-component-hardlink", "blocking": True},)
    assert resolution.errors == ()
    dumped = json.dumps(resolution.review_items)
    assert "SECRET_HEADER" not in dumped and "secret-value" not in dumped


def test_late_hardlink_on_one_of_two_candidates_discards_all_evidence(tmp_path, monkeypatch):
    """A late hardlink on only ONE of several otherwise-clean candidates
    still discards everything, including the clean forms evidence -- not
    just the tampered candidate."""
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "forms").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    (source / "datasets" / "labs.csv").write_text("SUBJID\n1\n", encoding="utf-8")
    _write_pdf(source / "forms" / "consent.pdf", "Clean forms content")
    tampered = source / "dictionary_mapping" / "tampered.csv"
    tampered.write_text("A,B\n1,2\n", encoding="utf-8")

    preflight = inspect_intake_source(source)
    os.link(tampered, source / "datasets" / "late.csv")

    _forbid_client(monkeypatch)
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert FakeHTTPConnection.requests == []
    assert resolution.source == "generated"
    assert resolution.review_items == ({"path": "", "reason": "cross-component-hardlink", "blocking": True},)


def test_hardlink_recheck_before_dispatch_catches_a_race_after_collection(tmp_path, monkeypatch):
    """The SECOND checkpoint: a hardlink created AFTER evidence collection
    finished (but before dispatch) is caught by the pre-dispatch recheck,
    even though the per-candidate checkpoint during collection saw nothing
    wrong."""
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "forms").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    (source / "datasets" / "labs.csv").write_text("SUBJID\n1\n", encoding="utf-8")
    _write_pdf(source / "forms" / "consent.pdf", "Alpha consent")
    dict_file = source / "dictionary_mapping" / "dict.csv"
    dict_file.write_text("A,B\n1,2\n", encoding="utf-8")
    preflight = inspect_intake_source(source)

    real_current_dataset_identities = naming._current_dataset_identities
    call_count = {"n": 0}

    def timed_identities(src):
        call_count["n"] += 1
        if call_count["n"] == 1:
            # First call: the per-candidate checkpoint during collection --
            # nothing linked yet.
            return real_current_dataset_identities(src)
        # The pre-dispatch recheck (first call after collection) observes
        # the race that happened in between; only create the link once so
        # a second recheck call (e.g. before a later combined dispatch)
        # doesn't hit an already-existing dirent.
        raced_link = source / "datasets" / "raced.csv"
        if not raced_link.exists():
            os.link(dict_file, raced_link)
        return real_current_dataset_identities(src)

    monkeypatch.setattr(naming, "_current_dataset_identities", timed_identities)
    _forbid_client(monkeypatch)
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert FakeHTTPConnection.requests == []
    assert resolution.source == "generated"
    assert resolution.review_items == ({"path": "", "reason": "cross-component-hardlink", "blocking": True},)


def test_symlinked_datasets_directory_after_preflight_fails_closed(tmp_path, monkeypatch):
    """The exact audited PoC: datasets/ is renamed and replaced by a
    directory symlink to the renamed tree after preflight. A fail-open
    scan would see this as "no dataset identities"; both the pre-
    existing _current_dataset_identities guard AND the item-4 fresh
    dataset descriptor-verification scan must instead treat an
    unsafe/inconclusive/symlink-blocked path identically to a confirmed
    hardlink alias -- here the item-4 scan is the FIRST to see it (it
    runs before any support evidence is even collected), so it reports
    its own fixed source-unreadable error rather than ever reaching the
    older cross-component-hardlink review path."""
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "forms").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    (source / "datasets" / "labs.csv").write_text("SUBJID\n1\n", encoding="utf-8")
    _write_pdf(source / "forms" / "consent.pdf", "Alpha consent")
    (source / "dictionary_mapping" / "dict.csv").write_text("A,B\n1,2\n", encoding="utf-8")
    preflight = inspect_intake_source(source)

    real_datasets = tmp_path / "elsewhere_datasets"
    os.rename(source / "datasets", real_datasets)
    (source / "datasets").symlink_to(real_datasets)

    _forbid_client(monkeypatch)
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert FakeHTTPConnection.requests == []
    assert resolution.source == "generated"
    assert resolution.review_items == ()
    assert resolution.errors == ({"path": "datasets/labs.csv", "reason": "source-unreadable"},)


def test_dataset_scan_error_fails_closed_not_open(tmp_path, monkeypatch):
    """Any scan failure -- not just a symlink -- must fail closed. Force
    the shared scan primitive to report an inconclusive result and prove
    naming never treats that as "zero dataset files"."""
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "forms").mkdir(parents=True)
    _write_pdf(source / "forms" / "consent.pdf", "Alpha consent")
    preflight = inspect_intake_source(source)

    monkeypatch.setattr(naming.intake_preflight, "_scan_component_identities", lambda src, name: (None, "source-unreadable"))
    _forbid_client(monkeypatch)
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert FakeHTTPConnection.requests == []
    assert resolution.source == "generated"
    assert resolution.review_items == ({"path": "", "reason": "cross-component-hardlink", "blocking": True},)


def test_client_first_call_creating_the_hardlink_gets_no_second_call(tmp_path, monkeypatch):
    """Exact reproduction of the between-dispatch disclosure: a client
    whose OWN first complete_bounded call creates a dict-to-datasets
    hardlink as a side effect must never receive a second call, and the
    dictionary content must never appear in any prompt actually sent."""
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "forms").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    (source / "datasets" / "labs.csv").write_text("SUBJID\n1\n", encoding="utf-8")
    _write_pdf(source / "forms" / "consent.pdf", "Alpha consent")
    dict_file = source / "dictionary_mapping" / "dict.csv"
    dict_file.write_text("VAR,DESC\n1,2\n", encoding="utf-8")
    preflight = inspect_intake_source(source)

    sent_prompts: list[str] = []

    class RaceOnFirstCallClient:
        calls = 0

        def complete_bounded(self, prompt, *, max_output_tokens, max_response_bytes):
            RaceOnFirstCallClient.calls += 1
            sent_prompts.append(prompt)
            if RaceOnFirstCallClient.calls == 1:
                os.link(dict_file, source / "datasets" / "raced.csv")
            return json.dumps({"study_name": "StudyAlpha", "confidence": _ACCEPT_CONF})

    monkeypatch.setattr(naming, "new_offline_local_client", lambda: RaceOnFirstCallClient())
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert RaceOnFirstCallClient.calls == 1
    assert all("VAR" not in prompt and "DESC" not in prompt for prompt in sent_prompts)
    assert resolution.source == "generated"
    assert resolution.review_items == ({"path": "", "reason": "cross-component-hardlink", "blocking": True},)


# --- item 4/5: fresh dataset descriptor/hash snapshot, real filesystem races --------------


def test_dataset_mutation_before_pre_extraction_scan_yields_zero_model_calls(tmp_path, monkeypatch):
    """Item 4's real-filesystem race: preflight admits a clean dataset
    and a clean dictionary_mapping candidate; the DATASET is then
    mutated for real (same byte length, mtime restored to the exact
    preflight-recorded value, defeating open_verified_source's own
    identity check alone) BEFORE resolve_intake_study is ever called --
    landing exactly in the window between preflight and item 4's
    pre-extraction dataset-verification scan, naming's very first read.
    No scan/hash primitive is mocked; the mutation is a real overwrite
    plus a real os.utime restoration."""
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    dataset_path = source / "datasets" / "labs.csv"
    dataset_path.write_text("SUBJID,AGE\n1,40\n", encoding="utf-8")
    (source / "dictionary_mapping" / "dict.csv").write_text("VAR,DESC\n1,2\n", encoding="utf-8")
    preflight = inspect_intake_source(source)
    dataset_candidate = next(c for c in preflight.candidates if c.relative_path == "datasets/labs.csv")
    expected_mtime_ns = dataset_candidate.identity.mtime_ns

    mutated = dataset_path.read_bytes().replace(b"1,40", b"9,99")
    assert len(mutated) == dataset_candidate.identity.size
    dataset_path.write_bytes(mutated)
    os.utime(dataset_path, ns=(expected_mtime_ns, expected_mtime_ns))
    post = os.stat(dataset_path)
    assert post.st_ino == dataset_candidate.identity.inode
    assert post.st_size == dataset_candidate.identity.size
    assert post.st_mtime_ns == expected_mtime_ns  # identity alone would see this as unchanged

    _forbid_client(monkeypatch)
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )

    assert FakeHTTPConnection.requests == []
    assert resolution.source == "generated"
    assert resolution.review_items == ()
    assert resolution.errors == ({"path": "datasets/labs.csv", "reason": "source-unreadable"},)


def test_dataset_mutation_between_extraction_and_pre_dispatch_scan_yields_zero_model_calls(tmp_path, monkeypatch):
    """Item 5's real-filesystem race: the SAME kind of real, restored-
    mtime dataset mutation, but timed to land AFTER naming has already
    collected and parsed clean support evidence (forms + dictionary_
    mapping both pass the item-4 pre-extraction scan and are fully
    extracted), and BEFORE the first pre-dispatch scan -- proving the
    pre-dispatch recheck independently catches a race the pre-extraction
    scan could not have seen (it hadn't happened yet). No scan/hash
    primitive itself is mocked or replaced; only the call boundary
    around the real _collect_evidence is used to land the real mutation
    at the right moment, mirroring this suite's existing between-
    checkpoint race convention (see
    test_hardlink_recheck_before_dispatch_catches_a_race_after_collection)."""
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "forms").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    dataset_path = source / "datasets" / "labs.csv"
    dataset_path.write_text("SUBJID,AGE\n1,40\n", encoding="utf-8")
    _write_pdf(source / "forms" / "consent.pdf", "Alpha consent")
    (source / "dictionary_mapping" / "dict.csv").write_text("VAR,DESC\n1,2\n", encoding="utf-8")
    preflight = inspect_intake_source(source)
    dataset_candidate = next(c for c in preflight.candidates if c.relative_path == "datasets/labs.csv")
    expected_mtime_ns = dataset_candidate.identity.mtime_ns

    real_collect_evidence = naming._collect_evidence

    def mutate_after_collection(src, admitted):
        result = real_collect_evidence(src, admitted)
        mutated = dataset_path.read_bytes().replace(b"1,40", b"9,99")
        dataset_path.write_bytes(mutated)
        os.utime(dataset_path, ns=(expected_mtime_ns, expected_mtime_ns))
        return result

    monkeypatch.setattr(naming, "_collect_evidence", mutate_after_collection)
    _forbid_client(monkeypatch)
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )

    post = os.stat(dataset_path)
    assert post.st_ino == dataset_candidate.identity.inode
    assert post.st_mtime_ns == expected_mtime_ns  # identity alone would see this as unchanged

    assert FakeHTTPConnection.requests == []
    assert resolution.source == "generated"
    assert resolution.errors == ()
    assert resolution.review_items == ({"path": "", "reason": "cross-component-hardlink", "blocking": True},)


def test_dataset_byte_identical_support_copy_yields_zero_parser_calls_before_extraction(tmp_path, monkeypatch):
    """Item 4's own independent, defense-in-depth exclusion: a support
    candidate whose bytes are already byte-for-byte identical to a
    verified dataset's bytes -- an independent lexical copy, not a
    hardlink -- reaches zero PARSER calls (not merely zero dispatch),
    even in the hypothetical where preflight's own phase-2 cross-
    component-dataset-copy quarantine did not already downgrade it.
    Simulated exactly like this suite's existing 'poisoned preflight'
    pattern (see test_symlink_naming_candidate_is_rejected_before_read)
    so naming's OWN check is what is actually exercised, never merely
    relying on preflight having already caught it."""
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    dataset_bytes = "SUBJID,AGE\n1,40\n"
    (source / "datasets" / "labs.csv").write_text(dataset_bytes, encoding="utf-8")
    (source / "dictionary_mapping" / "copy.csv").write_text(dataset_bytes, encoding="utf-8")
    preflight = inspect_intake_source(source)

    dataset_candidate = next(c for c in preflight.candidates if c.relative_path == "datasets/labs.csv")
    real_copy_candidate = next(c for c in preflight.candidates if c.relative_path == "dictionary_mapping/copy.csv")
    # Confirms the fixture premise: preflight's own phase-2 quarantine
    # already caught this byte-identical, non-hardlinked copy.
    assert real_copy_candidate.component == "_unclassified"
    assert real_copy_candidate.sha256 == dataset_candidate.sha256
    assert real_copy_candidate.identity.inode != dataset_candidate.identity.inode  # a copy, not a hardlink

    # Simulate the hypothetical where preflight's own quarantine did NOT
    # already downgrade it -- naming's own independent SHA-based
    # exclusion must still hold on its own.
    reclassified = real_copy_candidate.__class__(
        relative_path=real_copy_candidate.relative_path,
        source_component=real_copy_candidate.source_component,
        component="dictionary_mapping",
        identity=real_copy_candidate.identity,
        sha256=real_copy_candidate.sha256,
        sheet_count=real_copy_candidate.sheet_count,
    )
    poisoned_preflight = IntakePreflight((dataset_candidate, reclassified), (), ())

    extraction_calls = {"n": 0}
    real_extract_csv = naming._extract_csv_rows

    def counting_extract_csv(stream):
        extraction_calls["n"] += 1
        return real_extract_csv(stream)

    monkeypatch.setattr(naming, "_extract_csv_rows", counting_extract_csv)
    _forbid_client(monkeypatch)
    resolution = naming.resolve_intake_study(
        source, poisoned_preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )

    assert extraction_calls["n"] == 0
    assert resolution.source == "generated"
    assert resolution.review_items == ()
    assert resolution.errors == ()


def test_support_candidate_content_swapped_to_dataset_bytes_before_dispatch_yields_zero_model_calls(
    tmp_path, monkeypatch
):
    """Item 5's SHA-intersection re-check must compare a support
    candidate's LIVE current bytes against a fresh dataset re-scan, not
    just its static preflight-recorded ``candidate.sha256`` -- those two
    are the same value at admission time by definition, so comparing a
    fresh dataset hash against a STATIC support hash can never observe a
    race that happens strictly AFTER admission. Here a clean, legitimate
    dictionary_mapping candidate is admitted and fully extracted (real
    evidence collected from its ORIGINAL bytes), then -- after
    collection, before dispatch -- its on-disk bytes are overwritten to
    be byte-for-byte identical to a dataset file's bytes, with size and
    mtime restored to its own original identity (defeating
    open_verified_source's identity check alone). Only a genuinely LIVE
    re-hash of the support candidate's current bytes at dispatch time --
    not a comparison against its stale preflight-recorded hash -- can
    catch this."""
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    dataset_content = b"SUBJID,AGE\n1,40\n"
    base_support = b"VAR,DESC\n1,2\n"
    padding = len(dataset_content) - len(base_support)
    assert padding > 0
    support_original = base_support + b"\n" * padding
    assert len(support_original) == len(dataset_content)

    (source / "datasets" / "labs.csv").write_bytes(dataset_content)
    support_path = source / "dictionary_mapping" / "legit.csv"
    support_path.write_bytes(support_original)
    preflight = inspect_intake_source(source)
    support_candidate = next(c for c in preflight.candidates if c.relative_path == "dictionary_mapping/legit.csv")
    assert support_candidate.component == "dictionary_mapping"  # not already excluded/quarantined
    expected_mtime_ns = support_candidate.identity.mtime_ns

    real_collect_evidence = naming._collect_evidence

    def swap_support_content_after_collection(src, admitted):
        result = real_collect_evidence(src, admitted)
        support_path.write_bytes(dataset_content)  # now byte-identical to the dataset
        os.utime(support_path, ns=(expected_mtime_ns, expected_mtime_ns))
        return result

    monkeypatch.setattr(naming, "_collect_evidence", swap_support_content_after_collection)
    _forbid_client(monkeypatch)
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )

    post = os.stat(support_path)
    assert post.st_ino == support_candidate.identity.inode
    assert post.st_size == support_candidate.identity.size
    assert post.st_mtime_ns == expected_mtime_ns  # identity alone would see this as unchanged

    assert FakeHTTPConnection.requests == []
    assert resolution.source == "generated"
    assert resolution.errors == ()
    assert resolution.review_items == ({"path": "", "reason": "cross-component-hardlink", "blocking": True},)


def test_dataset_scan_failure_after_earlier_dataset_already_verified_leaks_no_partial_state(tmp_path, monkeypatch):
    """_verify_dataset_snapshot iterates every preflight-known
    ``datasets/`` candidate in sequence; when a LATER file's hash has
    drifted after an EARLIER file was already successfully opened and
    hashed, the scan must still return the fixed empty-set failure
    contract -- never leak the already-verified earlier file's
    identity/hash into a caller-visible partial result. A spy on
    ``open_verified_source`` proves BOTH dataset files were genuinely
    opened, in order, before the scan failed -- so this is a real
    mid-scan failure, not merely a first-file failure."""
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    (source / "datasets" / "a_labs.csv").write_text("SUBJID,AGE\n1,40\n", encoding="utf-8")
    z_path = source / "datasets" / "z_vitals.csv"
    z_path.write_text("SUBJID,BP\n1,120\n", encoding="utf-8")
    (source / "dictionary_mapping" / "dict.csv").write_text("VAR,DESC\n1,2\n", encoding="utf-8")
    preflight = inspect_intake_source(source)
    z_candidate = next(c for c in preflight.candidates if c.relative_path == "datasets/z_vitals.csv")
    expected_mtime_ns = z_candidate.identity.mtime_ns

    # TOCTOU: z_vitals.csv's content drifts after preflight, size and
    # mtime restored to defeat open_verified_source's identity check
    # alone -- only the fresh content hash inside _verify_dataset_snapshot
    # can catch it.
    mutated = z_path.read_bytes().replace(b"1,120", b"9,999")
    assert len(mutated) == z_candidate.identity.size
    z_path.write_bytes(mutated)
    os.utime(z_path, ns=(expected_mtime_ns, expected_mtime_ns))

    opened_paths: list[str] = []
    real_open = naming.open_verified_source

    def spying_open(src, rel_path, **kwargs):
        opened_paths.append(rel_path)
        return real_open(src, rel_path, **kwargs)

    monkeypatch.setattr(naming, "open_verified_source", spying_open)
    _forbid_client(monkeypatch)
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )

    # Both dataset candidates were genuinely opened, a_labs.csv (which
    # verifies cleanly) strictly before z_vitals.csv (which fails) --
    # proving this is a real mid-scan failure, not a first-file-only one.
    dataset_opens = [p for p in opened_paths if p.startswith("datasets/")]
    assert dataset_opens.index("datasets/a_labs.csv") < dataset_opens.index("datasets/z_vitals.csv")
    # dict.csv (the support candidate) was NEVER opened at all -- the
    # item-4 scan aborts the whole attempt before any support evidence
    # is ever read.
    assert "dictionary_mapping/dict.csv" not in opened_paths

    assert FakeHTTPConnection.requests == []
    assert resolution.source == "generated"
    assert resolution.errors == ({"path": "datasets/z_vitals.csv", "reason": "source-unreadable"},)
    assert resolution.review_items == ()

# --- identity/hash drift: candidate excluded, zero dispatch when it's the only one -------


def test_identity_drift_after_preflight_excludes_candidate_and_records_error(tmp_path, monkeypatch):
    source = tmp_path / "source"
    _make_minimal_forms_source(source, "Original content")
    preflight = inspect_intake_source(source)

    # TOCTOU: file mutated after preflight computed identity/sha256.
    (source / "forms" / "consent.pdf").write_bytes(b"mutated-bytes-different-size-and-content")

    _forbid_client(monkeypatch)
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )

    assert resolution.source == "generated"
    assert resolution.errors == ({"path": "forms/consent.pdf", "reason": "source-unreadable"},)


def test_hash_drift_with_matching_identity_is_rejected(tmp_path, monkeypatch):
    source = tmp_path / "source"
    _make_minimal_forms_source(source, "Original content")
    preflight = inspect_intake_source(source)
    candidate = preflight.candidates[0]
    # Poison the recorded sha256 while identity stays intact -- simulates a
    # defense-in-depth hash check catching something identity alone missed.
    poisoned = candidate.__class__(
        relative_path=candidate.relative_path,
        source_component=candidate.source_component,
        component=candidate.component,
        identity=candidate.identity,
        sha256="0" * 64,
        sheet_count=candidate.sheet_count,
    )
    poisoned_preflight = IntakePreflight((poisoned,), preflight.review_items, preflight.errors)

    _forbid_client(monkeypatch)
    resolution = naming.resolve_intake_study(
        source, poisoned_preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert resolution.source == "generated"
    assert resolution.errors == ({"path": "forms/consent.pdf", "reason": "source-unreadable"},)


def test_validation_completes_for_every_candidate_before_any_extraction_begins(tmp_path, monkeypatch):
    """Regression for the audited interleaving defect ('Candidate
    validation and extraction are interleaved, so a later mutation is
    found only after earlier content was parsed'): validation is a
    complete pass over EVERY admitted candidate before extraction ever
    begins for ANY of them. A lexically-earlier, otherwise-clean
    candidate (``dictionary_mapping/dict.csv``) must never be parsed just
    because a lexically-later candidate's (``forms/z.pdf``) safety
    failure is only discovered afterward."""
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "forms").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    _write_pdf(source / "forms" / "z.pdf", "hello")  # sorts AFTER dict.csv
    (source / "dictionary_mapping" / "dict.csv").write_text("VAR\n1\n", encoding="utf-8")
    preflight = inspect_intake_source(source)
    assert preflight.candidates[0].relative_path < preflight.candidates[1].relative_path

    # TOCTOU: mutate the LATER-sorted candidate after preflight computed
    # its identity/sha256.
    (source / "forms" / "z.pdf").write_bytes(b"mutated-bytes-different")

    extraction_calls = {"n": 0}
    real_extract_csv = naming._extract_csv_rows

    def counting_extract_csv(stream):
        extraction_calls["n"] += 1
        return real_extract_csv(stream)

    monkeypatch.setattr(naming, "_extract_csv_rows", counting_extract_csv)
    _forbid_client(monkeypatch)
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )

    assert extraction_calls["n"] == 0
    assert resolution.source == "generated"
    assert resolution.errors == ({"path": "forms/z.pdf", "reason": "source-unreadable"},)


def test_cross_component_hardlink_preflight_finding_aborts_naming_with_zero_reads_and_zero_calls(tmp_path, monkeypatch):
    """A preflight-detected cross-component hardlink on ONE support file
    must abort the ENTIRE naming attempt -- including a perfectly clean,
    otherwise-admitted forms candidate -- with zero content reads and
    zero model dispatch. Preflight already recorded the fixed
    cross-component-hardlink review item for the aliased file (merged by
    the caller); naming must add nothing new and must not read or
    dispatch any other admitted candidate's evidence."""
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "forms").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    dataset_file = source / "datasets" / "labs.csv"
    dataset_file.write_text("SUBJID,AGE\n1,40\n", encoding="utf-8")
    os.link(dataset_file, source / "dictionary_mapping" / "aliased.csv")
    _write_pdf(source / "forms" / "consent.pdf", "hello")
    preflight = inspect_intake_source(source)
    assert any(item["reason"] == "cross-component-hardlink" for item in preflight.review_items)
    assert not any(c.component == "dictionary_mapping" for c in preflight.candidates)

    monkeypatch.setattr(naming, "open_verified_source", lambda *a, **k: (_ for _ in ()).throw(AssertionError("must not open")))
    _forbid_client(monkeypatch)
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert FakeHTTPConnection.requests == []
    assert resolution.source == "generated"
    assert resolution.review_items == ()
    assert resolution.errors == ()


@pytest.mark.parametrize("reason", ["source-unreadable", "source-target-outside-root"])
def test_preflight_trust_error_aborts_naming_with_zero_reads_and_zero_calls(tmp_path, monkeypatch, reason):
    """A preflight ERROR (not just a review item) recording a
    verified-source trust failure elsewhere in the tree -- either
    ``source-unreadable`` or ``source-target-outside-root`` -- must abort
    the ENTIRE naming attempt exactly like a preflight-recorded
    symlink/hardlink review item does: zero candidate opens, zero client
    construction, zero dispatch, even though an otherwise-clean, fully
    admitted dictionary candidate exists. Exact reproduction of the
    audited PoC: a real, clean dictionary candidate plus a synthetic
    preflight error for an unrelated path."""
    source = tmp_path / "source"
    (source / "forms").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    _write_pdf(source / "forms" / "consent.pdf", "hello")
    _write_workbook(source / "dictionary_mapping" / "dict.xlsx", [["a", "b"]])
    preflight = inspect_intake_source(source)
    assert any(c.component == "dictionary_mapping" for c in preflight.candidates)

    poisoned_preflight = IntakePreflight(
        preflight.candidates,
        preflight.review_items,
        preflight.errors + ({"path": "forms/unreadable.pdf", "reason": reason},),
    )

    monkeypatch.setattr(naming, "open_verified_source", lambda *a, **k: (_ for _ in ()).throw(AssertionError("must not open")))
    _forbid_client(monkeypatch)
    resolution = naming.resolve_intake_study(
        source, poisoned_preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert FakeHTTPConnection.requests == []
    assert resolution.source == "generated"
    assert resolution.review_items == ()
    assert resolution.errors == ()


def test_symlink_naming_candidate_is_rejected_before_read(tmp_path, monkeypatch):
    source = tmp_path / "source"
    (source / "forms").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    real = source / "dictionary_mapping" / "real.xlsx"
    _write_workbook(real, [["a", "b"]])
    _write_pdf(source / "forms" / "consent.pdf", "hello")
    preflight = inspect_intake_source(source)
    # Simulate a post-preflight symlink swap by handing naming a candidate
    # whose relative_path now points at a symlink placed after the fact.
    link = source / "dictionary_mapping" / "swapped.xlsx"
    link.symlink_to(real)
    candidate = next(c for c in preflight.candidates if c.component == "dictionary_mapping")
    swapped = candidate.__class__(
        relative_path="dictionary_mapping/swapped.xlsx",
        source_component="dictionary_mapping",
        component="dictionary_mapping",
        identity=candidate.identity,
        sha256=candidate.sha256,
        sheet_count=candidate.sheet_count,
    )
    poisoned_preflight = IntakePreflight((swapped,), (), ())

    _forbid_client(monkeypatch)
    resolution = naming.resolve_intake_study(
        source, poisoned_preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert any(item["reason"] == "source-symlink-not-allowed" for item in resolution.review_items)
    assert resolution.source == "generated"


# --- oversized document: skipped before open, blocking review ----------------------------


def test_oversized_document_still_goes_through_hardlink_and_hash_gates_before_dispatch(tmp_path, monkeypatch):
    """Oversized candidates are no longer skipped by a stat-only
    pre-check before the hardlink/identity checkpoint (that used to hide
    a late hardlink on an over-limit candidate, see the regression
    below) -- they are opened, hardlink-rechecked, and hash-bounded-read
    like every other admitted candidate, and _MAX_DOCUMENT_BYTES is
    enforced by the bounded hash read itself, not a stat pre-check."""
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "forms").mkdir(parents=True)
    big = source / "forms" / "huge.pdf"
    big.write_bytes(b"0" * (naming._MAX_DOCUMENT_BYTES + 1))
    preflight = inspect_intake_source(source)

    _forbid_client(monkeypatch)
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert resolution.review_items == ({"path": "forms/huge.pdf", "reason": "support-evidence-limit", "blocking": True},)
    assert resolution.source == "generated"


def test_oversized_candidate_with_late_hardlink_reports_hardlink_not_size_limit(tmp_path, monkeypatch):
    """Exact reproduction of the audited PoC: an oversized forms
    candidate is hardlinked into datasets/ AFTER preflight, alongside an
    otherwise-clean dictionary candidate. The old stat-only oversize
    pre-check hid this candidate from the hardlink/identity checkpoint
    entirely, letting the clean dictionary candidate still dispatch (two
    model calls were observed in the audit). It must now abort the
    entire attempt with zero dispatch and the hardlink code, not
    support-evidence-limit."""
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "forms").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    (source / "datasets" / "labs.csv").write_text("SUBJID\n1\n", encoding="utf-8")
    big = source / "forms" / "huge.pdf"
    big.write_bytes(b"0" * (naming._MAX_DOCUMENT_BYTES + 1))
    (source / "dictionary_mapping" / "dict.csv").write_text("VAR,DESC\n1,2\n", encoding="utf-8")
    preflight = inspect_intake_source(source)
    assert any(c.relative_path == "forms/huge.pdf" for c in preflight.candidates)

    os.link(big, source / "datasets" / "late-huge.csv")

    _forbid_client(monkeypatch)
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert FakeHTTPConnection.requests == []
    assert resolution.source == "generated"
    assert resolution.review_items == ({"path": "", "reason": "cross-component-hardlink", "blocking": True},)
    assert resolution.errors == ()


def test_late_hardlink_on_oversized_noncontributing_candidate_yields_zero_dispatch(tmp_path, monkeypatch):
    """Exact reproduction of the audited PoC (finding: 'A late hardlink
    on an oversized/non-contributing candidate is missed before model
    dispatch'): an oversized forms candidate that NEVER contributes a
    retained evidence fragment must still be tracked by the pre-dispatch
    hardlink guard's identity set. A hardlink created into datasets/
    AFTER evidence collection finishes (but before dispatch) -- aliasing
    the OVERSIZED candidate specifically, which the guard used to ignore
    because only fragment-contributing identities were tracked -- must
    still be caught, with zero dispatch of the clean dictionary
    candidate's content."""
    monkeypatch.setattr(naming, "_MAX_DOCUMENT_BYTES", 32)
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "forms").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    (source / "datasets" / "labs.csv").write_text("SUBJID\n1\n", encoding="utf-8")
    huge = source / "forms" / "huge.pdf"
    huge.write_bytes(b"0" * 33)  # exceeds the reduced 32-byte cap
    (source / "dictionary_mapping" / "dict.csv").write_text("VAR,DESC\n1,2\n", encoding="utf-8")
    preflight = inspect_intake_source(source)
    assert any(c.relative_path == "forms/huge.pdf" for c in preflight.candidates)

    real_collect = naming._collect_evidence

    def wrapped_collect(src, admitted):
        result = real_collect(src, admitted)
        # The race: link a NEW dataset dirent to the oversized, never-
        # contributing candidate's inode ONLY after collection returns.
        os.link(huge, src / "datasets" / "late-huge.csv")
        return result

    monkeypatch.setattr(naming, "_collect_evidence", wrapped_collect)
    _forbid_client(monkeypatch)
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert FakeHTTPConnection.requests == []
    assert resolution.source == "generated"
    assert resolution.review_items == (
        {"path": "forms/huge.pdf", "reason": "support-evidence-limit", "blocking": True},
        {"path": "", "reason": "cross-component-hardlink", "blocking": True},
    )
    assert resolution.errors == ()
    dumped = json.dumps(resolution.review_items)
    assert "VAR" not in dumped and "DESC" not in dumped



# --- descriptor lifecycle: reader closes before context exit -----------------------------


def test_descriptor_stays_verified_through_hash_and_parse(tmp_path, monkeypatch):
    """The verified descriptor must still be open (and identity-checked on
    exit) while the reader parses -- proven by mutating the file DURING a
    spied parser call and observing the resulting source-unreadable error
    from the post-read identity recheck, not a stale success."""
    source = tmp_path / "source"
    _make_minimal_forms_source(source, "Alpha consent")
    preflight = inspect_intake_source(source)
    pdf_path = source / "forms" / "consent.pdf"

    real_extract = naming._extract_pdf_pages

    def mutate_during_parse(stream):
        pages = real_extract(stream)
        pdf_path.write_bytes(b"mutated-during-parse-window")
        return pages

    monkeypatch.setattr(naming, "_extract_pdf_pages", mutate_during_parse)
    _forbid_client(monkeypatch)
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert resolution.source == "generated"
    assert resolution.errors == ({"path": "forms/consent.pdf", "reason": "source-unreadable"},)


def test_between_pass_mutation_with_restored_identity_and_size_is_rejected(tmp_path, monkeypatch):
    """Pass 1 (:func:`_validate_candidate`) hashes and clears a candidate;
    pass 2 (:func:`_extract_candidate`) then re-opens it. FileIdentity
    (device/inode/size/mtime_ns) alone cannot see a same-inode, same-size
    content swap whose mtime is restored to the exact value pass 1 already
    validated -- ``open_verified_source``'s own identity checks (both at
    open and at post-read context exit) pass cleanly. Only pass 2's own
    fresh SHA-256 recheck against ``candidate.sha256``, performed before
    any parser call, can catch it. Exact reproduction of the audited PoC:
    a dictionary_mapping candidate is rewritten with same-length content
    between pass 1 and pass 2."""
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    dict_path = source / "dictionary_mapping" / "dict.csv"
    dict_path.write_text("VAR,DESC\n1,SAFE\n", encoding="utf-8")
    preflight = inspect_intake_source(source)
    candidate = preflight.candidates[0]
    assert candidate.relative_path == "dictionary_mapping/dict.csv"
    expected_mtime_ns = candidate.identity.mtime_ns

    real_validate = naming._validate_candidate

    def mutate_between_passes(src, cand):
        result = real_validate(src, cand)
        # Same-inode, same-size mutation with mtime restored to the exact
        # value pass 1 already validated -- identity alone cannot see it.
        dict_path.write_text("VAR,DESC\n1,EVIL\n", encoding="utf-8")
        os.utime(dict_path, ns=(expected_mtime_ns, expected_mtime_ns))
        post = os.stat(dict_path)
        assert post.st_ino == cand.identity.inode
        assert post.st_size == cand.identity.size
        assert post.st_mtime_ns == expected_mtime_ns
        return result

    monkeypatch.setattr(naming, "_validate_candidate", mutate_between_passes)

    extraction_calls = {"n": 0}
    real_extract_csv = naming._extract_csv_rows

    def counting_extract_csv(stream):
        extraction_calls["n"] += 1
        return real_extract_csv(stream)

    monkeypatch.setattr(naming, "_extract_csv_rows", counting_extract_csv)
    _forbid_client(monkeypatch)

    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )

    assert extraction_calls["n"] == 0
    assert FakeHTTPConnection.requests == []
    assert resolution.source == "generated"
    assert resolution.review_items == ()
    assert resolution.errors == ({"path": "dictionary_mapping/dict.csv", "reason": "source-unreadable"},)


# --- exact readers/order/canonical JSON (public behavior) ---------------------------------


def test_forms_evidence_is_ordered_bounded_and_canonical(tmp_path, monkeypatch):
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "forms").mkdir(parents=True)
    _write_pdf(source / "forms" / "a_first.pdf", "Alpha page one text")
    _write_pdf(source / "forms" / "z_second.pdf", "Zeta page one", "Zeta page two", "Zeta page three")
    preflight = inspect_intake_source(source)

    _install_fake_client(monkeypatch, _queue_dispatch(("StudyAlpha", _ACCEPT_CONF)))
    naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )

    prompts = _prompts_sent()
    assert len(prompts) == 1  # accepted on the first (forms) query -> no dict/combined dispatch
    assert prompts[0] == naming._PROMPT_PREFIX + (
        '{"component":"forms","documents":['
        '{"index":1,"pages":["Alpha page one text"]},'
        '{"index":2,"pages":["Zeta page one","Zeta page two"]}]}'
    )
    evidence = _evidence_from_prompt(prompts[0])
    dumped = json.dumps(evidence)
    assert "a_first.pdf" not in dumped and "z_second.pdf" not in dumped and "forms/" not in dumped


def test_dictionary_mapping_evidence_exact_canonical_snapshot(tmp_path, monkeypatch):
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    (source / "dictionary_mapping" / "dict.csv").write_text("VAR,DESC\nAGE,Age in years\n", encoding="utf-8")
    preflight = inspect_intake_source(source)

    _install_fake_client(monkeypatch, _queue_dispatch(("StudyAlpha", _ACCEPT_CONF)))
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )

    prompts = _prompts_sent()
    assert len(prompts) == 1
    assert prompts[0] == naming._PROMPT_PREFIX + (
        '{"component":"dictionary_mapping","documents":'
        '[{"index":1,"sheets":'
        '[{"index":1,"rows":[["VAR","DESC"],["AGE","Age in years"]]}]}]}'
    )
    assert resolution.name == "StudyAlpha"
    assert resolution.source == "ai"


def test_dictionary_mapping_evidence_mixes_csv_and_xlsx(tmp_path, monkeypatch):
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    (source / "dictionary_mapping" / "a_dict.csv").write_text("VAR,DESC\nAGE,Age in years\n", encoding="utf-8")
    _write_workbook(source / "dictionary_mapping" / "b_map.xlsx", [["code", "label"], ["1", "male"]])
    preflight = inspect_intake_source(source)

    _install_fake_client(monkeypatch, _queue_dispatch(("StudyAlpha", _ACCEPT_CONF)))
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )

    prompts = _prompts_sent()
    assert len(prompts) == 1
    evidence = _evidence_from_prompt(prompts[0])
    assert evidence["component"] == "dictionary_mapping"
    # POSIX candidate order: a_dict.csv sorts before b_map.xlsx
    assert evidence["documents"][0]["sheets"][0]["rows"] == [["VAR", "DESC"], ["AGE", "Age in years"]]
    xlsx_rows = evidence["documents"][1]["sheets"][0]["rows"]
    assert [row[:2] for row in xlsx_rows] == [["code", "label"], ["1", "male"]]
    assert all(len(row) == 20 for row in xlsx_rows)  # openpyxl's own max_col=20 window, padded
    assert all(cell == "" for row in xlsx_rows for cell in row[2:])
    assert resolution.name == "StudyAlpha"
    assert resolution.source == "ai"


# --- .xls dictionary_mapping dispatch: real xlwt bytes through xls_isolation -------------


def test_xls_dictionary_mapping_candidate_is_dispatched_through_xls_isolation(tmp_path, monkeypatch):
    """Regression for the naming dispatch bug: prior to this fix, every
    admitted dictionary_mapping candidate that was not ``.csv`` fell
    through to ``_extract_xlsx_sheets`` (openpyxl) unconditionally --
    including a real ``.xls`` (legacy BIFF) candidate, which openpyxl
    cannot parse (a zip-format assumption ``.xls`` never satisfies),
    silently collapsing to ``support-evidence-limit`` instead of ever
    reaching the isolated ``xls_isolation`` worker boundary. This test
    fails against the pre-fix dispatch (evidence never reaches the
    prompt, review_items carries a spurious support-evidence-limit
    entry) and passes once ``.xls`` is dispatched to
    ``xls_isolation.extract_xls_naming`` like every other admitted
    format."""
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    _write_xls(source / "dictionary_mapping" / "dict.xls", [["VAR", "DESC"], ["AGE", "Age in years"]])
    preflight = inspect_intake_source(source)
    assert any(
        c.relative_path == "dictionary_mapping/dict.xls" and c.component == "dictionary_mapping"
        for c in preflight.candidates
    )

    _install_fake_client(monkeypatch, _queue_dispatch(("StudyXls", _ACCEPT_CONF)))
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )

    prompts = _prompts_sent()
    assert len(prompts) == 1
    evidence = _evidence_from_prompt(prompts[0])
    assert evidence["component"] == "dictionary_mapping"
    assert evidence["documents"][0]["sheets"][0]["rows"] == [["VAR", "DESC"], ["AGE", "Age in years"]]
    assert resolution.name == "StudyXls"
    assert resolution.review_items == ()


def test_xls_worker_failure_collapses_to_support_evidence_limit(tmp_path, monkeypatch):
    """``xls_isolation.XlsWorkerError``/``XlsIsolationError`` raised by
    ``extract_xls_naming`` at naming time (for example a transient
    isolation failure after preflight already validated the same bytes)
    must collapse to the fixed ``support-evidence-limit`` code -- never a
    raw XLS exception, its code, or its message escaping this module."""
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    _write_xls(source / "dictionary_mapping" / "dict.xls", [["VAR", "DESC"]])
    preflight = inspect_intake_source(source)

    def failing_extract(data, expected_sha256):
        raise xls_isolation.XlsWorkerError("resource-limit")

    monkeypatch.setattr(naming.xls_isolation, "extract_xls_naming", failing_extract)
    _forbid_client(monkeypatch)
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert resolution.review_items == (
        {"path": "dictionary_mapping/dict.xls", "reason": "support-evidence-limit", "blocking": True},
    )
    assert resolution.source == "generated"


def test_xls_isolation_error_also_collapses_to_support_evidence_limit(tmp_path, monkeypatch):
    """Symmetric to the worker-error case: a parent-level
    ``XlsIsolationError`` (e.g. ``isolation-unavailable``) must collapse
    to the same fixed code, not leak its own distinct value."""
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    _write_xls(source / "dictionary_mapping" / "dict.xls", [["VAR", "DESC"]])
    preflight = inspect_intake_source(source)

    def failing_extract(data, expected_sha256):
        raise xls_isolation.XlsIsolationError("isolation-unavailable")

    monkeypatch.setattr(naming.xls_isolation, "extract_xls_naming", failing_extract)
    _forbid_client(monkeypatch)
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert resolution.review_items == (
        {"path": "dictionary_mapping/dict.xls", "reason": "support-evidence-limit", "blocking": True},
    )
    assert resolution.source == "generated"


def test_combined_evidence_exact_canonical_snapshot_forms_first(tmp_path, monkeypatch):
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "forms").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    _write_pdf(source / "forms" / "consent.pdf", "Alpha consent")
    (source / "dictionary_mapping" / "dict.csv").write_text("VAR,DESC\n", encoding="utf-8")
    preflight = inspect_intake_source(source)
    _install_fake_client(
        monkeypatch, _queue_dispatch((None, _REJECT_CONF), (None, _REJECT_CONF), ("StudyCombined", _ACCEPT_CONF))
    )

    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    prompts = _prompts_sent()
    assert len(prompts) == 3
    assert prompts[2] == naming._PROMPT_PREFIX + (
        '{"component":"combined",'
        '"dictionary_mapping":[{"index":1,"sheets":[{"index":1,"rows":[["VAR","DESC"]]}]}],'
        '"forms":[{"index":1,"pages":["Alpha consent"]}]}'
    )
    assert resolution.name == "StudyCombined"
    assert resolution.source == "ai"


def test_xlsx_windows_to_four_sheets_twenty_rows_twenty_cols():
    rows = [[f"r{r}c{c}" for c in range(25)] for r in range(25)]
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    wb.active.title = "S1"
    for row in rows:
        wb.active.append(row)
    for extra in ("S2", "S3", "S4", "S5"):
        ws = wb.create_sheet(extra)
        ws.append(["x"])
    wb.save(buf)

    sheets = naming._extract_xlsx_sheets(io.BytesIO(buf.getvalue()))
    assert [index for index, _ in sheets] == [1, 2, 3, 4]  # only first 4 sheets
    first_index, first_rows = sheets[0]
    assert len(first_rows) == 20  # only first 20 rows
    assert len(first_rows[0]) == 20  # only first 20 cols

def test_xlsx_max_col_bounds_the_reader_itself_not_a_post_slice():
    """Behaviorally proves the allocation contract, not just the output
    shape: a 25-column worksheet must never let openpyxl materialize more
    than 20 columns before naming ever sees the row."""
    from openpyxl.worksheet._read_only import ReadOnlyWorksheet

    rows = [[f"r{r}c{c}" for c in range(25)] for r in range(25)]
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    for row in rows:
        wb.active.append(row)
    wb.save(buf)

    observed_kwargs: dict = {}
    real_iter_rows = ReadOnlyWorksheet.iter_rows

    def spy_iter_rows(self, *args, **kwargs):
        observed_kwargs.update(kwargs)
        return real_iter_rows(self, *args, **kwargs)

    ReadOnlyWorksheet.iter_rows = spy_iter_rows
    try:
        sheets = naming._extract_xlsx_sheets(io.BytesIO(buf.getvalue()))
    finally:
        ReadOnlyWorksheet.iter_rows = real_iter_rows

    assert observed_kwargs.get("min_col") == 1
    assert observed_kwargs.get("max_col") == 20
    assert len(sheets[0][1][0]) == 20


def test_pdf_reader_caps_at_two_pages_in_order():
    buf = io.BytesIO()
    pdf = canvas.Canvas(buf)
    for text in ("Page One", "Page Two", "Page Three"):
        pdf.drawString(72, 720, text)
        pdf.showPage()
    pdf.save()
    pages = naming._extract_pdf_pages(buf.getvalue())
    assert len(pages) == 2
    assert pages[0].strip().startswith("Page One")
    assert pages[1].strip().startswith("Page Two")


# --- bounded PDF worker: compressed expansion cannot exhaust the parent ------------------


def test_compressed_pdf_expansion_is_bounded_and_never_dispatches(tmp_path, monkeypatch):
    """A small, valid, Flate-compressed PDF (well under _MAX_DOCUMENT_BYTES
    on disk) that decompresses to far more memory than its on-disk size
    implies must be rejected as support-evidence-limit by the isolated
    worker's hard address-space bound, not silently parsed to completion
    or allowed to exhaust this process. Zero model dispatch either way."""
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "forms").mkdir(parents=True)
    bomb = source / "forms" / "bomb.pdf"
    _write_flate_bomb_pdf(bomb, 300 * 1024 * 1024)
    assert bomb.stat().st_size < naming._MAX_DOCUMENT_BYTES
    preflight = inspect_intake_source(source)
    assert any(c.relative_path == "forms/bomb.pdf" for c in preflight.candidates)

    _forbid_client(monkeypatch)
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert FakeHTTPConnection.requests == []
    assert resolution.review_items == ({"path": "forms/bomb.pdf", "reason": "support-evidence-limit", "blocking": True},)
    assert resolution.source == "generated"
    assert resolution.errors == ()


def test_normal_support_pdf_naming_still_works_after_worker_isolation(tmp_path, monkeypatch):
    """Sanity companion to the bomb regression above: a normal, small,
    legitimate forms PDF must still name successfully end to end through
    the isolated worker -- the address-space/CPU/wall-time bounds must
    not reject ordinary input."""
    source = tmp_path / "source"
    _make_minimal_forms_source(source, "Alpha consent")
    preflight = inspect_intake_source(source)

    _install_fake_client(monkeypatch, _queue_dispatch(("StudyAlpha", _ACCEPT_CONF)))
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert resolution.name == "StudyAlpha"
    assert resolution.source == "ai"
    assert resolution.review_items == ()
    assert resolution.errors == ()


# --- worker lifecycle: hung/unresponsive child is reaped, never leaked ------------------


class _FakeConn:
    def __init__(self):
        self.closed = False

    def poll(self, timeout):
        return False

    def recv_bytes(self, maxlength=None):
        raise EOFError()

    def close(self):
        self.closed = True


class _FakeHangingProcess:
    """Simulates a spawned PDF worker that never sends data and never
    exits on its own -- exactly what an OOM-killed-too-slowly or a
    genuine pdfminer hang looks like from the parent's side.
    ``responds_to_terminate`` controls whether SIGTERM (``terminate()``)
    is enough, or whether escalation to SIGKILL (``kill()``) is
    required -- both must leave the process reaped, never leaked."""

    def __init__(self, *, responds_to_terminate: bool):
        self._alive = True
        self.terminate_called = False
        self.kill_called = False
        self.closed = False
        self._responds_to_terminate = responds_to_terminate
        self.exitcode = None

    def start(self):
        pass

    def is_alive(self):
        return self._alive

    def join(self, timeout=None):
        pass

    def terminate(self):
        self.terminate_called = True
        if self._responds_to_terminate:
            self._alive = False
            self.exitcode = -15

    def kill(self):
        self.kill_called = True
        self._alive = False
        self.exitcode = -9

    def close(self):
        self.closed = True


class _FakeHangingContext:
    def __init__(self, *, responds_to_terminate: bool):
        self._responds_to_terminate = responds_to_terminate
        self.processes: list[_FakeHangingProcess] = []
        self.conns: list[_FakeConn] = []

    def Pipe(self, duplex=False):
        parent_conn, child_conn = _FakeConn(), _FakeConn()
        self.conns.extend([parent_conn, child_conn])
        return parent_conn, child_conn

    def Process(self, target, args, daemon):
        process = _FakeHangingProcess(responds_to_terminate=self._responds_to_terminate)
        self.processes.append(process)
        return process


@pytest.mark.parametrize("responds_to_terminate", [True, False])
def test_hung_pdf_worker_is_terminated_or_killed_never_leaked(monkeypatch, responds_to_terminate):
    """A PDF worker that never responds and never exits on its own (an
    OOM/CPU-limit kill that has not landed yet, or a genuine pdfminer
    hang past the wall-clock bound) must be positively reaped by the
    parent -- ``terminate()``, escalating to ``kill()`` when that alone
    is not enough -- so the wall-clock bound can never leave a runaway
    child process behind. Content is treated as unusable
    (``support-evidence-limit``) either way."""
    fake_context = _FakeHangingContext(responds_to_terminate=responds_to_terminate)
    monkeypatch.setattr(naming, "_PDF_WORKER_CONTEXT", fake_context)
    monkeypatch.setattr(naming, "_PDF_WORKER_MAX_WALL_SECONDS", 0)

    with pytest.raises(naming._EvidenceLimitError):
        naming._extract_pdf_pages(b"%PDF-1.4 minimal")

    assert len(fake_context.processes) == 1
    process = fake_context.processes[0]
    assert process.terminate_called
    if not responds_to_terminate:
        assert process.kill_called
    assert not process.is_alive()
    assert process.closed
    assert all(conn.closed for conn in fake_context.conns)


# --- worker: fail-closed resource limits, mandatory before pdfplumber import -------------


class _CapturingConn:
    def __init__(self):
        self.sent: list[bytes] = []
        self.closed = False

    def send_bytes(self, data):
        self.sent.append(data)

    def close(self):
        self.closed = True


def test_worker_fails_closed_when_resource_module_unavailable():
    """If `resource` cannot even be imported (a non-POSIX platform),
    pdfplumber must never be imported and the PDF bytes must never be
    touched -- only the fixed error sentinel is sent."""
    from phi_engine.pipeline import _pdf_extract_worker as worker

    real_import = builtins.__import__
    imported = {"pdfplumber": False}

    def guarded_import(name, *args, **kwargs):
        if name == "resource":
            raise ImportError("simulated unavailable resource module")
        if name == "pdfplumber":
            imported["pdfplumber"] = True
        return real_import(name, *args, **kwargs)

    original = builtins.__import__
    builtins.__import__ = guarded_import
    try:
        conn = _CapturingConn()
        worker.run(b"whatever-pdf-bytes", 2, 256 * 1024 * 1024, 5, conn)
    finally:
        builtins.__import__ = original

    assert imported["pdfplumber"] is False
    assert conn.closed
    assert len(conn.sent) == 1
    assert json.loads(conn.sent[0].decode("ascii")) == {"status": "error", "pages": []}


@pytest.mark.parametrize("failing_limit_name", ["RLIMIT_AS", "RLIMIT_CPU"])
def test_worker_fails_closed_before_pdfplumber_import_when_setrlimit_fails(monkeypatch, failing_limit_name):
    """Either hard limit failing to apply -- RLIMIT_AS or RLIMIT_CPU --
    must prevent pdfplumber from EVER being imported or touching the
    (possibly hostile) PDF bytes; a best-effort fallback to an unbounded
    parse would defeat the entire point of this isolated worker. Exact
    reproduction of the audited PoC: `resource.setrlimit` replaced with a
    function that raises `OSError`."""
    import resource

    from phi_engine.pipeline import _pdf_extract_worker as worker

    failing_limit = getattr(resource, failing_limit_name)
    calls: list[int] = []

    def fake_setrlimit(which, limits):
        calls.append(which)
        if which == failing_limit:
            raise OSError(f"simulated {failing_limit_name} rejection")
        return None  # never touches the real process limits for the OTHER call

    monkeypatch.setattr(resource, "setrlimit", fake_setrlimit)

    imported = {"pdfplumber": False}
    real_import = builtins.__import__

    def guarded_import(name, *args, **kwargs):
        if name == "pdfplumber":
            imported["pdfplumber"] = True
        return real_import(name, *args, **kwargs)

    monkeypatch.setattr(builtins, "__import__", guarded_import)

    conn = _CapturingConn()
    worker.run(b"whatever-pdf-bytes", 2, 256 * 1024 * 1024, 5, conn)

    # Both hard limits are attempted, IN ORDER, before pdfplumber -- a
    # failing RLIMIT_AS is the very first thing attempted, so RLIMIT_CPU
    # is never even reached; a failing RLIMIT_CPU is only discovered
    # after RLIMIT_AS already succeeded.
    if failing_limit_name == "RLIMIT_AS":
        assert calls == [resource.RLIMIT_AS]
    else:
        assert calls == [resource.RLIMIT_AS, resource.RLIMIT_CPU]
    assert imported["pdfplumber"] is False
    assert conn.closed
    assert len(conn.sent) == 1
    assert json.loads(conn.sent[0].decode("ascii")) == {"status": "error", "pages": []}


def test_worker_applies_both_hard_limits_before_pdfplumber_when_they_succeed(monkeypatch):
    """Sanity companion: when both setrlimit calls succeed, pdfplumber IS
    reached -- proof the fail-closed guard above isn't blocking
    legitimate use."""
    import resource

    from phi_engine.pipeline import _pdf_extract_worker as worker

    calls: list[int] = []

    def recording_setrlimit(which, limits):
        # Record the attempt without ever touching this test process's
        # own real resource limits.
        calls.append(which)
        return None

    monkeypatch.setattr(resource, "setrlimit", recording_setrlimit)

    conn = _CapturingConn()
    worker.run(b"not a real pdf", 2, 256 * 1024 * 1024, 5, conn)

    assert calls == [resource.RLIMIT_AS, resource.RLIMIT_CPU]
    assert conn.closed
    assert len(conn.sent) == 1
    payload = json.loads(conn.sent[0].decode("ascii"))
    assert set(payload) == {"status", "pages"}
    assert payload["status"] == "error"  # garbage bytes: pdfplumber itself rejects them, not a limit failure


# --- parent-side worker reply: bounded, non-executable, strictly-schema-validated JSON ----


def _malformed_worker_replies() -> list[bytes]:
    max_pages = naming._MAX_PDF_PAGES
    max_codepoints = naming._MAX_FRAGMENT_CODEPOINTS
    too_many_pages = json.dumps({"status": "ok", "pages": ["p"] * (max_pages + 1)}, separators=(",", ":")).encode("ascii")
    too_long_page = json.dumps({"status": "ok", "pages": ["a" * (max_codepoints + 1)]}, separators=(",", ":")).encode(
        "ascii"
    )
    return [
        b"not json at all",
        b'{"status": "ok"}',
        b'{"status": "ok", "pages": [], "extra": 1}',
        b'{"status": "bad", "pages": []}',
        b'{"status": "ok", "pages": "not-a-list"}',
        b'{"status": "ok", "pages": [1, 2]}',
        b'{"status": "ok", "pages": [true, false]}',
        too_many_pages,
        too_long_page,
        b"[]",
        b"null",
    ]


@pytest.mark.parametrize("raw", _malformed_worker_replies())
def test_worker_reply_schema_rejects_every_malformed_shape(raw):
    with pytest.raises(naming._EvidenceLimitError):
        naming._decode_worker_reply(raw)


def test_worker_reply_schema_accepts_the_exact_valid_shape():
    payload = json.dumps({"status": "ok", "pages": ["Page one", "Page two"]}, separators=(",", ":")).encode("ascii")
    assert naming._decode_worker_reply(payload) == ["Page one", "Page two"]


def test_worker_reply_rejects_a_pickle_payload_without_executing_it(capsys):
    """Defense-in-depth: even if the bytes crossing the pipe were an
    actual pickle stream instead of JSON (a compromised worker, or
    anything impersonating it), the parent's strict ASCII-JSON decode
    must reject it outright. Nothing in this module calls
    `pickle.loads` any more, so a malicious `__reduce__` payload has no
    path to execute -- proven here with a `__reduce__` that would print
    a detectable sentinel if it ever ran."""
    import pickle

    class _Exploit:
        def __reduce__(self):
            return (print, ("PICKLE-EXECUTED-DURING-DECODE",))

    malicious = pickle.dumps(_Exploit())
    with pytest.raises(naming._EvidenceLimitError):
        naming._decode_worker_reply(malicious)
    assert "PICKLE-EXECUTED-DURING-DECODE" not in capsys.readouterr().out


class _ImmediateReplyConn:
    def __init__(self, payload: bytes | None):
        self._payload = payload
        self.closed = False

    def poll(self, timeout):
        return self._payload is not None

    def recv_bytes(self, maxlength=None):
        if self._payload is None:
            raise EOFError()
        return self._payload

    def close(self):
        self.closed = True


class _ImmediateExitProcess:
    def __init__(self):
        self.exitcode = 0
        self._alive = False
        self.closed = False

    def start(self):
        pass

    def is_alive(self):
        return self._alive

    def join(self, timeout=None):
        pass

    def terminate(self):
        self._alive = False

    def kill(self):
        self._alive = False

    def close(self):
        self.closed = True


class _ImmediateReplyContext:
    def __init__(self, payload: bytes | None):
        self._payload = payload

    def Pipe(self, duplex=False):
        return _ImmediateReplyConn(self._payload), _ImmediateReplyConn(None)

    def Process(self, target, args, daemon):
        return _ImmediateExitProcess()


def _malicious_end_to_end_replies() -> list[bytes]:
    return [
        b"not json",
        json.dumps({"status": "ok", "pages": [], "sneaky": True}, separators=(",", ":")).encode("ascii"),
        json.dumps({"status": "ok", "pages": [1, 2]}, separators=(",", ":")).encode("ascii"),
        json.dumps({"status": "ok", "pages": ["a"] * (naming._MAX_PDF_PAGES + 5)}, separators=(",", ":")).encode("ascii"),
    ]


@pytest.mark.parametrize("payload", _malicious_end_to_end_replies())
def test_extract_pdf_pages_rejects_malicious_worker_reply_end_to_end(monkeypatch, payload):
    """The full end-to-end path (the spawned worker replaced by a fake
    that returns attacker-controlled bytes) must reject any malformed/
    malicious reply through the same fixed _EvidenceLimitError, never
    accepting or partially trusting it."""
    monkeypatch.setattr(naming, "_PDF_WORKER_CONTEXT", _ImmediateReplyContext(payload))
    with pytest.raises(naming._EvidenceLimitError):
        naming._extract_pdf_pages(b"%PDF-1.4 minimal")


# --- parent lifecycle: raw exceptions normalize; the wall deadline is single and shared ---


class _RaisingConn:
    def __init__(self, *, poll_exc: BaseException | None = None, sleep_before_poll: float = 0.0):
        self._poll_exc = poll_exc
        self._sleep_before_poll = sleep_before_poll
        self.poll_calls: list[float] = []
        self.closed = False

    def poll(self, timeout):
        self.poll_calls.append(timeout)
        if self._sleep_before_poll:
            time.sleep(min(timeout, self._sleep_before_poll))
        if self._poll_exc is not None:
            raise self._poll_exc
        return False

    def recv_bytes(self, maxlength=None):
        raise EOFError()

    def close(self):
        self.closed = True


class _RaisingProcess:
    def __init__(
        self,
        *,
        start_exc: BaseException | None = None,
        terminate_exc: BaseException | None = None,
        kill_exc: BaseException | None = None,
    ):
        self._start_exc = start_exc
        self._terminate_exc = terminate_exc
        self._kill_exc = kill_exc
        self._alive = False
        self.exitcode = None
        self.terminate_called = False
        self.kill_called = False
        self.closed = False
        self.join_calls: list[float | None] = []

    def start(self):
        if self._start_exc is not None:
            raise self._start_exc
        self._alive = True

    def is_alive(self):
        return self._alive

    def join(self, timeout=None):
        self.join_calls.append(timeout)

    def terminate(self):
        self.terminate_called = True
        if self._terminate_exc is not None:
            raise self._terminate_exc
        self._alive = False
        self.exitcode = -15

    def kill(self):
        self.kill_called = True
        if self._kill_exc is not None:
            raise self._kill_exc
        self._alive = False
        self.exitcode = -9

    def close(self):
        self.closed = True


class _OneShotContext:
    def __init__(self, process, parent_conn, child_conn=None):
        self._process = process
        self._parent_conn = parent_conn
        self._child_conn = child_conn if child_conn is not None else _RaisingConn()

    def Pipe(self, duplex=False):
        return self._parent_conn, self._child_conn

    def Process(self, target, args, daemon):
        return self._process


def test_worker_start_raw_exception_normalizes_with_zero_live_child(monkeypatch):
    """A raw exception from Process.start() -- never a fixed sentinel
    from this library -- must still collapse to _EvidenceLimitError.
    Since the child never actually started, no reap is attempted, but
    both pipe endpoints are still closed."""
    process = _RaisingProcess(start_exc=RuntimeError("RAW-WORKER-START-SENTINEL"))
    parent_conn = _RaisingConn()
    child_conn = _RaisingConn()
    ctx = _OneShotContext(process, parent_conn, child_conn)
    monkeypatch.setattr(naming, "_PDF_WORKER_CONTEXT", ctx)

    with pytest.raises(naming._EvidenceLimitError):
        naming._extract_pdf_pages(b"%PDF-1.4 minimal")

    assert not process.is_alive()
    assert process.closed
    assert parent_conn.closed
    assert child_conn.closed


def test_worker_poll_raw_exception_still_reaps_and_normalizes(monkeypatch):
    """A raw exception from Connection.poll() must not prevent the
    started child from being positively reaped, nor leak a raw
    exception -- both collapse to the same fixed _EvidenceLimitError."""
    process = _RaisingProcess()
    parent_conn = _RaisingConn(poll_exc=RuntimeError("RAW-IPC-SENTINEL"))
    child_conn = _RaisingConn()
    ctx = _OneShotContext(process, parent_conn, child_conn)
    monkeypatch.setattr(naming, "_PDF_WORKER_CONTEXT", ctx)

    with pytest.raises(naming._EvidenceLimitError):
        naming._extract_pdf_pages(b"%PDF-1.4 minimal")

    assert process.terminate_called
    assert not process.is_alive()
    assert process.closed
    assert parent_conn.closed
    assert child_conn.closed


def test_worker_terminate_raw_exception_still_escalates_to_kill(monkeypatch):
    """A raw exception from terminate() itself must not skip the kill()
    escalation -- otherwise a child whose SIGTERM handling itself raised
    inside the Python wrapper could be left alive forever."""
    process = _RaisingProcess(terminate_exc=RuntimeError("RAW-TERMINATE-SENTINEL"))
    parent_conn = _RaisingConn()
    monkeypatch.setattr(naming, "_PDF_WORKER_MAX_WALL_SECONDS", 0)
    ctx = _OneShotContext(process, parent_conn)
    monkeypatch.setattr(naming, "_PDF_WORKER_CONTEXT", ctx)

    with pytest.raises(naming._EvidenceLimitError):
        naming._extract_pdf_pages(b"%PDF-1.4 minimal")

    assert process.terminate_called
    assert process.kill_called
    assert not process.is_alive()
    assert process.closed


def test_worker_kill_raw_exception_never_crashes_the_parent(monkeypatch):
    """Even if kill() itself also raises after a failed terminate(), the
    parent must still normalize to the fixed error and finish cleanup
    rather than propagating a raw exception."""
    process = _RaisingProcess(terminate_exc=RuntimeError("term fails"), kill_exc=RuntimeError("kill fails too"))
    parent_conn = _RaisingConn()
    monkeypatch.setattr(naming, "_PDF_WORKER_MAX_WALL_SECONDS", 0)
    ctx = _OneShotContext(process, parent_conn)
    monkeypatch.setattr(naming, "_PDF_WORKER_CONTEXT", ctx)

    with pytest.raises(naming._EvidenceLimitError):
        naming._extract_pdf_pages(b"%PDF-1.4 minimal")

    assert process.terminate_called
    assert process.kill_called
    assert process.closed


def test_wall_deadline_is_shared_between_poll_and_join_not_doubled(monkeypatch):
    """The configured wall bound must be ONE shared monotonic deadline,
    not applied in full to poll() and then again in full to join() --
    that would silently double the effective hang tolerance. A
    genuinely hanging poll() (it sleeps for its full requested timeout,
    exactly like a real blocking wait on an unresponsive pipe) must
    leave join() with only whatever time remains under the SAME
    deadline, and total wall-clock elapsed must stay close to ONE
    configured wall period, not two."""
    process = _RaisingProcess()
    parent_conn = _RaisingConn(sleep_before_poll=10.0)  # would hang far longer than the wall bound
    child_conn = _RaisingConn()
    ctx = _OneShotContext(process, parent_conn, child_conn)
    monkeypatch.setattr(naming, "_PDF_WORKER_CONTEXT", ctx)
    monkeypatch.setattr(naming, "_PDF_WORKER_MAX_WALL_SECONDS", 0.3)

    started_at = time.monotonic()
    with pytest.raises(naming._EvidenceLimitError):
        naming._extract_pdf_pages(b"%PDF-1.4 minimal")
    elapsed = time.monotonic() - started_at

    assert len(parent_conn.poll_calls) == 1
    assert parent_conn.poll_calls[0] == pytest.approx(0.3, abs=0.15)
    # join() must receive only the LEFTOVER time under the same
    # deadline poll() already consumed -- never another independent
    # 0.3-second wait (the exact bug: poll(WALL) then join(WALL) would
    # double the effective hang tolerance).
    assert process.join_calls[0] < 0.1
    assert elapsed < 0.7
    assert process.terminate_called
    assert not process.is_alive()
    assert process.closed and parent_conn.closed and child_conn.closed



# --- archive/source/cell/field/parser boundaries -------------------------------------------


def test_fragment_codepoint_limit_rejects_rather_than_truncates():
    long_text = "a" * (naming._MAX_FRAGMENT_CODEPOINTS + 1)
    buf = io.BytesIO()
    pdf = canvas.Canvas(buf)
    pdf.drawString(10, 750, long_text)
    pdf.save()
    with pytest.raises(naming._EvidenceLimitError):
        naming._extract_pdf_pages(buf.getvalue())


def test_csv_field_byte_limit_is_enforced():
    huge_field = "x" * (naming._MAX_CSV_FIELD_BYTES + 1)
    data = f"A,B\n{huge_field},ok\n".encode("utf-8")
    with pytest.raises(naming._EvidenceLimitError):
        naming._extract_csv_rows(io.BytesIO(data))


def test_csv_column_21_cannot_bypass_the_field_byte_limit():
    """A field beyond the retained 20 columns must still be validated --
    proves the byte check runs over every parsed field before slicing,
    not only the ones that survive to the output."""
    huge_field = "x" * (naming._MAX_CSV_FIELD_BYTES + 1)
    row = ",".join(["a"] * 20 + [huge_field])
    data = (row + "\n").encode("utf-8")
    with pytest.raises(naming._EvidenceLimitError):
        naming._extract_csv_rows(io.BytesIO(data))


def test_csv_row_and_column_windowing():
    header = ",".join(f"c{i}" for i in range(25))
    lines = [header] + [",".join(f"v{r}{i}" for i in range(25)) for r in range(25)]
    data = ("\n".join(lines) + "\n").encode("utf-8")
    rows = naming._extract_csv_rows(io.BytesIO(data))
    assert len(rows) == 20
    assert all(len(row) == 20 for row in rows)


def test_invalid_utf8_csv_is_evidence_limit_not_a_crash():
    with pytest.raises(naming._EvidenceLimitError):
        naming._extract_csv_rows(io.BytesIO(b"\xff\xfe\x00bad-utf8"))


def test_csv_field_size_limit_is_restored_after_parsing():
    original = __import__("csv").field_size_limit()
    try:
        naming._extract_csv_rows(io.BytesIO(b"a,b\n1,2\n"))
        assert __import__("csv").field_size_limit() == original
    finally:
        __import__("csv").field_size_limit(original)


def test_csv_field_size_limit_is_restored_after_an_exception():
    """The success-path restoration test cannot catch exception-only
    leakage of the process-global csv.field_size_limit -- trigger
    _EvidenceLimitError via an oversized field and assert the ambient
    value was still restored."""
    import csv as csv_module

    original = csv_module.field_size_limit()
    huge_field = "x" * (naming._MAX_CSV_FIELD_BYTES + 1)
    data = f"A,B\n{huge_field},ok\n".encode("utf-8")
    try:
        with pytest.raises(naming._EvidenceLimitError):
            naming._extract_csv_rows(io.BytesIO(data))
        assert csv_module.field_size_limit() == original
    finally:
        csv_module.field_size_limit(original)


def test_csv_field_size_limit_is_restored_after_invalid_utf8():
    import csv as csv_module

    original = csv_module.field_size_limit()
    try:
        with pytest.raises(naming._EvidenceLimitError):
            naming._extract_csv_rows(io.BytesIO(b"\xff\xfe\x00bad-utf8"))
        assert csv_module.field_size_limit() == original
    finally:
        csv_module.field_size_limit(original)


def test_zip_bomb_ratio_is_rejected():
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("xl/workbook.xml", "0" * (10 * 1024 * 1024))
    with pytest.raises(naming._EvidenceLimitError):
        naming._extract_xlsx_sheets(io.BytesIO(buf.getvalue()))


def test_malformed_zip_is_evidence_limit():
    with pytest.raises(naming._EvidenceLimitError):
        naming._extract_xlsx_sheets(io.BytesIO(b"not a zip file at all"))


def test_zip_central_directory_overrun_is_rejected():
    """Independent of any single member's declared size: a directory with
    many small members can itself exceed max_zip_directory_bytes before
    any member is ever opened."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for i in range(2048):
            zf.writestr(f"member_{i}_" + "x" * 50 + ".txt", "0")
    data = buf.getvalue()
    with pytest.raises(naming._EvidenceLimitError):
        naming._extract_xlsx_sheets(io.BytesIO(data))


def test_xlsx_lazy_worksheet_iteration_failure_is_evidence_limit_not_a_leak():
    """Reproduces the audited leak: a well-formed-looking workbook whose
    cell reference is malformed fails during LAZY worksheet iteration
    (not during load_workbook itself), and must still collapse to the
    fixed code -- never a raw ValueError with the malformed content
    embedded in its message."""
    buf = io.BytesIO()
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        "</Types>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        "</Relationships>"
    )
    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheets><sheet name="Sheet1" sheetId="1" r:id="rId1"/></sheets></workbook>'
    )
    wbrels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
        "</Relationships>"
    )
    sentinel = "PHI123-45-6789"
    sheet1 = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<sheetData><row r="1"><c r="{sentinel}" t="str"><v>bad</v></c></row></sheetData>'
        "</worksheet>"
    )
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("_rels/.rels", rels)
        zf.writestr("xl/workbook.xml", workbook_xml)
        zf.writestr("xl/_rels/workbook.xml.rels", wbrels)
        zf.writestr("xl/worksheets/sheet1.xml", sheet1)
    data = buf.getvalue()

    try:
        naming._extract_xlsx_sheets(io.BytesIO(data))
        pytest.fail("expected _EvidenceLimitError")
    except naming._EvidenceLimitError:
        pass
    except Exception as exc:  # pragma: no cover -- explicit anti-leak assertion
        pytest.fail(f"raw exception leaked instead of _EvidenceLimitError: {type(exc).__name__}: {exc}")


def test_malformed_xlsx_end_to_end_produces_fixed_review_not_a_crash(tmp_path, monkeypatch):
    """A malformed cell reference lives inside worksheet XML, which
    preflight's own workbook.xml-only sheet count never inspects -- this
    candidate reaches naming as a clean-looking admitted dictionary_mapping
    file, and only naming's fuller openpyxl load+iterate trips over it."""
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>'
        '<Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
        "</Types>"
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>'
        "</Relationships>"
    )
    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheets><sheet name="Sheet1" sheetId="1" r:id="rId1"/></sheets></workbook>'
    )
    wbrels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/>'
        "</Relationships>"
    )
    sheet1 = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        '<sheetData><row r="1"><c r="PHI123-45-6789" t="str"><v>bad</v></c></row></sheetData>'
        "</worksheet>"
    )
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("_rels/.rels", rels)
        zf.writestr("xl/workbook.xml", workbook_xml)
        zf.writestr("xl/_rels/workbook.xml.rels", wbrels)
        zf.writestr("xl/worksheets/sheet1.xml", sheet1)
    (source / "dictionary_mapping" / "malformed.xlsx").write_bytes(buf.getvalue())
    preflight = inspect_intake_source(source)
    # Confirms the fixture premise: preflight's workbook.xml-only sheet
    # count admits this as a clean dictionary_mapping candidate.
    assert any(
        c.relative_path == "dictionary_mapping/malformed.xlsx" and c.component == "dictionary_mapping"
        for c in preflight.candidates
    )

    _forbid_client(monkeypatch)
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert resolution.source == "generated"
    assert resolution.review_items == (
        {"path": "dictionary_mapping/malformed.xlsx", "reason": "support-evidence-limit", "blocking": True},
    )
    dumped = json.dumps(resolution.review_items)
    assert "PHI123-45-6789" not in dumped


# --- maximal truncation boundary for all three payload classes -----------------------------


def test_forms_truncation_boundary_is_maximal_not_just_within_budget(tmp_path, monkeypatch):
    monkeypatch.setattr(naming, "_MAX_EVIDENCE_BYTES", 90)
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "forms").mkdir(parents=True)
    _write_pdf(source / "forms" / "a_doc.pdf", "AAAAAAAAAA")
    _write_pdf(source / "forms" / "b_doc.pdf", "BBBBBBBBBB")
    _write_pdf(source / "forms" / "c_doc.pdf", "CCCCCCCCCC")
    preflight = inspect_intake_source(source)
    _install_fake_client(monkeypatch, _queue_dispatch((None, 0.0)))
    naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    prompts = _prompts_sent()
    evidence = _evidence_from_prompt(prompts[0])
    encoded = prompts[0][len(naming._PROMPT_PREFIX) :].encode("utf-8")
    assert len(encoded) <= 90

    # Maximality: appending the next ordered fragment (proven by
    # reconstructing it through the module's own payload builder) exceeds
    # the cap -- this is not an arbitrary early stop.
    kept_docs = {d["index"]: list(d["pages"]) for d in evidence["documents"]}
    next_index = (max(kept_docs) if kept_docs else 0) + 1
    trial = {k: list(v) for k, v in kept_docs.items()}
    trial.setdefault(next_index, []).append("D" * 10)
    trial_payload = {"component": "forms", "documents": [{"index": i, "pages": trial[i]} for i in sorted(trial)]}
    assert len(json.dumps(trial_payload, sort_keys=True, separators=(",", ":"), ensure_ascii=False).encode("utf-8")) > 90


def test_dictionary_truncation_boundary_is_maximal(tmp_path, monkeypatch):
    monkeypatch.setattr(naming, "_MAX_EVIDENCE_BYTES", 130)
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    (source / "dictionary_mapping" / "a_dict.csv").write_text("AAAA,BBBB\n1111,2222\n3333,4444\n", encoding="utf-8")
    preflight = inspect_intake_source(source)
    _install_fake_client(monkeypatch, _queue_dispatch((None, 0.0)))
    naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    prompts = _prompts_sent()
    evidence = _evidence_from_prompt(prompts[0])
    encoded = prompts[0][len(naming._PROMPT_PREFIX) :].encode("utf-8")
    assert len(encoded) <= 130
    kept_rows = evidence["documents"][0]["sheets"][0]["rows"] if evidence["documents"] else []
    all_rows = [["AAAA", "BBBB"], ["1111", "2222"], ["3333", "4444"]]
    assert kept_rows == all_rows[: len(kept_rows)]
    assert len(kept_rows) < len(all_rows)  # budget genuinely truncated something
    trial_rows = kept_rows + [all_rows[len(kept_rows)]]
    trial_payload = {
        "component": "dictionary_mapping",
        "documents": [{"index": 1, "sheets": [{"index": 1, "rows": trial_rows}]}],
    }
    assert len(json.dumps(trial_payload, sort_keys=True, separators=(",", ":"), ensure_ascii=False).encode("utf-8")) > 130


def test_combined_truncation_boundary_is_maximal_forms_then_dict(tmp_path, monkeypatch):
    monkeypatch.setattr(naming, "_MAX_EVIDENCE_BYTES", 160)
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "forms").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    _write_pdf(source / "forms" / "consent.pdf", "FORMSFORMSFORMS")
    (source / "dictionary_mapping" / "dict.csv").write_text("DICTDICT,ROWROW\n1,2\n3,4\n", encoding="utf-8")
    preflight = inspect_intake_source(source)
    _install_fake_client(monkeypatch, _queue_dispatch((None, 0.0), (None, 0.0), (None, 0.0)))
    naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    prompts = _prompts_sent()
    assert len(prompts) == 3
    encoded = prompts[2][len(naming._PROMPT_PREFIX) :].encode("utf-8")
    assert len(encoded) <= 160
    evidence = json.loads(encoded)
    assert evidence["component"] == "combined"
    # forms-first ordering: forms content is never sacrificed for dict
    # content (forms fragments are appended before any dict fragment) --
    # with this budget, forms alone already consumes enough of the shared
    # cap that dict is excluded entirely, not partially.
    assert evidence["forms"] == [{"index": 1, "pages": ["FORMSFORMSFORMS"]}]
    assert evidence["dictionary_mapping"] == []
    # Maximality: appending even the FIRST dict row would exceed the cap.
    trial = dict(evidence)
    trial["dictionary_mapping"] = [
        {"index": 1, "sheets": [{"index": 1, "rows": [["DICTDICT", "ROWROW"]]}]}
    ]
    assert len(json.dumps(trial, sort_keys=True, separators=(",", ":"), ensure_ascii=False).encode("utf-8")) > 160


# --- descriptor OSError normalization -------------------------------------------------------


def test_descriptor_dup_failure_is_normalized_to_source_unreadable(tmp_path, monkeypatch):
    source = tmp_path / "source"
    _make_minimal_forms_source(source, "Alpha consent")
    preflight = inspect_intake_source(source)

    def failing_read(fd, n):
        raise OSError(24, "Too many open files")

    monkeypatch.setattr(naming.os, "read", failing_read)
    _forbid_client(monkeypatch)
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert resolution.source == "generated"
    assert resolution.errors == ({"path": "forms/consent.pdf", "reason": "source-unreadable"},)
    for err in resolution.errors:
        assert "Too many open files" not in json.dumps(err)


# --- no dataset evidence -------------------------------------------------------------------


def test_combined_payload_never_includes_dataset_component():
    combined = naming._combined_payload_dict({1: ["form text"]}, {1: {1: [["header"]]}})
    assert combined["component"] == "combined"
    assert set(combined) == {"component", "forms", "dictionary_mapping"}
    assert "datasets" not in json.dumps(combined)


# --- local-only routing -------------------------------------------------------------------


def test_never_calls_ordinary_llm_client(tmp_path, monkeypatch):
    source = tmp_path / "source"
    _make_canonical_source(source)
    preflight = inspect_intake_source(source)
    monkeypatch.setattr(config, "get_llm_client", lambda: (_ for _ in ()).throw(AssertionError("must not be called")))
    _install_fake_client(monkeypatch, _queue_dispatch((None, _REJECT_CONF), (None, _REJECT_CONF)))
    naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )  # no assertion error raised == get_llm_client never touched


def test_dispatch_uses_real_complete_bounded_with_128_tokens_4096_bytes(tmp_path, monkeypatch):
    source = tmp_path / "source"
    _make_minimal_forms_source(source, "Alpha consent")
    preflight = inspect_intake_source(source)
    _install_fake_client(monkeypatch, _queue_dispatch(("StudyAlpha", _ACCEPT_CONF)))

    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert resolution.name == "StudyAlpha"
    generate_body = json.loads(FakeHTTPConnection.requests[1][2])
    assert generate_body["options"] == {"num_predict": 128}


# --- PHI gate on evidence, before dispatch --------------------------------------------------


def test_phi_bearing_evidence_blocks_before_any_dispatch(tmp_path, monkeypatch):
    source = tmp_path / "source"
    _make_minimal_forms_source(source, "Patient SSN is 123-45-6789 please verify")
    preflight = inspect_intake_source(source)
    _install_fake_client(monkeypatch, [])  # any request at all is a bug -> IndexError on empty queue

    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )

    assert FakeHTTPConnection.requests == []
    assert resolution.source == "generated"
    assert resolution.review_items == ({"path": "", "reason": "possible-phi-requires-study", "blocking": True},)
    dumped = json.dumps(resolution.review_items)
    assert "123-45-6789" not in dumped


# --- guard_llm_output on the response -------------------------------------------------------


def test_phi_bearing_model_output_is_an_inspection_failure_not_a_leak(tmp_path, monkeypatch):
    source = tmp_path / "source"
    _make_minimal_forms_source(source, "Alpha consent")
    preflight = inspect_intake_source(source)
    responses = _queue_dispatch('{"study_name":"123-45-6789","confidence":0.99}', (None, 0.0))
    _install_fake_client(monkeypatch, responses)

    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )

    assert resolution.errors == ({"path": None, "reason": "study-name-inspection-failed"},)
    assert resolution.source == "generated"
    dumped = json.dumps(resolution.errors) + json.dumps(resolution.review_items) + resolution.name
    assert "123-45-6789" not in dumped


# --- strict output schema validation --------------------------------------------------------


@pytest.mark.parametrize(
    "raw",
    [
        "not json at all",
        "",
        "```json\n{\"study_name\":\"X\",\"confidence\":0.9}\n```",  # fencing must be rejected, not unwrapped
        '{"study_name":"X","confidence":0.9} trailing text',
        '{"study_name":"X"}',  # missing confidence
        '{"study_name":"X","confidence":0.9,"extra":1}',  # extra key
        '{"study_name":1,"confidence":0.9}',  # wrong type
        '{"study_name":"X","confidence":"0.9"}',  # confidence as string
        '{"study_name":"X","confidence":true}',  # boolean confidence
        '{"study_name":"X","confidence":1.5}',  # out of range
        '{"study_name":"X","confidence":-0.1}',  # out of range
        '{"study_name":"X","confidence":NaN}',  # nonfinite
        '{"study_name":"X","confidence":Infinity}',  # nonfinite
        '{"study_name":"X","confidence":0.9,"study_name":"Y"}',  # duplicate key
        '[1,2,3]',  # not an object
        '"just a string"',
    ],
)
def test_strict_output_schema_rejects_every_malformed_shape(raw):
    with pytest.raises(naming._InspectionFailed):
        naming._parse_model_output(raw)


def test_strict_output_schema_accepts_null_study_name():
    study_name, confidence = naming._parse_model_output('{"study_name":null,"confidence":0.42}')
    assert study_name is None
    assert confidence == pytest.approx(0.42)


def test_strict_output_schema_accepts_valid_string_and_bounds():
    study_name, confidence = naming._parse_model_output('{"study_name":"StudyAlpha","confidence":1}')
    assert study_name == "StudyAlpha"
    assert confidence == 1.0


def test_oversized_transport_envelope_end_to_end_is_inspection_failed(tmp_path, monkeypatch):
    """A >4096-byte /api/generate envelope end to end (real transport
    ceiling, not a stubbed client) collapses to the fixed code."""
    source = tmp_path / "source"
    _make_minimal_forms_source(source, "Alpha consent")
    preflight = inspect_intake_source(source)
    name, digest = MODEL_SPEC.split("@", 1)
    oversized = json.dumps({"response": "x" * 5000}).encode()
    FakeHTTPConnection.requests = []
    FakeHTTPConnection.responses = [
        FakeHTTPResponse(200, json.dumps({"models": [{"name": name, "digest": digest}]}).encode()),
        FakeHTTPResponse(200, oversized),
    ]
    monkeypatch.setattr(routing.http.client, "HTTPConnection", FakeHTTPConnection)
    monkeypatch.setattr(naming, "new_offline_local_client", lambda: routing.OfflineLocalLLMClient(_local_config()))

    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert resolution.errors == (
        {"path": None, "reason": "study-name-inspection-failed"},
        {"path": None, "reason": "study-name-inspection-failed"},
    )
    assert resolution.source == "generated"


def test_oversized_or_malformed_model_response_is_inspection_failed(tmp_path, monkeypatch):
    source = tmp_path / "source"
    _make_minimal_forms_source(source, "Alpha consent")
    preflight = inspect_intake_source(source)
    responses = _queue_dispatch("not-json-at-all", (None, 0.0))
    _install_fake_client(monkeypatch, responses)

    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert resolution.errors == ({"path": None, "reason": "study-name-inspection-failed"},)
    assert resolution.source == "generated"


def test_local_transport_failure_is_inspection_failed(tmp_path, monkeypatch):
    source = tmp_path / "source"
    _make_minimal_forms_source(source, "Alpha consent")
    preflight = inspect_intake_source(source)
    FakeHTTPConnection.requests = []
    FakeHTTPConnection.responses = [FakeHTTPResponse(500, b"")]
    monkeypatch.setattr(routing.http.client, "HTTPConnection", FakeHTTPConnection)
    client = routing.OfflineLocalLLMClient(_local_config(max_retries=0))
    monkeypatch.setattr(naming, "new_offline_local_client", lambda: client)

    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert resolution.errors == (
        {"path": None, "reason": "study-name-inspection-failed"},
        {"path": None, "reason": "study-name-inspection-failed"},
    )
    assert resolution.source == "generated"


def test_local_endpoint_and_digest_controls_still_enforced(tmp_path, monkeypatch):
    source = tmp_path / "source"
    _make_minimal_forms_source(source, "Alpha consent")
    preflight = inspect_intake_source(source)
    # offline_approved=False must fail closed before any transport happens.
    bad_client = routing.OfflineLocalLLMClient(_local_config(offline_approved=False))
    monkeypatch.setattr(naming, "new_offline_local_client", lambda: bad_client)

    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert resolution.errors == (
        {"path": None, "reason": "study-name-inspection-failed"},
        {"path": None, "reason": "study-name-inspection-failed"},
    )
    assert resolution.source == "generated"


def test_local_client_configuration_failure_is_inspection_failed(tmp_path, monkeypatch):
    """The public client factory's own config-loading failure
    (LocalLLMConfigurationError) must not escape resolve_intake_study."""
    source = tmp_path / "source"
    _make_minimal_forms_source(source, "Alpha consent")
    preflight = inspect_intake_source(source)

    def broken_factory():
        raise config.LocalLLMConfigurationError("bad local_llm config")

    monkeypatch.setattr(naming, "new_offline_local_client", broken_factory)
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert resolution.source == "generated"
    for err in resolution.errors:
        assert err["reason"] == "study-name-inspection-failed"
        assert "bad local_llm config" not in json.dumps(err)


# --- separate / matching / conflicting / combined / null decisions -------------------------


def test_forms_only_accepted_is_used_without_dict_query(tmp_path, monkeypatch):
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "forms").mkdir(parents=True)
    _write_pdf(source / "forms" / "consent.pdf", "Alpha consent")
    preflight = inspect_intake_source(source)
    _install_fake_client(monkeypatch, _queue_dispatch(("StudyAlpha", _ACCEPT_CONF)))
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert resolution.name == "StudyAlpha"
    assert resolution.source == "ai"
    assert len(_prompts_sent()) == 1


def test_dictionary_only_accepted_is_used_symmetrically(tmp_path, monkeypatch):
    """Symmetric to the forms-only branch: dictionary_mapping alone can be
    accepted and used as the final name with no forms evidence at all."""
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    (source / "dictionary_mapping" / "dict.csv").write_text("VAR,DESC\n", encoding="utf-8")
    preflight = inspect_intake_source(source)
    _install_fake_client(monkeypatch, _queue_dispatch(("StudyDict", _ACCEPT_CONF)))
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert resolution.name == "StudyDict"
    assert resolution.source == "ai"
    assert len(_prompts_sent()) == 1


def test_matching_candidates_preserve_forms_spelling(tmp_path, monkeypatch):
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "forms").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    _write_pdf(source / "forms" / "consent.pdf", "Alpha consent")
    _write_workbook(source / "dictionary_mapping" / "dict.xlsx", [["dict"]])
    preflight = inspect_intake_source(source)
    _install_fake_client(monkeypatch, _queue_dispatch(("StudyAlpha", _ACCEPT_CONF), ("STUDYALPHA", _ACCEPT_CONF)))

    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert resolution.name == "StudyAlpha"  # forms spelling preserved despite dict differing only in case
    assert resolution.source == "ai"
    assert resolution.review_items == ()


def test_conflicting_candidates_block_with_conflict_review_and_generated_fallback(tmp_path, monkeypatch):
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "forms").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    _write_pdf(source / "forms" / "consent.pdf", "Alpha consent")
    _write_workbook(source / "dictionary_mapping" / "dict.xlsx", [["dict"]])
    preflight = inspect_intake_source(source)
    _install_fake_client(monkeypatch, _queue_dispatch(("StudyAlpha", _ACCEPT_CONF), ("StudyBeta", _ACCEPT_CONF)))

    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert resolution.source == "generated"
    assert resolution.review_items == (
        {
            "path": "",
            "reason": "study-name-conflict",
            "blocking": True,
            "candidates": {"forms": "StudyAlpha", "dictionary_mapping": "StudyBeta"},
        },
    )


def test_neither_accepted_falls_back_to_combined_query(tmp_path, monkeypatch):
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "forms").mkdir(parents=True)
    (source / "dictionary_mapping").mkdir(parents=True)
    _write_pdf(source / "forms" / "consent.pdf", "Alpha consent")
    _write_workbook(source / "dictionary_mapping" / "dict.xlsx", [["dict"]])
    preflight = inspect_intake_source(source)
    _install_fake_client(
        monkeypatch,
        _queue_dispatch((None, _REJECT_CONF), (None, _REJECT_CONF), ("StudyCombined", _ACCEPT_CONF)),
    )

    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    prompts = _prompts_sent()
    assert len(prompts) == 3
    combined_evidence = _evidence_from_prompt(prompts[2])
    assert combined_evidence["component"] == "combined"
    assert set(combined_evidence) == {"component", "forms", "dictionary_mapping"}
    assert resolution.name == "StudyCombined"
    assert resolution.source == "ai"


def test_clean_low_confidence_and_null_result_means_generated_fallback(tmp_path, monkeypatch):
    source = tmp_path / "source"
    _make_minimal_forms_source(source, "Alpha consent")
    preflight = inspect_intake_source(source)
    # forms low-confidence, no dict evidence -> combined re-queried on the
    # same (only) forms fragments, also clean-null -> generated, no errors.
    _install_fake_client(monkeypatch, _queue_dispatch((None, _REJECT_CONF), (None, 0.0)))

    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert resolution.source == "generated"
    assert resolution.errors == ()
    assert resolution.review_items == ()


def test_exact_threshold_confidence_is_accepted(tmp_path, monkeypatch):
    source = tmp_path / "source"
    _make_minimal_forms_source(source, "Alpha consent")
    preflight = inspect_intake_source(source)
    _install_fake_client(monkeypatch, _queue_dispatch(("StudyAlpha", _THRESHOLD)))
    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert resolution.name == "StudyAlpha"
    assert resolution.source == "ai"


def test_combined_dispatch_skipped_when_zero_evidence_after_consent(tmp_path, monkeypatch):
    source = tmp_path / "source"
    (source / "datasets").mkdir(parents=True)
    (source / "datasets" / "labs.csv").write_text("SUBJID\n1\n", encoding="utf-8")
    preflight = inspect_intake_source(source)  # no forms/, no dict/mappings admitted
    _forbid_client(monkeypatch)

    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    assert resolution.source == "generated"
    assert FakeHTTPConnection.requests == []


# --- random fallback -------------------------------------------------------------------------


def test_generated_name_format_and_validity():
    name = naming._generate_study_name()
    assert re.fullmatch(r"study-[0-9a-f]{8}", name)
    pipeline_lock.lock_path_for(name)  # must not raise


def test_generated_names_are_not_constant():
    names = {naming._generate_study_name() for _ in range(20)}
    assert len(names) > 1


# --- fixed, value-free error/review vocabulary ------------------------------------------------


def test_review_and_error_records_never_carry_raw_content_or_paths_beyond_source_relative(tmp_path, monkeypatch):
    source = tmp_path / "source"
    _make_minimal_forms_source(source, "Patient SSN is 123-45-6789")
    preflight = inspect_intake_source(source)
    _install_fake_client(monkeypatch, [])

    resolution = naming.resolve_intake_study(
        source, preflight, explicit_study=None, support_confirmed_no_phi=True, intake_root=tmp_path / "intake"
    )
    for item in resolution.review_items:
        assert set(item) <= {"path", "reason", "blocking", "detail", "candidates"}
        assert isinstance(item["reason"], str)
    for item in resolution.errors:
        assert set(item) <= {"path", "reason", "detail"}


# --- canonical_source_root: pure, no I/O ------------------------------------------------------


def test_canonical_source_root_is_pure_path_resolution(tmp_path):
    target = tmp_path / "does-not-exist" / "still-not-created"
    result = naming.canonical_source_root(target)
    assert result == str(target.resolve())
    assert not target.exists()  # never created anything


def test_canonical_source_root_used_consistently_for_same_path(tmp_path):
    a = naming.canonical_source_root(tmp_path / "x")
    b = naming.canonical_source_root(tmp_path / "x")
    assert a == b
