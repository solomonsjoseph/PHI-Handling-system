"""Deterministic column-transform reference semantics (rewrite plan step 11).

Moved verbatim out of ``agents/reasoning.py`` when ``Executor`` became a
code-writing agent: the live execution path no longer drives a fixed,
hand-coded action table -- Executor's generated ``transformations.py`` +
``apply_<opaque_file_id>.py`` do that now, checked by the step 9 codegen
chain and (once landed) Task 12's ``OutputVerifier``.

What still needs this module, and why it is not simply deleted:

- ``control/deterministic_verifier.py``'s ``_source_value_mismatch_problem``
  recomputes the expected transform of every source cell via
  :func:`_apply_action` and compares it to what Executor's generated code
  actually wrote. This is exactly the point of an independent reference
  oracle: the generated code's prompt states the same semantics defined
  here, so this module is the ground truth the verifier checks generated
  output against, not a second copy of production logic. Deleted once
  Task 12 folds ``DeterministicVerifier`` into ``OutputVerifier`` and
  replaces this style of check with per-column effect-ledger diffing.
- ``phi_corpus/replay.py`` (a corpus-benchmark/scoring harness, not part
  of the live agent pipeline) calls :func:`apply_column_actions_to_dataset`
  directly to score decisions deterministically offline, with no LLM or
  container involved. Deleted with the rest of ``phi_corpus/`` in a later
  rewrite-plan step.
- Several tests exercise these semantics directly.

Executor itself no longer imports anything from this module.
"""
from __future__ import annotations

import csv
import hashlib
import hmac
import os
import re
from pathlib import Path
from typing import Any

import openpyxl

from ..detectors import detect_text

# Literals this pipeline itself produces and must be able to talk about.
# 45 CFR 164.514(b)(2)(i)(C) mandates the exact string "90+" for an
# aggregated age, and (B) mandates "000" for a restricted ZIP3, so both
# appear verbatim in generated code, in self-test vectors, and in the
# diagnostics fed back to a model. Presidio reads "90+" as a date and the
# ZIP3 codes as postal codes, which is correct in isolation and wrong
# here: redacting them turns a self-test report into
# "[REDACTED] != [REDACTED]" and the model can no longer see its own
# mistake. Exempted only where the span is exactly one of these, never as
# a substring.
_RESTRICTED_ZIP3 = {"036", "059", "063", "102", "203", "556", "692", "790",
                     "821", "823", "830", "831", "878", "879", "884", "890", "893"}

FIRST_PARTY_CONSTANTS: set[str] = {"90+", "000"} | set(_RESTRICTED_ZIP3)


def register_first_party_constants(values: "frozenset[str] | set[str]") -> None:
    """Add literals another first-party module is known to emit.

    `agents/reasoning.py` registers its code-generation self-test vectors
    here at import time. They cannot be declared in this module (they are
    the codegen layer's own fixtures) and this module cannot import that
    one (the gateway would then depend on the agents package, which the
    architecture test forbids), so the codegen layer pushes them down.
    """
    FIRST_PARTY_CONSTANTS.update(values)


class PseudonymRegistry:
    """Study-scoped, exact-value pseudonym registry.

    The SAME real value produces the SAME pseudonym across the entire study
    (all files, all columns). Different values produce different pseudonyms
    even if they occupy the same column role in different files.
    """
    def __init__(self, salt: str = ""):
        self._map: dict[str, str] = {}
        self._salt = salt

    def get(self, value: str) -> str:
        if not value:
            return value
        if value in self._map:
            return self._map[value]
        # deterministic 8-hex digest, salted per study so cross-study linkage is impossible
        digest = hashlib.sha256(f"{self._salt}:{value}".encode()).hexdigest()[:8]
        token = f"P{digest}"
        self._map[value] = token
        return token

    def digest(self, column: str, value: str) -> str:
        """Keyed digest for the `hash` action. HMAC over 'column:value' under
        the per-study salt, so the output cannot be reproduced without the
        server-held key even when the salt input (session id) is public."""
        return hmac.new(self._salt.encode(), f"{column}:{value}".encode(), hashlib.sha256).hexdigest()[:16]

    def save(self) -> str:
        """Encrypt this study's real-value -> pseudonym map plus its salt
        into one opaque blob. Pure function: no DB access here, so this
        class stays exactly what it was -- an in-memory registry -- and the
        caller decides where the blob lives and for how long.

        This is the reversal key: the one piece of data that makes a
        pseudonymized export re-identifiable. It must never be written next
        to ``exports`` or into the publication bundle.
        """
        from ..crypto import encrypt_reversal_map
        return encrypt_reversal_map({"salt": self._salt, "map": self._map})


def _scrub_text_cell(value: str) -> str:
    """Run Presidio + regex against a free-text cell. LLM never sees this.

    Replaces every detected PHI substring with a category token. Non-PHI
    text is preserved so clinicians retain the sentence around the redaction.
    """
    if not value:
        return value
    spans = detect_text(value, detectors=("presidio", "rule"))
    if not spans:
        return value
    spans_sorted = sorted(spans, key=lambda s: s.start, reverse=True)
    out = value
    for s in spans_sorted:
        cat = s.hipaa_category or "X"
        end = s.end
        # A detector span can overrun into adjacent markup (e.g. eating
        # part of a closing HTML tag after a name). PHI values don't
        # contain a raw '<', so clip the span at the first one found
        # inside it rather than let the substitution corrupt structure.
        lt = out.find("<", s.start, end)
        if lt != -1:
            end = lt
        out = out[: s.start] + f"[{cat}]" + out[end:]
    return out


def _apply_action(value: str, action: str, column: str, registry: "PseudonymRegistry | None" = None) -> str:
    if value is None or value == "":
        return value
    if action == "keep":
        return value
    if action == "drop":
        return ""
    if action == "cap_age_90":
        try:
            n = int(re.sub(r"[^0-9-]", "", value))
            return "90+" if n > 89 else str(n)
        except Exception:
            # Fail closed like year_only's malformed-input branch: a
            # non-numeric age (free text, "N/A", transcription artifact)
            # must not ship the original value verbatim.
            return ""
    if action == "year_only":
        m = re.search(r"(\d{4})", value)
        return m.group(1) if m else ""
    if action == "zip3_truncate":
        z = re.sub(r"[^0-9]", "", value)[:3]
        if z in _RESTRICTED_ZIP3:
            return "000"
        return z.ljust(3, "0")
    if action == "hash":
        if registry is not None:
            return registry.digest(column, value)
        return "[HASH]"
    if action == "pseudonymize":
        if registry is not None:
            return registry.get(value)
        return "[PSEUDONYM]"
    if action == "scrub_text":
        return _scrub_text_cell(value)
    if action == "human_review":
        return "[HUMAN_REVIEW_PENDING]"
    raise ValueError(f"unhandled action {action!r} for column {column!r}")


_FORMULA_LEAD_CHARS = ("=", "+", "-", "@", "\t", "\r")


def _neutralise_formula(value: str) -> str:
    """Prefix a spreadsheet-formula-shaped value with a leading apostrophe
    so a cell beginning with ``=``, ``+``, ``-``, ``@``, tab, or carriage
    return lands as inert text rather than an executable formula when the
    recipient opens the export in a spreadsheet application."""
    if value and value[0] in _FORMULA_LEAD_CHARS:
        return "'" + value
    return value


def apply_column_actions_to_dataset(src: Path, dst: Path, ext: str, decisions: list[dict[str, Any]],
                                    registry: "PseudonymRegistry | None" = None,
                                    omit_columns: set[str] | None = None) -> None:
    """Apply per-column actions to CSV or XLSX with an optional study-wide pseudonym registry.

    SEC-004 fail-closed: any column present in the source but WITHOUT a
    Judge/Sentinel decision is treated as ``drop`` (empty) rather than passed
    through verbatim. Override via env ``PHI_UNMAPPED_COLUMN_ACTION`` to
    ``scrub_text`` if the operator prefers redacted-in-place free-text.

    ``omit_columns`` (deferred human-review columns) are excluded from the
    output entirely -- never routed through ``_apply_action``, never
    written. For XLSX this deletes the column before any row is read, so a
    deferred cell's value is never even loaded into memory, not merely left
    unwritten.

    Not part of the live Executor path (see module docstring): this is the
    deterministic reference implementation the verifier checks generated
    code's output against, and the corpus replay harness's offline scorer.
    """
    omit_columns = set(omit_columns or ())
    action_by_col: dict[str, dict[str, Any]] = {}
    _dupes: list[str] = []
    for d in decisions:
        col = d.get("column", "")
        if col in action_by_col:
            _dupes.append(f"{col!r} ({action_by_col[col].get('action')!r} vs {d.get('action')!r})")
        action_by_col[col] = d
    if _dupes:
        # A duplicate decision for one column is a Judge/Sentinel/human-review
        # merge bug upstream. Silently picking whichever sorted last risked
        # shipping the looser of two conflicting actions -- fail loud instead.
        raise ValueError(f"duplicate decisions for column(s): {'; '.join(_dupes)}")
    _default_action = os.environ.get("PHI_UNMAPPED_COLUMN_ACTION", "drop").strip() or "drop"
    if _default_action not in {"drop", "scrub_text"}:
        _default_action = "drop"

    def _decision_for(col: str) -> dict[str, Any]:
        d = action_by_col.get(col)
        if d is not None:
            return d
        return {"action": _default_action, "column": col, "reason": "SEC-004 fail-closed default"}

    # Write to a temp path in the same directory and rename into place only
    # on clean completion, so a mid-write exception (detector error, corrupt
    # xlsx) never leaves a partially-transformed file at the real export
    # path -- the caller sees the exception and the tmp file is removed.
    tmp = dst.with_name(dst.name + ".tmp")
    try:
        if ext in ("csv", "tsv"):
            delim = "\t" if ext == "tsv" else ","
            with src.open("r", encoding="utf-8", errors="replace", newline="") as fin, \
                 tmp.open("w", encoding="utf-8", newline="") as fout:
                reader = csv.DictReader(fin, delimiter=delim)
                fieldnames = reader.fieldnames or []
                surviving = [c for c in fieldnames if c not in omit_columns]
                writer = csv.DictWriter(fout, fieldnames=surviving, delimiter=delim)
                writer.writeheader()
                for row in reader:
                    out_row: dict[str, str] = {}
                    for col in surviving:
                        d = _decision_for(col)
                        transformed = _apply_action(row.get(col) or "", d.get("action", "drop"), col, registry)
                        out_row[col] = _neutralise_formula(transformed)
                    writer.writerow(out_row)
        elif ext in ("xlsx", "xls"):
            wb = openpyxl.load_workbook(src)
            ws = wb[wb.sheetnames[0]]
            headers: list[str] = []
            for r in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = [str(c) if c is not None else "" for c in r]
                break
            if omit_columns:
                omit_positions = sorted(
                    (j for j, col in enumerate(headers, start=1) if col in omit_columns),
                    reverse=True,
                )
                for pos in omit_positions:
                    ws.delete_cols(pos, 1)
                headers = [c for c in headers if c not in omit_columns]
            for i in range(2, (ws.max_row or 1) + 1):
                for j, col in enumerate(headers, start=1):
                    d = _decision_for(col)
                    cell = ws.cell(row=i, column=j)
                    transformed = _apply_action(str(cell.value) if cell.value is not None else "",
                                               d.get("action", "drop"), col, registry)
                    cell.value = _neutralise_formula(transformed)  # type: ignore[attr-defined]
            wb.save(tmp)
        else:
            # Unknown extension - SEC-004 fail closed: refuse to emit verbatim.
            # Write a single-line marker file so the operator sees the block.
            tmp.write_text(
                f"[REDACTED] source extension {ext!r} not supported by executor; "
                f"content withheld to prevent PHI leak.\n",
                encoding="utf-8",
            )
    except Exception:
        tmp.unlink(missing_ok=True)
        raise
    os.replace(tmp, dst)
