"""File readers with strict dataset/narrative separation.

Core PHI constraint: DATASETS expose column headers only to any LLM. Row values
never leave process boundaries for LLM inspection. NARRATIVES may be read in
full by the LLM.

Kinds:
  - dataset:   csv, xlsx, parquet
  - narrative: pdf, docx, txt, md, eml
"""
from __future__ import annotations

import csv
import email
from itertools import islice
from pathlib import Path
from typing import Iterator

import openpyxl

try:
    import pyarrow.parquet as pq
except Exception:  # pragma: no cover
    pq = None

try:
    from docx import Document as DocxDocument
except Exception:  # pragma: no cover
    DocxDocument = None

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover
    PdfReader = None

try:
    import pytesseract  # type: ignore
    from pdf2image import convert_from_path  # type: ignore
except Exception:  # pragma: no cover
    pytesseract = None
    convert_from_path = None

try:
    import pdfplumber  # type: ignore
except Exception:  # pragma: no cover
    pdfplumber = None


# Phase C: threshold below which we treat a PDF as image-only and run OCR.
# Real-world scanned CRFs / consent forms yield near-zero text via `pypdf`.
OCR_TEXT_THRESHOLD = 50
OCR_MAX_PAGES = 100  # bounded so a giant scan can't wedge the pipeline


DATASET_EXTS = {"csv", "tsv", "xlsx", "xls", "parquet"}
NARRATIVE_EXTS = {"pdf", "docx", "txt", "md", "eml", "html", "htm"}


def classify_ext(name: str) -> tuple[str, str]:
    ext = name.rsplit(".", 1)[-1].lower() if "." in name else ""
    if ext in DATASET_EXTS:
        return "dataset", ext
    if ext in NARRATIVE_EXTS:
        return "narrative", ext
    return "narrative", ext or "txt"


# --- Dataset readers -------------------------------------------------------

def read_csv_columns(path: Path) -> tuple[list[str], int]:
    """Returns (columns, row_count). Does not retain row values."""
    with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.reader(f)
        try:
            columns = next(reader)
        except StopIteration:
            return [], 0
        rows = sum(1 for _ in reader)
    return [c.strip() for c in columns], rows


def read_xlsx_columns(path: Path) -> tuple[list[str], int]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    columns: list[str] = []
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        columns = [str(c) if c is not None else "" for c in row]
        break
    row_count = max(0, (ws.max_row or 0) - 1)
    wb.close()
    return columns, row_count

def _iter_table_rows(path: Path, ext: str) -> Iterator[list[str]]:
    if ext in ("csv", "tsv"):
        delimiter = "\t" if ext == "tsv" else ","
        with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
            for row in csv.reader(f, delimiter=delimiter):
                values = [str(cell) for cell in row]
                if any(cell.strip() for cell in values):
                    yield values
        return
    if ext in ("xlsx", "xls"):
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook[workbook.sheetnames[0]]
            for row in worksheet.iter_rows(values_only=True):
                values = [str(cell) if cell is not None else "" for cell in row]
                if any(cell.strip() for cell in values):
                    yield values
        finally:
            workbook.close()
        return
    raise ValueError(f"unsupported table ext {ext!r}")


def _table_column_key(column: str) -> str:
    return column.strip().casefold()


def _read_table_rows(
    path: Path, ext: str, max_rows: int
) -> tuple[list[str], list[list[str]]]:
    if max_rows < 0:
        raise ValueError("max_rows must be non-negative")
    rows = _iter_table_rows(path, ext)
    try:
        try:
            headers = next(rows)
        except StopIteration:
            return [], []
        if ext in ("csv", "tsv"):
            headers = [cell.strip() for cell in headers]
        return headers, list(islice(rows, max_rows))
    finally:
        rows.close()

def read_table_rows(path: Path, max_rows: int = 5000) -> tuple[list[str], list[list[str]]]:
    """Return a CSV, TSV, or workbook header and up to ``max_rows`` string rows."""
    return _read_table_rows(path, path.suffix.lstrip(".").lower(), max_rows)


def column_value_stats(
    path: Path,
    ext: str,
    columns: list[str],
    max_rows: int = 50_000,
) -> dict[str, dict[str, int]]:
    """Return bounded per-column distinct-value and row counts without retaining values."""
    if max_rows < 0:
        raise ValueError("max_rows must be non-negative")
    normalized_ext = ext.lower().lstrip(".")
    table_rows = _iter_table_rows(path, normalized_ext)
    try:
        try:
            headers = next(table_rows)
        except StopIteration:
            return {column: {"distinct": 0, "rows": 0} for column in columns}
        if normalized_ext in ("csv", "tsv"):
            headers = [cell.strip() for cell in headers]
        positions = {_table_column_key(header): index for index, header in enumerate(headers)}
        distinct_values = {column: set() for column in columns}
        row_counts = {column: 0 for column in columns}
        tracked_columns = [
            (column, positions.get(_table_column_key(column)), distinct_values[column])
            for column in distinct_values
        ]
        for row in islice(table_rows, max_rows):
            for column, index, values in tracked_columns:
                if index is None:
                    continue
                row_counts[column] += 1
                values.add(row[index] if index < len(row) else "")
        return {
            column: {"distinct": len(distinct_values[column]), "rows": row_counts[column]}
            for column in columns
        }
    finally:
        table_rows.close()


def read_parquet_columns(path: Path) -> tuple[list[str], int]:
    if pq is None:
        raise RuntimeError("pyarrow is not installed")
    pf = pq.ParquetFile(path)
    return list(pf.schema_arrow.names), pf.metadata.num_rows


def iter_dataset_rows(path: Path, ext: str) -> Iterator[tuple[int, dict[str, str]]]:
    """Yield (row_index, {column: value}). Consumer applies detection per cell."""
    if ext in ("csv", "tsv"):
        delim = "\t" if ext == "tsv" else ","
        with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
            reader = csv.DictReader(f, delimiter=delim)
            for i, row in enumerate(reader):
                yield i, {k: (v or "") for k, v in row.items()}
    elif ext in ("xlsx", "xls"):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        header: list[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                header = [str(c) if c is not None else "" for c in row]
                continue
            yield i - 1, {header[j] if j < len(header) else f"col_{j}": ("" if v is None else str(v)) for j, v in enumerate(row)}
        wb.close()
    elif ext == "parquet":
        if pq is None:
            raise RuntimeError("pyarrow is not installed")
        table = pq.read_table(path)
        cols = table.column_names
        rows = table.to_pylist()
        for i, r in enumerate(rows):
            yield i, {c: ("" if r.get(c) is None else str(r.get(c))) for c in cols}
    else:
        raise ValueError(f"unsupported dataset ext {ext!r}")


# --- Narrative readers -----------------------------------------------------

def read_txt(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def read_pdf(path: Path) -> str:
    """Extract text from a PDF.

    Phase C: if the digital text-layer is missing or nearly empty
    (e.g. scanned CRFs, image-only consent forms) we fall back to OCR
    via ``pytesseract`` on rasterised pages. OCR output is then fed
    through the same PHI scrubbing pipeline as any narrative text.
    """
    if PdfReader is None:
        raise RuntimeError("pypdf is not installed")
    reader = PdfReader(str(path))
    text = "\n\n".join((page.extract_text() or "") for page in reader.pages)
    # If we got real digital text, use it as-is.
    if len(text.strip()) >= OCR_TEXT_THRESHOLD:
        return text
    # Otherwise: image-only PDF -> OCR every page.
    if pytesseract is None or convert_from_path is None:
        # OCR stack unavailable at runtime; return whatever we had.
        return text
    try:
        ocr_pages: list[str] = []
        images = convert_from_path(str(path), dpi=200, last_page=OCR_MAX_PAGES)
        for img in images:
            ocr_pages.append(pytesseract.image_to_string(img) or "")
        ocr_text = "\n\n".join(ocr_pages).strip()
        if ocr_text:
            return ocr_text
    except Exception:
        # Best-effort: never crash the pipeline; fall through with digital text
        # (which is what a purely-image PDF would have been anyway).
        pass
    return text


def read_pdf_form_fields(path: Path) -> list[dict[str, str | None]] | None:
    """Tier-1 deterministic extraction for true fillable (AcroForm) PDFs.

    Returns ``None`` when the PDF carries no AcroForm fields -- the caller
    should fall back to LLM extraction on flattened/OCR'd text (``read_pdf``)
    for flat or scanned forms, which have no widgets to read.

    Each real field name comes straight from the PDF's own AcroForm
    dictionary via ``pypdf`` -- never inferred, never guessed. The printed
    label is the nearest line of extracted text directly above the field's
    widget, matched by page position via ``pdfplumber``. Adapted from the
    annotation-geometry proximity-matching technique in
    ``solomonsjoseph/RePORT-AI-Portal`` (``PHI_handing_review`` branch,
    ``plugins/report-ai-study-pipeline/skills/sot-lean-generator/scripts/
    generate_pdf_aware_candidate.py``) -- that script's per-study alias
    tables and field overrides are not reused, only the general technique.
    """
    if PdfReader is None or pdfplumber is None:
        return None
    raw_fields = PdfReader(str(path)).get_fields()
    if not raw_fields:
        return None

    fields: list[dict[str, str | None]] = []
    with pdfplumber.open(str(path)) as doc:
        for page in doc.pages:
            words = sorted(
                page.extract_words(x_tolerance=1, y_tolerance=3) or [],
                key=lambda w: (float(w.get("top") or 0), float(w.get("x0") or 0)),
            )
            for annot in page.annots or []:
                data = annot.get("data") or {}
                name = data.get("T")
                if name is None:
                    continue
                name = name.decode("utf-8", "replace") if isinstance(name, bytes) else str(name)
                ax = (float(annot["x0"]) + float(annot["x1"])) / 2
                ay = (float(annot["top"]) + float(annot["bottom"])) / 2
                fields.append({"label": _nearest_label(words, ax, ay), "collected_variable": name})
    return fields or None


def _nearest_label(words: list[dict], ax: float, ay: float) -> str | None:
    """Nearest line of text above (falling back to left of) a widget's
    center point, joined left-to-right. Simple line-grouping by shared
    ``top`` within a small tolerance, not the reference script's full
    per-segment scoring -- adequate for a single label per widget; a form
    with dense multi-column layouts may need the fuller algorithm later."""
    above = [w for w in words if 0 < ay - float(w.get("top") or 0) < 40]
    same_row = [w for w in words if abs(ay - float(w.get("top") or 0)) < 6 and float(w.get("x1") or 0) <= ax]
    candidates = above or same_row
    if not candidates:
        return None
    target_top = float(candidates[-1].get("top") or 0)
    line = [w for w in candidates if abs(float(w.get("top") or 0) - target_top) < 4]
    line.sort(key=lambda w: float(w.get("x0") or 0))
    text = " ".join(str(w.get("text") or "") for w in line).strip()
    return text or None


# SEC-001 (audit iter_22): narrative-path docx bomb defence lives in
# phi_core.docx_safe so it stays in lock-step with the dictionary-path
# reader in phi_core.agents.specialists (iter_22 root-caused this drift).


def read_docx(path: Path) -> str:
    """Extract paragraph text from a .docx narrative file.

    Uses ``python-docx`` for the actual paragraph walk, but first
    verifies the archive is safe (size-capped, valid ZIP) via
    ``docx_safe.is_safe_docx``. Returns "" on any unsafe archive so the
    pipeline continues with no narrative text rather than OOMing.
    """
    if DocxDocument is None:
        raise RuntimeError("python-docx is not installed")
    from .docx_safe import is_safe_docx
    if not is_safe_docx(path):
        return ""
    d = DocxDocument(str(path))
    return "\n".join(p.text for p in d.paragraphs)


def read_eml(path: Path) -> str:
    msg = email.message_from_bytes(path.read_bytes())
    hdrs = "\n".join(f"{k}: {v}" for k, v in msg.items())
    body_parts: list[str] = []
    if msg.is_multipart():
        for part in msg.walk():
            if part.get_content_type() == "text/plain":
                payload = part.get_payload(decode=True)
                if payload:
                    body_parts.append(payload.decode("utf-8", errors="replace"))
    else:
        payload = msg.get_payload(decode=True)
        if payload:
            body_parts.append(payload.decode("utf-8", errors="replace"))
    return hdrs + "\n\n" + "\n\n".join(body_parts)


def read_narrative(path: Path, ext: str) -> str:
    if ext == "pdf":
        return read_pdf(path)
    if ext == "docx":
        return read_docx(path)
    if ext == "eml":
        return read_eml(path)
    return read_txt(path)
