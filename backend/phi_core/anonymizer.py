"""Anonymizer: apply review decisions and produce PHI-scrubbed output.

Strategies per span:
  - accept + no replacement -> [REDACTED:<hipaa>]
  - accept + replacement    -> use replacement
  - reject                  -> keep original
  - reclassify              -> use new_category label, still redact with tag
Deterministic: no randomness during application.
"""
from __future__ import annotations

import csv
from pathlib import Path

import openpyxl

from .models import DetectedSpan


def _tag(span: DetectedSpan) -> str:
    if span.replacement:
        return span.replacement
    hipaa = span.hipaa_category or "X"
    return f"[REDACTED:{hipaa}:{span.entity_type}]"


def apply_to_text(text: str, spans: list[DetectedSpan]) -> str:
    """Apply accepted spans to text. Reject-kept spans are ignored."""
    active = [s for s in spans if s.review_status in ("accepted", "reclassified")]
    active.sort(key=lambda s: s.start, reverse=True)
    out = text
    for s in active:
        out = out[:s.start] + _tag(s) + out[s.end:]
    return out


def apply_to_dataset(
    src_path: Path,
    dst_path: Path,
    ext: str,
    spans_by_cell: dict[tuple[int, str], list[DetectedSpan]],
    columns_full_redact: dict[str, DetectedSpan],
) -> None:
    """Copy dataset src -> dst with per-cell redactions applied.

    spans_by_cell: {(row_index, column): [spans]} for cell-level partial redactions.
    columns_full_redact: {column: template_span} for header-flagged columns
        whose entire value is replaced.
    """
    if ext in ("csv", "tsv"):
        delim = "\t" if ext == "tsv" else ","
        with src_path.open("r", encoding="utf-8", errors="replace", newline="") as fin, \
             dst_path.open("w", encoding="utf-8", newline="") as fout:
            reader = csv.DictReader(fin, delimiter=delim)
            fieldnames = reader.fieldnames or []
            writer = csv.DictWriter(fout, fieldnames=fieldnames, delimiter=delim)
            writer.writeheader()
            for i, row in enumerate(reader):
                for col, ts in columns_full_redact.items():
                    if col in row and row[col]:
                        row[col] = _tag(ts)
                for (ri, col), cell_spans in spans_by_cell.items():
                    if ri == i and col in row and row[col]:
                        row[col] = apply_to_text(row[col], cell_spans)
                writer.writerow(row)
    elif ext in ("xlsx", "xls"):
        wb = openpyxl.load_workbook(src_path)
        ws = wb[wb.sheetnames[0]]
        headers: list[str] = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(c) if c is not None else "" for c in row]
            break
        for i in range(2, (ws.max_row or 1) + 1):
            for j, col in enumerate(headers, start=1):
                cell = ws.cell(row=i, column=j)
                val = cell.value
                if val is None:
                    continue
                s = str(val)
                if col in columns_full_redact:
                    cell.value = _tag(columns_full_redact[col])
                    continue
                key = (i - 2, col)
                if key in spans_by_cell:
                    cell.value = apply_to_text(s, spans_by_cell[key])
        wb.save(dst_path)
    else:
        raise ValueError(f"anonymize: unsupported ext {ext!r}")
