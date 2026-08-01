"""Intake edge-case audit. Runs against the live backend on :8001.

Categories tested:
  A. Happy paths (multiple valid combinations)
  B. Missing component paths
  C. Unsafe ZIP paths (traversal, symlinks, absolute)
  D. Type mismatches (.json in datasets, .csv in forms, multi-sheet xlsx)
  E. Aliases and case variations (Datasets/, DATA/, codebook/)
  F. Structural quirks (single-root wrapper, nested subdirs, zero-byte file, empty zip)
  G. Non-ZIP payloads
  H. Re-intake replaces prior intake
"""
from __future__ import annotations

import io
import os
import subprocess
import sys
import zipfile
from pathlib import Path

import openpyxl
import requests
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter


API = "http://localhost:8001"
ROOT = Path("/tmp/intake_tests")
ROOT.mkdir(parents=True, exist_ok=True)


# ---------- fixture helpers ----------

def make_csv(path: Path, headers: list[str], rows: int = 2) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [",".join(headers)]
    for i in range(rows):
        lines.append(",".join(f"v{i}_{j}" for j in range(len(headers))))
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def make_pdf(path: Path, text: str = "Consent Form") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    c = canvas.Canvas(str(path), pagesize=letter)
    c.drawString(72, 720, text)
    c.save()


def make_xlsx_single(path: Path, headers: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "sheet1"
    ws.append(headers)
    ws.append([f"v{i}" for i in range(len(headers))])
    wb.save(path)


def make_xlsx_multi(path: Path, headers: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.active.title = "sheet1"
    wb.active.append(headers)
    wb.create_sheet("sheet2").append(["a", "b"])
    wb.save(path)


def build_zip(zip_path: Path, files_map: dict[str, Path]) -> None:
    """Build a zip with arcname -> filesystem path mapping."""
    zip_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for arcname, src in files_map.items():
            z.write(src, arcname=arcname)


def build_zip_with_raw(zip_path: Path, entries: dict[str, bytes]) -> None:
    """Build a zip with arcname -> bytes."""
    zip_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        for arcname, data in entries.items():
            z.writestr(arcname, data)


def build_zip_with_symlink(zip_path: Path, real_files: dict[str, bytes], symlink: tuple[str, str]) -> None:
    """Build a zip that contains a symlink entry (mode 0xA000)."""
    zip_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zip_path, "w") as z:
        for arcname, data in real_files.items():
            z.writestr(arcname, data)
        arcname, target = symlink
        info = zipfile.ZipInfo(arcname)
        info.external_attr = (0xA1FF << 16)   # symlink
        z.writestr(info, target)


def new_session() -> str:
    r = requests.post(f"{API}/api/sessions", json={"jurisdiction": "us"}, timeout=10)
    r.raise_for_status()
    return r.json()["id"]


def intake(sid: str, zip_path: Path) -> dict:
    with zip_path.open("rb") as f:
        r = requests.post(
            f"{API}/api/sessions/{sid}/intake",
            files={"file": (zip_path.name, f, "application/zip")},
            timeout=30,
        )
    if r.status_code != 200:
        return {"http": r.status_code, "detail": r.json().get("detail", r.text)}
    return r.json()


# ---------- prebuild small assets ----------

def _prebuild():
    make_csv(ROOT / "assets" / "patients.csv", ["patient_name", "dob", "ssn", "mrn"])
    make_csv(ROOT / "assets" / "codebook.csv", ["column_name", "description"])
    make_pdf(ROOT / "assets" / "consent.pdf")
    make_xlsx_single(ROOT / "assets" / "patients.xlsx", ["patient_name", "dob"])
    make_xlsx_multi(ROOT / "assets" / "multi.xlsx", ["patient_name", "dob"])


PASS = "PASS"
FAIL = "FAIL"
_results: list[tuple[str, str, str, str]] = []


def _run(name: str, expected_status: str, expected_missing_contains: list[str] | None,
         expected_review_contains: list[str] | None, builder) -> None:
    zpath = ROOT / f"{name}.zip"
    if zpath.exists():
        zpath.unlink()
    try:
        builder(zpath)
        sid = new_session()
        r = intake(sid, zpath)
        status = r.get("status")
        missing = r.get("missing_components") or []
        review_reasons = [e["reason"] for e in r.get("review_entries") or []]
        ok = True
        why = []
        if status != expected_status:
            ok = False
            why.append(f"status={status!r} want {expected_status!r}")
        if expected_missing_contains is not None:
            for m in expected_missing_contains:
                if not any(m in x for x in missing):
                    ok = False
                    why.append(f"missing lacks {m!r}: {missing}")
        if expected_review_contains is not None:
            for m in expected_review_contains:
                if not any(m in x for x in review_reasons):
                    ok = False
                    why.append(f"review lacks {m!r}: {review_reasons}")
        _results.append((name, PASS if ok else FAIL, status or "?", "; ".join(why) or f"ok linked={r.get('linked')} review={r.get('review')}"))
    except Exception as e:
        _results.append((name, FAIL, "-", f"exception: {type(e).__name__}: {e}"))


_prebuild()

# A. Happy paths -----------------------------------------------------------
def A1(z):
    build_zip(z, {
        "datasets/patients.csv": ROOT / "assets/patients.csv",
        "forms/consent.pdf": ROOT / "assets/consent.pdf",
    })
_run("A1_datasets_plus_forms", "ready", None, None, A1)

def A2(z):
    build_zip(z, {
        "datasets/patients.csv": ROOT / "assets/patients.csv",
        "data_dictionary/codebook.csv": ROOT / "assets/codebook.csv",
    })
_run("A2_datasets_plus_dict", "ready", None, None, A2)

def A3(z):
    build_zip(z, {
        "datasets/patients.csv": ROOT / "assets/patients.csv",
        "mappings/codebook.csv": ROOT / "assets/codebook.csv",
    })
_run("A3_datasets_plus_mappings", "ready", None, None, A3)

def A4(z):
    build_zip(z, {
        "datasets/patients.xlsx": ROOT / "assets/patients.xlsx",
        "forms/consent.pdf": ROOT / "assets/consent.pdf",
    })
_run("A4_xlsx_single_sheet_dataset", "ready", None, None, A4)

def A5(z):  # nested subdirs inside a component
    build_zip(z, {
        "datasets/2024/q1/patients.csv": ROOT / "assets/patients.csv",
        "forms/site_a/consent.pdf": ROOT / "assets/consent.pdf",
    })
_run("A5_nested_subdirs", "ready", None, None, A5)

def A6(z):  # single-root wrapper
    build_zip(z, {
        "study_xyz/datasets/patients.csv": ROOT / "assets/patients.csv",
        "study_xyz/forms/consent.pdf": ROOT / "assets/consent.pdf",
    })
_run("A6_single_root_wrapper", "ready", None, None, A6)

# B. Missing components ----------------------------------------------------
def B1(z):
    build_zip(z, {"forms/consent.pdf": ROOT / "assets/consent.pdf"})
_run("B1_no_datasets", "failed", ["datasets"], None, B1)

def B2(z):
    build_zip(z, {"datasets/patients.csv": ROOT / "assets/patients.csv"})
_run("B2_no_forms_no_dict", "failed", ["one_of_forms"], None, B2)

def B3(z):  # empty zip
    build_zip_with_raw(z, {})
_run("B3_empty_zip", "failed", ["datasets"], None, B3)

# C. Unsafe zip -----------------------------------------------------------
def C1(z):
    build_zip_with_raw(z, {
        "datasets/patients.csv": b"name,age\n",
        "forms/consent.pdf": b"%PDF-1.4\n",
        "../etc_passwd": b"root:x:0:0:root\n",
    })
_run("C1_path_traversal", "failed", None, None, C1)

def C2(z):
    build_zip_with_symlink(z, {
        "datasets/patients.csv": b"name,age\n",
        "forms/consent.pdf": b"%PDF-1.4\n",
    }, symlink=("forms/link_to_etc", "/etc/hosts"))
_run("C2_symlink_entry", "failed", None, None, C2)

def C3(z):  # absolute path
    build_zip_with_raw(z, {
        "/tmp/patients.csv": b"name,age\n",
        "datasets/patients.csv": b"name,age\n",
        "forms/consent.pdf": b"%PDF-1.4\n",
    })
_run("C3_absolute_path", "failed", None, None, C3)

# D. Type mismatches ------------------------------------------------------
def D1(z):
    build_zip_with_raw(z, {"datasets/patients.json": b"{\"a\":1}",
                            "forms/consent.pdf": b"%PDF-1.4\n"})
    # datasets must have at least one accepted file; here .json is unclassified,
    # so there is no ACCEPTED dataset -> datasets component missing.
_run("D1_json_in_datasets", "failed", ["datasets"], [".json"], D1)

def D2(z):
    build_zip(z, {
        "datasets/patients.csv": ROOT / "assets/patients.csv",
        "forms/patients.csv": ROOT / "assets/patients.csv",  # csv in forms not allowed
    })
_run("D2_csv_in_forms", "failed", ["one_of_forms"], [".csv"], D2)

def D3(z):
    build_zip(z, {
        "datasets/multi.xlsx": ROOT / "assets/multi.xlsx",  # multi-sheet not allowed
        "forms/consent.pdf": ROOT / "assets/consent.pdf",
    })
_run("D3_multi_sheet_xlsx", "failed", ["datasets"], ["single-sheet"], D3)

def D4(z):
    build_zip_with_raw(z, {
        "datasets/patients.csv": b"name,age\n",
        "unknown_top/random.txt": b"noise",  # unknown top dir
        "forms/consent.pdf": b"%PDF-1.4\n",
    })
_run("D4_unknown_top_dir", "review_required", None, ["unknown_top"], D4)

# E. Aliases and case variations ------------------------------------------
def E1(z):  # aliases: data/ -> datasets, dictionary/ -> data_dictionary
    build_zip_with_raw(z, {
        "data/patients.csv": b"name,age\n",
        "dictionary/codebook.csv": b"col,desc\n",
    })
_run("E1_alias_data_and_dictionary", "ready", None, None, E1)

def E2(z):  # case variation
    build_zip_with_raw(z, {
        "DataSets/patients.csv": b"name,age\n",
        "FORMS/consent.pdf": b"%PDF-1.4\n",
    })
_run("E2_uppercase_dirs", "ready", None, None, E2)

def E3(z):  # codebook alias
    build_zip_with_raw(z, {
        "datasets/patients.csv": b"name,age\n",
        "codebook/mapping.csv": b"col,desc\n",
    })
_run("E3_codebook_alias", "ready", None, None, E3)

# F. Structural quirks ----------------------------------------------------
def F1(z):  # files at zip root, no folder
    build_zip_with_raw(z, {
        "patients.csv": b"name,age\n",
        "consent.pdf": b"%PDF-1.4\n",
    })
_run("F1_no_component_dirs", "failed", ["datasets"], ["not a known component"], F1)

def F2(z):  # zero-byte file in datasets => blocks + also missing datasets
    build_zip_with_raw(z, {
        "datasets/empty.csv": b"",
        "forms/consent.pdf": b"%PDF-1.4\n",
    })
_run("F2_zero_byte_dataset", "failed", ["datasets"], ["empty"], F2)

def F3(z):  # duplicate content across components (data_dictionary sorts first,
    # so datasets/patients.csv becomes _unclassified; datasets/ ends up missing).
    body = b"col_a,col_b\n1,2\n"
    build_zip_with_raw(z, {
        "datasets/patients.csv": body,
        "data_dictionary/codebook.csv": body,
    })
_run("F3_duplicate_across_components", "failed", ["datasets"], ["duplicate content"], F3)

def F4(z):  # non-PDF file with .pdf extension
    build_zip_with_raw(z, {
        "datasets/patients.csv": b"name,age\n",
        "forms/fake.pdf": b"NOT A PDF",
    })
_run("F4_fake_pdf", "failed", ["one_of_forms"], ["not a PDF"], F4)

def F5(z):  # CSV with empty header line
    build_zip_with_raw(z, {
        "datasets/patients.csv": b"\n",
        "forms/consent.pdf": b"%PDF-1.4\n",
    })
_run("F5_empty_csv_header", "failed", ["datasets"], ["header"], F5)

# G. Non-ZIP ---------------------------------------------------------------
def G1(z):
    z.write_bytes(b"this is not a zip file")
_run("G1_not_a_zip", "failed", None, None, G1)

# H. Re-intake replaces prior --------------------------------------------
def H1(_):
    # First intake fails, then a fresh intake succeeds on the same session.
    zbad = ROOT / "H1_bad.zip"
    zgood = ROOT / "H1_good.zip"
    build_zip_with_raw(zbad, {"forms/consent.pdf": b"%PDF-1.4\n"})
    build_zip(zgood, {
        "datasets/patients.csv": ROOT / "assets/patients.csv",
        "forms/consent.pdf": ROOT / "assets/consent.pdf",
    })
    sid = new_session()
    r1 = intake(sid, zbad)
    r2 = intake(sid, zgood)
    ok = r1.get("status") == "failed" and r2.get("status") == "ready"
    _results.append((
        "H1_re_intake_replaces_prior", PASS if ok else FAIL,
        f"{r1.get('status')} -> {r2.get('status')}",
        "ok" if ok else f"unexpected outcome",
    ))

H1(None)

# ---------- report ----------
w = max(len(n) for n, *_ in _results) + 2
print(f"{'test':{w}}{'result':6} {'status':16}  detail")
print("-" * (w + 60))
fail_count = 0
for name, res, status, detail in _results:
    if res == FAIL:
        fail_count += 1
    print(f"{name:{w}}{res:6} {status:16}  {detail}")
print("-" * (w + 60))
print(f"total: {len(_results)}  passed: {len(_results) - fail_count}  failed: {fail_count}")
sys.exit(1 if fail_count else 0)
