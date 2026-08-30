"""Tests for the section-59 column ledger (Phase 11b, docs #59):
``build_column_ledger_rows`` correlation logic and
``write_column_ledger_xlsx``'s real xlsx output.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from phi_core.control.column_ledger import (
    COLUMN_LEDGER_HEADERS,
    build_column_ledger_rows,
    write_column_ledger_xlsx,
)
from phi_core.control.records import (
    ColumnDecision,
    ExecutionResult,
    HumanReviewEvent,
    ResolutionEntry,
    ReviewFinding,
    VerificationResult,
)


def _decision(**overrides) -> ColumnDecision:
    base = dict(
        run_id="r1", file_id="dataset_a", column_id="col_1", safe_display_name="Visit Date",
        semantic_meaning="Clinical visit date", sensitivity_classification="restricted_phi",
        applicable_rule="HIPAA Safe Harbor", operation="date_shift", method_id="m1", method_version=1,
        plain_language_reason="Dates identify patients", decision_status="verified",
    )
    base.update(overrides)
    return ColumnDecision(**base)


def test_one_row_per_logical_column_regardless_of_correction_chain_length():
    decisions = [
        _decision(operation="keep", decision_status="correction_required", created_at="2026-01-01T00:00:00Z"),
        _decision(operation="date_shift", decision_status="verified", created_at="2026-01-02T00:00:00Z"),
        _decision(file_id="dataset_a", column_id="col_2", safe_display_name="MRN", operation="pseudonymize",
                   created_at="2026-01-01T00:00:00Z"),
    ]
    rows = build_column_ledger_rows(decisions)
    assert len(rows) == 2  # two logical columns, not three decisions
    by_column = {r.column: r for r in rows}
    assert by_column["Visit Date"].correction_made == "Yes"
    assert by_column["MRN"].correction_made == "No"


def test_initial_vs_final_action_reflects_correction_chain_order():
    decisions = [
        _decision(operation="keep", created_at="2026-01-01T00:00:00Z"),
        _decision(operation="date_shift", created_at="2026-01-02T00:00:00Z"),
    ]
    row = build_column_ledger_rows(decisions)[0]
    assert row.initial_proposed_action == "keep"
    assert row.final_approved_action == "date_shift"
    assert row.what_we_did == "Shifted dates by a per-subject random offset"


def test_no_correction_single_decision_has_matching_initial_and_final():
    row = build_column_ledger_rows([_decision(operation="keep")])[0]
    assert row.initial_proposed_action == row.final_approved_action == "keep"
    assert row.correction_made == "No"


def test_reviewer_result_matches_by_file_and_column():
    decisions = [_decision()]
    findings = [ReviewFinding(verdict="CORRECTION_REQUIRED", file_id="dataset_a", column="col_1", detail="needs fix")]
    row = build_column_ledger_rows(decisions, review_findings=findings)[0]
    assert row.reviewer_result == "CORRECTION_REQUIRED: needs fix"


def test_reviewer_result_is_na_when_no_finding_recorded():
    row = build_column_ledger_rows([_decision()])[0]
    assert row.reviewer_result == "N/A"


def test_human_review_fields_populated_from_matching_resolution():
    decisions = [_decision()]
    events = [HumanReviewEvent(
        request_id="req1", run_id="r1", session_id="s1", workflow_version="1", task_id="t1",
        seq=1, client_event_id="c1", principal="dr.smith", kind="resolution", body_hash="h",
        resolutions=[ResolutionEntry(file_id="dataset_a", column="col_1", mode="approve", comment="fine")],
    )]
    row = build_column_ledger_rows(decisions, human_review_events=events)[0]
    assert row.human_review == "Yes"
    assert row.human_decision == "approve: fine"


def test_human_review_is_no_when_no_resolution_matches_the_column():
    decisions = [_decision()]
    events = [HumanReviewEvent(
        request_id="req1", run_id="r1", session_id="s1", workflow_version="1", task_id="t1",
        seq=1, client_event_id="c1", principal="dr.smith", kind="resolution", body_hash="h",
        resolutions=[ResolutionEntry(file_id="other_file", column="other_col", mode="approve")],
    )]
    row = build_column_ledger_rows(decisions, human_review_events=events)[0]
    assert row.human_review == "No"
    assert row.human_decision == ""


def test_executor_and_verification_results_reflect_run_level_records():
    decisions = [_decision()]
    execution_result = ExecutionResult(task_id="t1", run_id="r1", manifest_id="m1", success=False,
                                        failure_class="io_error")
    verification_result = VerificationResult(run_id="r1", passed=False, failed_checks=["coverage_gap"])
    row = build_column_ledger_rows(
        decisions, execution_result=execution_result, verification_result=verification_result,
    )[0]
    assert row.executor_result == "failed: io_error"
    assert row.final_verification == "failed: coverage_gap"


def test_method_names_lookup_resolves_display_name():
    row = build_column_ledger_rows([_decision(method_id="m1")], method_names={"m1": "Date Shift v2"})[0]
    assert row.method_used == "Date Shift v2"


def test_method_names_lookup_falls_back_to_method_id_when_unresolved():
    row = build_column_ledger_rows([_decision(method_id="m1")])[0]
    assert row.method_used == "m1"


def test_evidence_references_join_semantic_and_regulatory_refs():
    row = build_column_ledger_rows(
        [_decision(semantic_evidence_refs=["ev_sem"], regulatory_evidence_refs=["ev_reg"])],
    )[0]
    assert row.evidence_references == "ev_sem; ev_reg"


def test_dataset_file_includes_dataset_part_id_when_present():
    row = build_column_ledger_rows([_decision(dataset_part_id="part_2")])[0]
    assert row.dataset_file == "dataset_a/part_2"


def test_write_column_ledger_xlsx_has_nineteen_columns_and_every_logical_column(tmp_path: Path):
    decisions = [
        _decision(),
        _decision(file_id="dataset_a", column_id="col_2", safe_display_name="MRN", operation="pseudonymize"),
        _decision(file_id="dataset_b", column_id="col_1", safe_display_name="Zip", operation="generalize"),
    ]
    rows = build_column_ledger_rows(decisions)
    path = write_column_ledger_xlsx(rows, tmp_path / "What_Happened_to_Each_Column.xlsx")

    assert path.exists()
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    values = list(ws.iter_rows(values_only=True))

    header = values[0]
    assert len(header) == 19 == len(COLUMN_LEDGER_HEADERS)
    assert list(header) == list(COLUMN_LEDGER_HEADERS)

    data_rows = values[1:]
    assert len(data_rows) == 3  # every logical column appears exactly once
    columns_seen = {(r[0], r[1]) for r in data_rows}
    assert columns_seen == {("dataset_a", "Visit Date"), ("dataset_a", "MRN"), ("dataset_b", "Zip")}


def test_write_column_ledger_xlsx_with_no_rows_still_writes_header_only(tmp_path: Path):
    path = write_column_ledger_xlsx([], tmp_path / "empty.xlsx")
    wb = openpyxl.load_workbook(path)
    values = list(wb.active.iter_rows(values_only=True))
    assert len(values) == 1
    assert list(values[0]) == list(COLUMN_LEDGER_HEADERS)
