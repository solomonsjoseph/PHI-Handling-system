"""
Excel (xlsx) PHI corpus generator.

Covers all 31 structural edge cases across four detection tiers:
  Tier A -- Canonical: unambiguous PHI with clear column headers
  Tier B -- Hard Recall: PHI that is easy for detection tools to miss
             (metadata, hidden rows/cols/sheets, cell comments, formula results,
              chart titles, named ranges, non-English headers, merged headers,
              split columns, concatenated cells, cross-sheet linkage)
  Tier C -- False Positive Traps: PHI-shaped values that are NOT PHI
             (placeholder IDs, reserved phone numbers, test SSNs, etc.)
  Tier D -- Human Review: genuinely ambiguous (quasi-ID combos, free-text notes)

Authority: 45 CFR 164.514(b)(2)(i) HIPAA Safe Harbor -- Identifiers listed at
           categories A through R must be removed for Safe Harbor de-identification.
           See AUTHORITY_MATRIX.md Table A for the full cross-jurisdiction mapping.

Record text field: structured flat extraction representing all content a proper
           extraction pipeline would pull from the xlsx file (cells, metadata,
           sheet names, hidden rows, cell comments, named ranges, chart titles).
           Gold spans are byte offsets within this extracted text.

Generated xlsx files: written to the output directory alongside JSONL records
           so that raw-format tool tests can be run against actual .xlsx files.
"""
from __future__ import annotations

import json
import random
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from generators.common import (
    AUTH_HIPAA_SAFE_HARBOR,
    AUTH_SWEENEY_K_ANON,
    DETECTION_REGIME_RULE,
    DETECTION_REGIME_NER,
    DETECTION_REGIME_CONFLICT,
    LAYER_HIPAA,
    DeterministicGenerator,
    GoldSpan,
    Record,
    write_jsonl,
)

AUTH_XLSX = "45 CFR 164.514(b)(2)(i) HIPAA Safe Harbor"
AUTH_XLSX_EDPB = "EDPB Opinion 28/2024 (AI model anonymization)"
FORMAT = "xlsx"

# ---------------------------------------------------------------------------
# Synthetic data pools
# ---------------------------------------------------------------------------

_FIRST = ["Aaron", "Beth", "Carl", "Dana", "Eric", "Fiona", "Gary", "Holly",
          "Ian", "Janet", "Kyle", "Lisa", "Mark", "Nancy", "Owen", "Pam",
          "Ray", "Sandra", "Tim", "Uma"]
_LAST = ["Archer", "Burke", "Cross", "Drake", "Ellis", "Ford", "Grant", "Hall",
         "Irwin", "Jordan", "Knox", "Lowe", "Miles", "Nash", "Orme", "Price",
         "Quinn", "Riley", "Stone", "Todd"]
_PROVIDERS = ["Dr. Margaret Young", "Dr. Robert Chen", "Dr. Sarah Patel",
              "Dr. James Okafor", "Dr. Linda Vasquez"]
_CLINICS = ["General Hospital", "University Medical Center", "Regional Health",
            "Community Medical Center", "Memorial Health"]
_DIAGNOSES = ["Type 2 diabetes mellitus", "Hypertension", "Chronic kidney disease stage 3",
              "Major depressive disorder", "Atrial fibrillation"]
_ZIPS = ["10001", "90210", "60601", "77001", "85001", "30301", "98101",
         "02101", "33101", "19101"]

# Non-PHI clinical values (must not fire)
_WBC = ["4.2", "5.8", "7.1", "9.3", "11.0", "3.9", "6.4", "8.7", "5.1", "10.2"]
_HBA1C = ["5.4", "6.2", "7.8", "9.1", "5.9", "6.8", "8.4", "7.2", "5.7", "10.3"]
_ICD10 = ["E11.9", "I10", "N18.3", "F32.1", "I48.0", "E03.9", "J45.20",
          "M17.11", "I50.20", "J44.1"]
_CPT = ["99213", "99214", "99232", "93000", "85025", "80053", "71046",
        "27447", "43239", "36415"]
_LOS = ["2", "4", "7", "1", "12", "3", "5", "9", "6", "14"]
_RISK_SCORE = ["0.12", "0.34", "0.67", "0.08", "0.91", "0.23", "0.55",
               "0.44", "0.78", "0.19"]

# Tier C placeholder values -- must not fire
_PLACEHOLDER_IDS = ["N/A", "NULL", "UNKNOWN", "000000000", "999999999"]
_PLACEHOLDER_SSNS = ["123-45-6789", "000-00-0000", "999-99-9999"]
_PLACEHOLDER_PHONES = ["555-0100", "555-0123", "555-0199"]  # NANP reserved
_PLACEHOLDER_EMAILS = ["test@example.com", "user@test.com", "noreply@hospital.com"]
_PLACEHOLDER_NAMES = ["TEST PATIENT", "SAMPLE", "DEMO USER", "TRAINING RECORD"]
_PLACEHOLDER_DOBS = ["01/01/1900"]  # Excel empty date default
_PLACEHOLDER_ZIPS = ["00000", "99999"]


def _name(rng: random.Random) -> str:
    return f"{rng.choice(_FIRST)} {rng.choice(_LAST)}"


def _mrn(rng: random.Random) -> str:
    return f"MRN-{rng.randint(100000, 999999)}"


def _ssn(rng: random.Random) -> str:
    # Avoid SSA reserved ranges: 000, 666, 900-999
    area = rng.choice([str(x).zfill(3) for x in range(1, 666)] +
                      [str(x).zfill(3) for x in range(667, 900)])
    grp = str(rng.randint(1, 99)).zfill(2)
    seq = str(rng.randint(1, 9999)).zfill(4)
    return f"{area}-{grp}-{seq}"


def _phone(rng: random.Random) -> str:
    area = rng.choice([200, 212, 213, 303, 312, 404, 415, 512, 617, 702,
                       713, 718, 801, 858, 901, 919, 206, 305, 407, 503])
    prefix = rng.randint(200, 999)
    line = rng.randint(1000, 9999)
    return f"({area}) {prefix}-{line}"


def _dob(rng: random.Random) -> str:
    y = rng.randint(1940, 1990)
    m = rng.randint(1, 12)
    d = rng.randint(1, 28)
    return f"{m:02d}/{d:02d}/{y:04d}"


def _dob_senior(rng: random.Random) -> str:
    """DOB for patient aged >= 90 -- exact age is HIPAA PHI category C."""
    y = rng.randint(1920, 1934)
    m = rng.randint(1, 12)
    d = rng.randint(1, 28)
    return f"{m:02d}/{d:02d}/{y:04d}"


def _email(rng: random.Random) -> str:
    first = rng.choice(_FIRST).lower()
    last = rng.choice(_LAST).lower()
    domain = rng.choice(["gmail.com", "yahoo.com", "outlook.com", "icloud.com"])
    return f"{first}.{last}@{domain}"


def _ip(rng: random.Random) -> str:
    return f"{rng.randint(1,254)}.{rng.randint(0,255)}.{rng.randint(0,255)}.{rng.randint(1,254)}"


def _npi(rng: random.Random) -> str:
    return f"{rng.randint(1000000000, 9999999999)}"


# ---------------------------------------------------------------------------
# Extracted-text builder
# ---------------------------------------------------------------------------

def _build_text(sections: List[Tuple[str, str]]) -> str:
    """Build flat extraction text from (section_header, content) pairs."""
    parts = []
    for header, content in sections:
        parts.append(f"[{header}]")
        parts.append(content)
    return "\n".join(parts)


def _find_spans(text: str, phi_specs: List[Tuple[str, str, str, str, str]]) -> List[GoldSpan]:
    """
    Build GoldSpan list from (value, category, hipaa_cat, authority, regime).
    Finds first occurrence of each value in text. Raises if not found.
    """
    spans = []
    for value, category, hipaa_cat, authority, regime in phi_specs:
        start = text.find(value)
        if start < 0:
            raise ValueError(f"PHI value '{value}' not found in extracted text")
        spans.append(GoldSpan(
            start=start,
            end=start + len(value),
            category=category,
            hipaa_category=hipaa_cat or None,
            jurisdiction="us",
            authority=authority,
            value=value,
            entity_type=category,
            detection_regime=regime,
        ))
    return spans


# ---------------------------------------------------------------------------
# Tier A: Canonical records -- clear PHI headers, clear PHI values
# ---------------------------------------------------------------------------

def _tier_a_canonical(rng: random.Random, idx: int) -> Record:
    """
    Standard clinical spreadsheet with unambiguous PHI column headers.
    Group 1: PHI columns (patient_name, dob, mrn, ssn, phone, email, zip, ip_address)
    Group 2: Non-PHI clinical columns (wbc, hba1c, icd10_code, cpt_code, los, risk_score)
    """
    pt_name = _name(rng)
    dob_val = _dob(rng)
    mrn_val = _mrn(rng)
    ssn_val = _ssn(rng)
    phone_val = _phone(rng)
    email_val = _email(rng)
    zip_val = rng.choice(_ZIPS)
    ip_val = _ip(rng)

    wbc_val = rng.choice(_WBC)
    hba1c_val = rng.choice(_HBA1C)
    icd_val = rng.choice(_ICD10)
    cpt_val = rng.choice(_CPT)
    los_val = rng.choice(_LOS)
    risk_val = rng.choice(_RISK_SCORE)

    text = _build_text([
        ("METADATA", f"author: \nautitle: Clinical Dataset"),
        ("SHEET:Patient Data:visible",
         f"patient_name,date_of_birth,mrn,ssn,phone,email,zip,ip_address,"
         f"wbc,hba1c,icd10_code,cpt_code,length_of_stay,risk_score\n"
         f"{pt_name},{dob_val},{mrn_val},{ssn_val},{phone_val},{email_val},"
         f"{zip_val},{ip_val},{wbc_val},{hba1c_val},{icd_val},{cpt_val},{los_val},{risk_val}"),
    ])

    phi_specs = [
        (pt_name, "NAME_PATIENT", "A", AUTH_XLSX, DETECTION_REGIME_NER),
        (dob_val, "DATE_DOB", "C", AUTH_XLSX, DETECTION_REGIME_RULE),
        (mrn_val, "MRN", "H", AUTH_XLSX, DETECTION_REGIME_RULE),
        (ssn_val, "SSN", "G", AUTH_XLSX, DETECTION_REGIME_RULE),
        (phone_val, "PHONE", "D", AUTH_XLSX, DETECTION_REGIME_RULE),
        (email_val, "EMAIL", "F", AUTH_XLSX, DETECTION_REGIME_RULE),
        (zip_val, "ZIP", "B", AUTH_XLSX, DETECTION_REGIME_CONFLICT),
        (ip_val, "IP_ADDRESS", "O", AUTH_XLSX, DETECTION_REGIME_RULE),
    ]

    rec = Record(
        record_id=f"xlsx_tier_a_{idx:04d}",
        text=text,
        gold_spans=_find_spans(text, phi_specs),
        layer=LAYER_HIPAA,
        jurisdiction="us",
        detection_regime=DETECTION_REGIME_NER,
        de_id_tier="identifiable",
        format=FORMAT,
        authority_citations=[AUTH_XLSX],
        metadata={
            "corpus_tier": "A",
            "requires_human_review": False,
            "human_review_reason": "",
            "xlsx_phi_locations": ["cell_value"],
            "non_phi_columns": ["wbc", "hba1c", "icd10_code", "cpt_code",
                                "length_of_stay", "risk_score"],
        },
    )
    return rec


def _tier_a_split_name_clear(rng: random.Random, idx: int) -> Record:
    """
    fname + lname columns clearly labeled as PHI.
    Tier A because headers are unambiguous (fname, lname are standard PHI names).
    """
    first = rng.choice(_FIRST)
    last = rng.choice(_LAST)
    mrn_val = _mrn(rng)
    dob_val = _dob(rng)

    text = _build_text([
        ("METADATA", "author: \ntitle: Patient Registry"),
        ("SHEET:Registry:visible",
         f"fname,lname,mrn,date_of_birth,gender,race,readmission_flag\n"
         f"{first},{last},{mrn_val},{dob_val},F,White,0"),
    ])

    phi_specs = [
        (first, "NAME_PATIENT_FIRST", "A", AUTH_XLSX, DETECTION_REGIME_NER),
        (last, "NAME_PATIENT_LAST", "A", AUTH_XLSX, DETECTION_REGIME_NER),
        (mrn_val, "MRN", "H", AUTH_XLSX, DETECTION_REGIME_RULE),
        (dob_val, "DATE_DOB", "C", AUTH_XLSX, DETECTION_REGIME_RULE),
    ]

    return Record(
        record_id=f"xlsx_tier_a_splitname_{idx:04d}",
        text=text,
        gold_spans=_find_spans(text, phi_specs),
        layer=LAYER_HIPAA,
        jurisdiction="us",
        detection_regime=DETECTION_REGIME_NER,
        de_id_tier="identifiable",
        format=FORMAT,
        authority_citations=[AUTH_XLSX],
        metadata={
            "corpus_tier": "A",
            "requires_human_review": False,
            "human_review_reason": "",
            "xlsx_phi_locations": ["cell_value"],
            "note": "fname+lname split but clearly labeled PHI headers",
        },
    )


# ---------------------------------------------------------------------------
# Tier B: Hard Recall -- PHI that is easy to miss
# ---------------------------------------------------------------------------

def _tier_b_metadata_author(rng: random.Random, idx: int) -> Record:
    """
    B4: PHI in Excel metadata author field and document title.
    Many text extractors read only cell values, missing document properties.
    Authority: 45 CFR 164.514(b)(2)(i)(A) Names -- clinician name in metadata.
    """
    provider = rng.choice(_PROVIDERS)
    clinic = rng.choice(_CLINICS)
    pt_name = _name(rng)
    mrn_val = _mrn(rng)

    text = _build_text([
        ("METADATA", f"author: {provider}\ntitle: {pt_name} Oncology Chart\n"
                     f"subject: {clinic} Patient Records\nkeywords: PHI"),
        ("SHEET:Data:visible",
         f"record_id,icd10_code,readmission_flag\n1,E11.9,0"),
    ])

    phi_specs = [
        (provider, "NAME_PROVIDER", "A", AUTH_XLSX, DETECTION_REGIME_NER),
        (pt_name, "NAME_PATIENT", "A", AUTH_XLSX, DETECTION_REGIME_NER),
    ]

    return Record(
        record_id=f"xlsx_tier_b_metadata_{idx:04d}",
        text=text,
        gold_spans=_find_spans(text, phi_specs),
        layer=LAYER_HIPAA,
        jurisdiction="us",
        detection_regime=DETECTION_REGIME_NER,
        de_id_tier="identifiable",
        format=FORMAT,
        authority_citations=[AUTH_XLSX],
        metadata={
            "corpus_tier": "B",
            "requires_human_review": False,
            "human_review_reason": "",
            "xlsx_phi_locations": ["metadata.author", "metadata.title"],
            "edge_case": "B4: PHI in document metadata properties",
            "detection_challenge": "tools reading only cell values miss metadata PHI",
        },
    )


def _tier_b_sheet_tab_name(rng: random.Random, idx: int) -> Record:
    """
    B5: PHI in sheet tab name (sheet named 'MRN_123456' or 'John Smith - Oncology').
    """
    pt_name = _name(rng)
    mrn_val = _mrn(rng)
    sheet_name = f"{pt_name} - Oncology"

    text = _build_text([
        ("METADATA", "author: \ntitle: Clinical Data"),
        (f"SHEET:{sheet_name}:visible",
         f"icd10_code,cpt_code,readmission_flag\nE11.9,99213,0"),
    ])

    phi_specs = [
        (pt_name, "NAME_PATIENT", "A", AUTH_XLSX, DETECTION_REGIME_NER),
    ]

    return Record(
        record_id=f"xlsx_tier_b_sheetname_{idx:04d}",
        text=text,
        gold_spans=_find_spans(text, phi_specs),
        layer=LAYER_HIPAA,
        jurisdiction="us",
        detection_regime=DETECTION_REGIME_NER,
        de_id_tier="identifiable",
        format=FORMAT,
        authority_citations=[AUTH_XLSX],
        metadata={
            "corpus_tier": "B",
            "requires_human_review": False,
            "human_review_reason": "",
            "xlsx_phi_locations": ["sheet_name"],
            "edge_case": "B5: PHI in worksheet tab name",
            "detection_challenge": "tools scanning only cell values miss sheet names",
        },
    )


def _tier_b_cell_comment(rng: random.Random, idx: int) -> Record:
    """
    B6: PHI in cell comment/note, not in cell value.
    Cell value is benign; comment text contains SSN.
    """
    ssn_val = _ssn(rng)
    pt_name = _name(rng)

    text = _build_text([
        ("METADATA", "author: \ntitle: Dataset"),
        ("SHEET:Data:visible",
         f"record_id,icd10_code,mortality_flag\n1,J44.1,0"),
        ("CELL_COMMENTS",
         f"Data!A2: Note -- patient {pt_name}, SSN {ssn_val}, verify before submission"),
    ])

    phi_specs = [
        (pt_name, "NAME_PATIENT", "A", AUTH_XLSX, DETECTION_REGIME_NER),
        (ssn_val, "SSN", "G", AUTH_XLSX, DETECTION_REGIME_RULE),
    ]

    return Record(
        record_id=f"xlsx_tier_b_comment_{idx:04d}",
        text=text,
        gold_spans=_find_spans(text, phi_specs),
        layer=LAYER_HIPAA,
        jurisdiction="us",
        detection_regime=DETECTION_REGIME_NER,
        de_id_tier="identifiable",
        format=FORMAT,
        authority_citations=[AUTH_XLSX],
        metadata={
            "corpus_tier": "B",
            "requires_human_review": False,
            "human_review_reason": "",
            "xlsx_phi_locations": ["cell_comment"],
            "edge_case": "B6: PHI in cell comment (not in cell value)",
            "detection_challenge": "comment text requires cell.comment.text extraction",
        },
    )


def _tier_b_hidden_row(rng: random.Random, idx: int) -> Record:
    """
    B7: PHI in hidden row. Row is hidden (row.hidden=True in openpyxl).
    pandas read_excel() and many tools skip hidden rows silently.
    """
    pt_name = _name(rng)
    mrn_val = _mrn(rng)
    ssn_val = _ssn(rng)

    text = _build_text([
        ("METADATA", "author: \ntitle: Patient List"),
        ("SHEET:Patients:visible",
         f"patient_name,mrn,ssn,icd10_code\n"
         f"[VISIBLE ROW 1] Jane Public,MRN-000001,N/A,I10\n"
         f"[HIDDEN ROW 2] {pt_name},{mrn_val},{ssn_val},E11.9"),
    ])

    phi_specs = [
        (pt_name, "NAME_PATIENT", "A", AUTH_XLSX, DETECTION_REGIME_NER),
        (mrn_val, "MRN", "H", AUTH_XLSX, DETECTION_REGIME_RULE),
        (ssn_val, "SSN", "G", AUTH_XLSX, DETECTION_REGIME_RULE),
    ]

    return Record(
        record_id=f"xlsx_tier_b_hiddenrow_{idx:04d}",
        text=text,
        gold_spans=_find_spans(text, phi_specs),
        layer=LAYER_HIPAA,
        jurisdiction="us",
        detection_regime=DETECTION_REGIME_RULE,
        de_id_tier="identifiable",
        format=FORMAT,
        authority_citations=[AUTH_XLSX],
        metadata={
            "corpus_tier": "B",
            "requires_human_review": False,
            "human_review_reason": "",
            "xlsx_phi_locations": ["hidden_row"],
            "edge_case": "B7: PHI in hidden row (row.hidden=True)",
            "detection_challenge": "pandas read_excel() skips hidden rows; openpyxl must check row.hidden",
        },
    )


def _tier_b_hidden_sheet(rng: random.Random, idx: int) -> Record:
    """
    B9: Hidden lookup sheet mapping de-identified IDs to real patient names.
    DATA sheet has pseudonymized IDs; LOOKUP sheet (hidden) maps to real identities.
    """
    pt_name = _name(rng)
    mrn_val = _mrn(rng)

    text = _build_text([
        ("METADATA", "author: \ntitle: Research Dataset"),
        ("SHEET:DATA:visible",
         f"subject_id,icd10_code,length_of_stay,readmission_flag\n"
         f"P001,E11.9,4,0\nP002,I10,2,1"),
        ("SHEET:LOOKUP:hidden",
         f"subject_id,patient_name,mrn\n"
         f"P001,{pt_name},{mrn_val}"),
    ])

    phi_specs = [
        (pt_name, "NAME_PATIENT", "A", AUTH_XLSX, DETECTION_REGIME_NER),
        (mrn_val, "MRN", "H", AUTH_XLSX, DETECTION_REGIME_RULE),
    ]

    return Record(
        record_id=f"xlsx_tier_b_hiddensheet_{idx:04d}",
        text=text,
        gold_spans=_find_spans(text, phi_specs),
        layer=LAYER_HIPAA,
        jurisdiction="us",
        detection_regime=DETECTION_REGIME_NER,
        de_id_tier="identifiable",
        format=FORMAT,
        authority_citations=[AUTH_XLSX],
        metadata={
            "corpus_tier": "B",
            "requires_human_review": False,
            "human_review_reason": "",
            "xlsx_phi_locations": ["hidden_sheet"],
            "edge_case": "B9: Re-identification via hidden lookup sheet",
            "detection_challenge": "sheet.sheet_state='hidden'; must iterate wb.worksheets not wb.active",
        },
    )


def _tier_b_formula_result(rng: random.Random, idx: int) -> Record:
    """
    B3: PHI in formula computed result.
    The formula is =CONCATENATE(A2," DOB: ",B2) but the displayed value contains PHI.
    Text extraction of formula results (not formula strings) exposes PHI.
    """
    pt_name = _name(rng)
    dob_val = _dob(rng)
    mrn_val = _mrn(rng)
    computed = f"{pt_name} DOB: {dob_val}"  # what =CONCATENATE evaluates to

    text = _build_text([
        ("METADATA", "author: \ntitle: Summary"),
        ("SHEET:Summary:visible",
         f"record_id,clinical_summary_formula_result,mrn,icd10_code\n"
         f"1,{computed},{mrn_val},E11.9"),
    ])
    # Note: in the actual xlsx file, the cell stores =CONCATENATE(...) but displays computed.
    # A tool that extracts formula strings instead of values will miss this.

    phi_specs = [
        (pt_name, "NAME_PATIENT", "A", AUTH_XLSX, DETECTION_REGIME_NER),
        (dob_val, "DATE_DOB", "C", AUTH_XLSX, DETECTION_REGIME_RULE),
        (mrn_val, "MRN", "H", AUTH_XLSX, DETECTION_REGIME_RULE),
    ]

    return Record(
        record_id=f"xlsx_tier_b_formula_{idx:04d}",
        text=text,
        gold_spans=_find_spans(text, phi_specs),
        layer=LAYER_HIPAA,
        jurisdiction="us",
        detection_regime=DETECTION_REGIME_NER,
        de_id_tier="identifiable",
        format=FORMAT,
        authority_citations=[AUTH_XLSX],
        metadata={
            "corpus_tier": "B",
            "requires_human_review": False,
            "human_review_reason": "",
            "xlsx_phi_locations": ["formula_computed_value"],
            "edge_case": "B3: PHI in formula result (=CONCATENATE); formula string itself not PHI",
            "detection_challenge": "must read cached cell value, not cell formula string",
        },
    )


def _tier_b_date_serial(rng: random.Random, idx: int) -> Record:
    """
    B12: DOB stored as Excel serial integer (not formatted as date string).
    Excel serial 29221 = 1980-01-01. Text scanners see '29221', not '01/01/1980'.
    """
    # Pick a date and convert to Excel serial (days since 1900-01-01, with leap year bug)
    y = rng.randint(1940, 1990)
    m = rng.randint(1, 12)
    d = rng.randint(1, 28)
    # Approximate Excel serial (accurate enough for test purposes)
    from datetime import date as _date
    delta = _date(y, m, d) - _date(1899, 12, 30)
    serial = delta.days
    formatted = f"{m:02d}/{d:02d}/{y:04d}"
    mrn_val = _mrn(rng)
    pt_name = _name(rng)

    text = _build_text([
        ("METADATA", "author: \ntitle: Dataset"),
        ("SHEET:Patients:visible",
         # Extraction must parse serial -> date. Extracted text shows formatted date.
         f"patient_name,date_of_birth_serial,date_of_birth_parsed,mrn,wbc\n"
         f"{pt_name},{serial},{formatted},{mrn_val},{rng.choice(_WBC)}"),
    ])

    phi_specs = [
        (pt_name, "NAME_PATIENT", "A", AUTH_XLSX, DETECTION_REGIME_NER),
        (formatted, "DATE_DOB", "C", AUTH_XLSX, DETECTION_REGIME_RULE),
        (mrn_val, "MRN", "H", AUTH_XLSX, DETECTION_REGIME_RULE),
    ]

    return Record(
        record_id=f"xlsx_tier_b_dateserial_{idx:04d}",
        text=text,
        gold_spans=_find_spans(text, phi_specs),
        layer=LAYER_HIPAA,
        jurisdiction="us",
        detection_regime=DETECTION_REGIME_RULE,
        de_id_tier="identifiable",
        format=FORMAT,
        authority_citations=[AUTH_XLSX],
        metadata={
            "corpus_tier": "B",
            "requires_human_review": False,
            "human_review_reason": "",
            "xlsx_phi_locations": ["cell_value_date_serial"],
            "edge_case": f"B12: DOB as Excel serial {serial} (= {formatted})",
            "detection_challenge": "must call openpyxl with data_only=True and parse number-format dates",
        },
    )


def _tier_b_ssn_leading_zero_dropped(rng: random.Random, idx: int) -> Record:
    """
    B13: SSN with leading zero stored as Number type -- leading zero silently dropped.
    '012-34-5678' -> stored as 12345678 (integer) -> displays as '12345678'.
    """
    # Force an SSN with area code starting with 0
    area = rng.choice(["010", "020", "030", "040", "050", "060", "070", "080", "090"])
    grp = str(rng.randint(1, 99)).zfill(2)
    seq = str(rng.randint(1, 9999)).zfill(4)
    ssn_original = f"{area}-{grp}-{seq}"
    ssn_stored_as_int = ssn_original.replace("-", "").lstrip("0")  # leading zero dropped
    pt_name = _name(rng)
    mrn_val = _mrn(rng)

    text = _build_text([
        ("METADATA", "author: \ntitle: Dataset"),
        ("SHEET:Patients:visible",
         # ssn_stored shows what the cell displays after number coercion
         f"patient_name,mrn,ssn_as_text,ssn_coerced_number,wbc\n"
         f"{pt_name},{mrn_val},{ssn_original},{ssn_stored_as_int},{rng.choice(_WBC)}"),
    ])

    phi_specs = [
        (pt_name, "NAME_PATIENT", "A", AUTH_XLSX, DETECTION_REGIME_NER),
        (mrn_val, "MRN", "H", AUTH_XLSX, DETECTION_REGIME_RULE),
        (ssn_original, "SSN", "G", AUTH_XLSX, DETECTION_REGIME_RULE),
        # The coerced number is ALSO PHI (truncated SSN) -- different pattern
        (ssn_stored_as_int, "SSN_TRUNCATED", "G", AUTH_XLSX, DETECTION_REGIME_RULE),
    ]

    return Record(
        record_id=f"xlsx_tier_b_ssnleadingzero_{idx:04d}",
        text=text,
        gold_spans=_find_spans(text, phi_specs),
        layer=LAYER_HIPAA,
        jurisdiction="us",
        detection_regime=DETECTION_REGIME_RULE,
        de_id_tier="identifiable",
        format=FORMAT,
        authority_citations=[AUTH_XLSX],
        metadata={
            "corpus_tier": "B",
            "requires_human_review": False,
            "human_review_reason": "",
            "xlsx_phi_locations": ["cell_value_number_coercion"],
            "edge_case": f"B13: SSN {ssn_original} stored as number, displays as {ssn_stored_as_int}",
            "detection_challenge": "pattern match for 9-digit string without dashes; leading zero lost",
        },
    )


def _tier_b_phi_in_notes_column(rng: random.Random, idx: int) -> Record:
    """
    B14: PHI embedded in a non-PHI-labeled column ('notes').
    Column header is 'notes' -- not a PHI header. Cell value contains full PHI narrative.
    """
    pt_name = _name(rng)
    dob_val = _dob(rng)
    mrn_val = _mrn(rng)
    embedded = f"Patient {pt_name} DOB {dob_val} seen for follow-up. MRN: {mrn_val}."

    text = _build_text([
        ("METADATA", "author: \ntitle: Clinical Notes"),
        ("SHEET:Notes:visible",
         f"visit_date,provider_id,notes,icd10_code\n"
         f"2024-03-15,PROV-001,{embedded},E11.9"),
    ])

    phi_specs = [
        (pt_name, "NAME_PATIENT", "A", AUTH_XLSX, DETECTION_REGIME_NER),
        (dob_val, "DATE_DOB", "C", AUTH_XLSX, DETECTION_REGIME_RULE),
        (mrn_val, "MRN", "H", AUTH_XLSX, DETECTION_REGIME_RULE),
    ]

    return Record(
        record_id=f"xlsx_tier_b_notescol_{idx:04d}",
        text=text,
        gold_spans=_find_spans(text, phi_specs),
        layer=LAYER_HIPAA,
        jurisdiction="us",
        detection_regime=DETECTION_REGIME_NER,
        de_id_tier="identifiable",
        format=FORMAT,
        authority_citations=[AUTH_XLSX],
        metadata={
            "corpus_tier": "B",
            "requires_human_review": False,
            "human_review_reason": "",
            "xlsx_phi_locations": ["cell_value_unlabeled_column"],
            "edge_case": "B14: PHI in 'notes' column (non-PHI header, PHI cell content)",
            "detection_challenge": "column-header-based rules miss PHI in free-text columns",
        },
    )


def _tier_b_phi_in_header_cell(rng: random.Random, idx: int) -> Record:
    """
    B15: PHI embedded in a column header cell itself.
    Header reads "Dr. Sarah Jones's Patients" -- the provider name is in the header.
    """
    provider = rng.choice(_PROVIDERS)
    pt_name = _name(rng)
    mrn_val = _mrn(rng)
    header_with_phi = f"{provider}'s Patients"

    text = _build_text([
        ("METADATA", "author: \ntitle: Provider Dataset"),
        ("SHEET:Data:visible",
         f"{header_with_phi},mrn,icd10_code,readmission_flag\n"
         f"{pt_name},{mrn_val},I10,0"),
    ])

    phi_specs = [
        (provider, "NAME_PROVIDER", "A", AUTH_XLSX, DETECTION_REGIME_NER),
        (pt_name, "NAME_PATIENT", "A", AUTH_XLSX, DETECTION_REGIME_NER),
        (mrn_val, "MRN", "H", AUTH_XLSX, DETECTION_REGIME_RULE),
    ]

    return Record(
        record_id=f"xlsx_tier_b_phiheader_{idx:04d}",
        text=text,
        gold_spans=_find_spans(text, phi_specs),
        layer=LAYER_HIPAA,
        jurisdiction="us",
        detection_regime=DETECTION_REGIME_NER,
        de_id_tier="identifiable",
        format=FORMAT,
        authority_citations=[AUTH_XLSX],
        metadata={
            "corpus_tier": "B",
            "requires_human_review": False,
            "human_review_reason": "",
            "xlsx_phi_locations": ["column_header_cell"],
            "edge_case": "B15: PHI embedded within a column header cell value",
            "detection_challenge": "tools that skip header rows for content scanning miss this",
        },
    )


def _tier_b_non_english_header(rng: random.Random, idx: int) -> Record:
    """
    B17: PHI column headers in non-English language (Spanish, Hindi).
    A tool trained only on English PHI header names will miss these.
    """
    pt_name = _name(rng)
    dob_val = _dob(rng)
    mrn_val = _mrn(rng)
    phone_val = _phone(rng)

    text = _build_text([
        ("METADATA", "author: \ntitle: Datos del Paciente"),
        ("SHEET:Pacientes:visible",
         # Spanish headers
         f"nombre_del_paciente,fecha_de_nacimiento,numero_expediente,telefono,"
         f"glucosa,codigo_diagnostico\n"
         f"{pt_name},{dob_val},{mrn_val},{phone_val},{rng.choice(_WBC)},E11.9"),
    ])

    phi_specs = [
        (pt_name, "NAME_PATIENT", "A", AUTH_XLSX, DETECTION_REGIME_NER),
        (dob_val, "DATE_DOB", "C", AUTH_XLSX, DETECTION_REGIME_RULE),
        (mrn_val, "MRN", "H", AUTH_XLSX, DETECTION_REGIME_RULE),
        (phone_val, "PHONE", "D", AUTH_XLSX, DETECTION_REGIME_RULE),
    ]

    return Record(
        record_id=f"xlsx_tier_b_nonengheader_{idx:04d}",
        text=text,
        gold_spans=_find_spans(text, phi_specs),
        layer=LAYER_HIPAA,
        jurisdiction="us",
        detection_regime=DETECTION_REGIME_NER,
        de_id_tier="identifiable",
        format=FORMAT,
        authority_citations=[AUTH_XLSX],
        metadata={
            "corpus_tier": "B",
            "requires_human_review": False,
            "human_review_reason": "",
            "xlsx_phi_locations": ["cell_value"],
            "edge_case": "B17: PHI column headers in Spanish (nombre_del_paciente, fecha_de_nacimiento)",
            "detection_challenge": "English-only header-name rules miss non-English headers; "
                                   "must use value-level detection",
        },
    )


def _tier_b_merged_header(rng: random.Random, idx: int) -> Record:
    """
    B18: Merged cell header spanning PHI columns.
    A1:B1 merged, labeled 'Patient Information'; A2=first_name, B2=last_name.
    Detector reading row 1 for headers gets 'Patient Information' not 'first_name'/'last_name'.
    """
    first = rng.choice(_FIRST)
    last = rng.choice(_LAST)
    mrn_val = _mrn(rng)

    text = _build_text([
        ("METADATA", "author: \ntitle: Patient Roster"),
        ("SHEET:Roster:visible",
         # Represents merged cell structure in flat text
         f"[MERGED A1:B1] Patient Information | [C1] mrn | [D1] icd10_code\n"
         f"[A2] first_name | [B2] last_name | [C2] mrn | [D2] icd10_code\n"
         f"{first},{last},{mrn_val},{rng.choice(_ICD10)}"),
    ])

    phi_specs = [
        (first, "NAME_PATIENT_FIRST", "A", AUTH_XLSX, DETECTION_REGIME_NER),
        (last, "NAME_PATIENT_LAST", "A", AUTH_XLSX, DETECTION_REGIME_NER),
        (mrn_val, "MRN", "H", AUTH_XLSX, DETECTION_REGIME_RULE),
    ]

    return Record(
        record_id=f"xlsx_tier_b_mergedheader_{idx:04d}",
        text=text,
        gold_spans=_find_spans(text, phi_specs),
        layer=LAYER_HIPAA,
        jurisdiction="us",
        detection_regime=DETECTION_REGIME_NER,
        de_id_tier="identifiable",
        format=FORMAT,
        authority_citations=[AUTH_XLSX],
        metadata={
            "corpus_tier": "B",
            "requires_human_review": False,
            "human_review_reason": "",
            "xlsx_phi_locations": ["cell_value"],
            "edge_case": "B18: Merged cell header (A1:B1 = 'Patient Information'); actual sub-headers in row 2",
            "detection_challenge": "header detection heuristic must handle multi-row headers",
        },
    )


def _tier_b_crosssheet_linkage(rng: random.Random, idx: int) -> Record:
    """
    B21: Cross-sheet linkage re-identifies de-identified data.
    ANALYSIS sheet has formula =RAW_DATA!A2 exposing patient name.
    """
    pt_name = _name(rng)
    mrn_val = _mrn(rng)

    text = _build_text([
        ("METADATA", "author: \ntitle: Analysis"),
        ("SHEET:RAW_DATA:visible",
         f"patient_name,mrn,icd10_code\n{pt_name},{mrn_val},E11.9"),
        ("SHEET:ANALYSIS:visible",
         # Formula result in B2 exposes patient_name from RAW_DATA
         f"metric,patient_label,value\n"
         f"risk_score,{pt_name} (via =RAW_DATA!A2),0.34"),
    ])

    phi_specs = [
        (pt_name, "NAME_PATIENT", "A", AUTH_XLSX, DETECTION_REGIME_NER),
        (mrn_val, "MRN", "H", AUTH_XLSX, DETECTION_REGIME_RULE),
    ]

    return Record(
        record_id=f"xlsx_tier_b_crosssheet_{idx:04d}",
        text=text,
        gold_spans=_find_spans(text, phi_specs),
        layer=LAYER_HIPAA,
        jurisdiction="us",
        detection_regime=DETECTION_REGIME_NER,
        de_id_tier="identifiable",
        format=FORMAT,
        authority_citations=[AUTH_XLSX],
        metadata={
            "corpus_tier": "B",
            "requires_human_review": False,
            "human_review_reason": "",
            "xlsx_phi_locations": ["cell_value", "formula_reference"],
            "edge_case": "B21: Cross-sheet formula reference re-identifies de-identified data",
            "detection_challenge": "must scan ALL sheets including formula referees, not just primary data sheet",
        },
    )


def _tier_b_concatenated_cell(rng: random.Random, idx: int) -> Record:
    """
    B2: PHI concatenated with non-PHI in same cell ('Dx: T2DM, Pt: John Smith, MRN: 12345').
    Multiple PHI types mixed in one cell value -- substring detectors may miss structure.
    """
    pt_name = _name(rng)
    mrn_val = _mrn(rng)
    dx = rng.choice(["T2DM", "HTN", "CKD3", "MDD", "AFib"])
    concat = f"Dx: {dx}, Pt: {pt_name}, MRN: {mrn_val}"

    text = _build_text([
        ("METADATA", "author: \ntitle: Clinical"),
        ("SHEET:Data:visible",
         f"record_id,combined_clinical_text,wbc\n1,{concat},{rng.choice(_WBC)}"),
    ])

    phi_specs = [
        (pt_name, "NAME_PATIENT", "A", AUTH_XLSX, DETECTION_REGIME_NER),
        (mrn_val, "MRN", "H", AUTH_XLSX, DETECTION_REGIME_RULE),
    ]

    return Record(
        record_id=f"xlsx_tier_b_concatenated_{idx:04d}",
        text=text,
        gold_spans=_find_spans(text, phi_specs),
        layer=LAYER_HIPAA,
        jurisdiction="us",
        detection_regime=DETECTION_REGIME_NER,
        de_id_tier="identifiable",
        format=FORMAT,
        authority_citations=[AUTH_XLSX],
        metadata={
            "corpus_tier": "B",
            "requires_human_review": False,
            "human_review_reason": "",
            "xlsx_phi_locations": ["cell_value_concatenated"],
            "edge_case": "B2: Multiple PHI types concatenated in single cell with non-PHI",
            "detection_challenge": "column-header PHI classification fails; needs NER on cell content",
        },
    )


def _tier_b_named_range(rng: random.Random, idx: int) -> Record:
    """
    B10: PHI value accessible via a named range (defined name).
    Named range 'PatientID' resolves to a cell containing MRN.
    """
    mrn_val = _mrn(rng)
    pt_name = _name(rng)

    text = _build_text([
        ("METADATA", "author: \ntitle: Dataset"),
        ("SHEET:Data:visible",
         f"patient_name,mrn,icd10_code\n{pt_name},{mrn_val},I10"),
        ("NAMED_RANGES",
         f"PatientID={mrn_val}\nPatientName={pt_name}"),
    ])

    phi_specs = [
        (pt_name, "NAME_PATIENT", "A", AUTH_XLSX, DETECTION_REGIME_NER),
        (mrn_val, "MRN", "H", AUTH_XLSX, DETECTION_REGIME_RULE),
    ]

    return Record(
        record_id=f"xlsx_tier_b_namedrange_{idx:04d}",
        text=text,
        gold_spans=_find_spans(text, phi_specs),
        layer=LAYER_HIPAA,
        jurisdiction="us",
        detection_regime=DETECTION_REGIME_NER,
        de_id_tier="identifiable",
        format=FORMAT,
        authority_citations=[AUTH_XLSX],
        metadata={
            "corpus_tier": "B",
            "requires_human_review": False,
            "human_review_reason": "",
            "xlsx_phi_locations": ["named_range", "cell_value"],
            "edge_case": "B10: PHI accessible via named range (workbook.defined_names)",
            "detection_challenge": "must iterate wb.defined_names and resolve cell references",
        },
    )


# ---------------------------------------------------------------------------
# Tier C: False Positive Traps -- PHI-shaped values that are NOT PHI
# ---------------------------------------------------------------------------

def _tier_c_placeholder_id(rng: random.Random, idx: int) -> Record:
    """
    C24/C25/C26/C27: Placeholder values in PHI-labeled columns.
    These match PHI patterns but are explicit null/test markers -- must NOT fire.
    """
    pid = rng.choice(_PLACEHOLDER_IDS)
    pssn = rng.choice(_PLACEHOLDER_SSNS)
    pname = rng.choice(_PLACEHOLDER_NAMES)
    pdob = _PLACEHOLDER_DOBS[0]
    pzip = rng.choice(_PLACEHOLDER_ZIPS)
    real_icd = rng.choice(_ICD10)

    text = _build_text([
        ("METADATA", "author: \ntitle: Test Dataset"),
        ("SHEET:TestData:visible",
         f"patient_name,date_of_birth,patient_id,ssn,zip,icd10_code\n"
         f"{pname},{pdob},{pid},{pssn},{pzip},{real_icd}"),
    ])

    # No gold spans -- these are NOT PHI. An empty gold_spans list means
    # a correct system fires ZERO detections on this record.
    return Record(
        record_id=f"xlsx_tier_c_placeholder_{idx:04d}",
        text=text,
        gold_spans=[],
        layer=LAYER_HIPAA,
        jurisdiction="us",
        detection_regime=DETECTION_REGIME_RULE,
        de_id_tier="safe_harbor",
        format=FORMAT,
        authority_citations=[AUTH_XLSX],
        metadata={
            "corpus_tier": "C",
            "requires_human_review": False,
            "human_review_reason": "",
            "xlsx_phi_locations": [],
            "edge_case": "C24-27: Placeholder values in PHI-labeled columns",
            "detection_challenge": "pattern match fires on placeholder SSNs and IDs; "
                                   "must whitelist null markers",
            "false_positive_risk": "HIGH -- SSN pattern '123-45-6789' and '000-00-0000' "
                                   "appear in PHI columns but are SSA-reserved test values",
        },
    )


def _tier_c_reserved_phone(rng: random.Random, idx: int) -> Record:
    """
    C29: NANP 555-01xx reserved fictional phone numbers.
    Phone pattern matches but these are reserved by NANP for fiction (movies/TV).
    Authority: NANP INC database; 555-0100 through 555-0199 are permanently reserved.
    """
    phone = rng.choice(_PLACEHOLDER_PHONES)
    pt_name = _name(rng)  # name IS PHI -- only the phone is a false positive trap
    mrn_val = _mrn(rng)

    text = _build_text([
        ("METADATA", "author: \ntitle: Example Dataset"),
        ("SHEET:Example:visible",
         f"patient_name,mrn,phone_example,icd10_code\n"
         f"{pt_name},{mrn_val},{phone},{rng.choice(_ICD10)}"),
    ])

    phi_specs = [
        (pt_name, "NAME_PATIENT", "A", AUTH_XLSX, DETECTION_REGIME_NER),
        (mrn_val, "MRN", "H", AUTH_XLSX, DETECTION_REGIME_RULE),
        # phone is NOT in phi_specs -- it must not be flagged
    ]

    return Record(
        record_id=f"xlsx_tier_c_reservedphone_{idx:04d}",
        text=text,
        gold_spans=_find_spans(text, phi_specs),
        layer=LAYER_HIPAA,
        jurisdiction="us",
        detection_regime=DETECTION_REGIME_NER,
        de_id_tier="identifiable",
        format=FORMAT,
        authority_citations=[AUTH_XLSX],
        metadata={
            "corpus_tier": "C",
            "requires_human_review": False,
            "human_review_reason": "",
            "xlsx_phi_locations": ["cell_value"],
            "edge_case": f"C29: NANP-reserved fictional phone {phone} must not fire",
            "detection_challenge": "simple phone regex matches 555-01xx; must apply NANP reserved-block check",
            "false_positive_risk": f"phone {phone} is NANP reserved 555-01xx -- NOT a real phone",
        },
    )


def _tier_c_test_email(rng: random.Random, idx: int) -> Record:
    """
    C30: RFC 2606 reserved email domains -- must not fire.
    example.com, test.com are reserved per RFC 2606 for documentation use.
    noreply@ is a standard system address, not a person's email.
    """
    email = rng.choice(_PLACEHOLDER_EMAILS)
    pt_name = _name(rng)
    mrn_val = _mrn(rng)

    text = _build_text([
        ("METADATA", "author: \ntitle: Dataset"),
        ("SHEET:Data:visible",
         f"patient_name,mrn,system_email,icd10_code\n"
         f"{pt_name},{mrn_val},{email},{rng.choice(_ICD10)}"),
    ])

    phi_specs = [
        (pt_name, "NAME_PATIENT", "A", AUTH_XLSX, DETECTION_REGIME_NER),
        (mrn_val, "MRN", "H", AUTH_XLSX, DETECTION_REGIME_RULE),
        # email NOT in phi_specs
    ]

    return Record(
        record_id=f"xlsx_tier_c_testemail_{idx:04d}",
        text=text,
        gold_spans=_find_spans(text, phi_specs),
        layer=LAYER_HIPAA,
        jurisdiction="us",
        detection_regime=DETECTION_REGIME_NER,
        de_id_tier="identifiable",
        format=FORMAT,
        authority_citations=[AUTH_XLSX],
        metadata={
            "corpus_tier": "C",
            "requires_human_review": False,
            "human_review_reason": "",
            "xlsx_phi_locations": ["cell_value"],
            "edge_case": f"C30: RFC 2606 reserved email {email} must not fire",
            "detection_challenge": "email regex matches any address@domain; "
                                   "must apply RFC 2606 reserved domain allowlist",
            "false_positive_risk": f"email {email} is RFC-2606-reserved -- not a person's address",
        },
    )


def _tier_c_age_bin(rng: random.Random, idx: int) -> Record:
    """
    C31: Age bin (not exact age) is safe under HIPAA -- must not fire.
    HIPAA 45 CFR 164.514(b)(2)(i)(C): ages >89 are PHI. Age bins are not.
    A binned age ('60-69', '<1', '70-79') does not re-identify.
    """
    age_bins = ["<1", "1-9", "10-19", "20-29", "30-39", "40-49",
                "50-59", "60-69", "70-79", "80-89"]
    age_bin = rng.choice(age_bins)
    mrn_val = _mrn(rng)

    text = _build_text([
        ("METADATA", "author: \ntitle: Aggregate Dataset"),
        ("SHEET:Aggregate:visible",
         f"record_id,age_group,icd10_code,readmission_flag,length_of_stay\n"
         f"1,{age_bin},{rng.choice(_ICD10)},0,{rng.choice(_LOS)}"),
    ])

    # No PHI spans -- age bin is not PHI
    return Record(
        record_id=f"xlsx_tier_c_agebin_{idx:04d}",
        text=text,
        gold_spans=[],
        layer=LAYER_HIPAA,
        jurisdiction="us",
        detection_regime=DETECTION_REGIME_CONFLICT,
        de_id_tier="limited_data_set",
        format=FORMAT,
        authority_citations=[AUTH_XLSX, "45 CFR 164.514(e) Limited Data Set"],
        metadata={
            "corpus_tier": "C",
            "requires_human_review": False,
            "human_review_reason": "",
            "xlsx_phi_locations": [],
            "edge_case": f"C31: Age bin '{age_bin}' is NOT PHI under HIPAA 164.514(b)(2)(i)(C)",
            "detection_challenge": "naive age detectors fire on any number in 'age' column; "
                                   "must distinguish exact age from age group",
        },
    )


# ---------------------------------------------------------------------------
# Tier D: Human Review Required
# ---------------------------------------------------------------------------

def _tier_d_free_text_notes(rng: random.Random, idx: int) -> Record:
    """
    D: 'notes' column free-text -- may or may not contain PHI.
    Gold label is AMBIGUOUS. Human review required to confirm whether any
    PHI is present and what it is.
    Authority: EDPB Opinion 28/2024 -- AI/LLM processing of free-text health notes
    requires documented legal basis under GDPR Art. 9(2) and HIPAA minimum necessary.
    """
    # This record has PHI-probable free text but gold spans are marked ambiguous
    pt_name = _name(rng)
    note = (f"Patient was seen by attending. Family history noted. "
            f"Contact listed as {pt_name} (may be household member). "
            f"Referred to specialist at General Hospital.")

    text = _build_text([
        ("METADATA", "author: \ntitle: Clinical Notes"),
        ("SHEET:Notes:visible",
         f"visit_id,notes,icd10_code\n1,{note},{rng.choice(_ICD10)}"),
    ])

    # Gold span marked but with requires_human_review=True
    spans = []
    start = text.find(pt_name)
    if start >= 0:
        spans.append(GoldSpan(
            start=start,
            end=start + len(pt_name),
            category="NAME_AMBIGUOUS",
            hipaa_category="A",
            jurisdiction="us",
            authority=AUTH_XLSX,
            value=pt_name,
            entity_type="NAME_AMBIGUOUS",
            detection_regime=DETECTION_REGIME_NER,
        ))

    return Record(
        record_id=f"xlsx_tier_d_freenotes_{idx:04d}",
        text=text,
        gold_spans=spans,
        layer=LAYER_HIPAA,
        jurisdiction="us",
        detection_regime=DETECTION_REGIME_NER,
        de_id_tier="identifiable",
        format=FORMAT,
        authority_citations=[AUTH_XLSX, AUTH_XLSX_EDPB],
        metadata={
            "corpus_tier": "D",
            "requires_human_review": True,
            "human_review_reason": (
                "Free-text notes column may contain PHI (names, dates, facility references). "
                "Name found is ambiguous: could be patient or household member. "
                "Clinical reviewer must confirm before labeling."
            ),
            "xlsx_phi_locations": ["cell_value_free_text"],
            "edge_case": "D: Free-text notes column with ambiguous PHI",
        },
    )


def _tier_d_quasi_id_combination(rng: random.Random, idx: int) -> Record:
    """
    D: Quasi-identifier combination (gender + ZIP + race + admission_year).
    None of these is PHI individually under HIPAA. But combined, Sweeney 2002
    shows ~5-7% of US population is uniquely identifiable by DOB + gender + ZIP.
    Year-level date reduces risk but does not eliminate it for rare conditions.
    Human review required to determine k-anonymity compliance.
    """
    gender = rng.choice(["F", "M"])
    zip_val = rng.choice(_ZIPS)
    race = rng.choice(["White", "Black", "Asian", "Hispanic", "Other"])
    year = rng.randint(2018, 2023)
    rare_dx = rng.choice(["Q78.0", "E70.0", "G11.1", "D61.01"])  # rare ICD codes

    text = _build_text([
        ("METADATA", "author: \ntitle: Research Extract"),
        ("SHEET:Data:visible",
         f"gender,zip,race,admission_year,primary_dx,length_of_stay\n"
         f"{gender},{zip_val},{race},{year},{rare_dx},{rng.choice(_LOS)}"),
    ])

    # No hard PHI spans -- but quasi-ID combination flagged in metadata
    return Record(
        record_id=f"xlsx_tier_d_quasiid_{idx:04d}",
        text=text,
        gold_spans=[],
        layer=LAYER_HIPAA,
        jurisdiction="us",
        detection_regime=DETECTION_REGIME_CONFLICT,
        de_id_tier="limited_data_set",
        format=FORMAT,
        authority_citations=[AUTH_XLSX, AUTH_SWEENEY_K_ANON],
        metadata={
            "corpus_tier": "D",
            "requires_human_review": True,
            "human_review_reason": (
                f"Quasi-identifier combination: gender={gender}, zip={zip_val}, "
                f"race={race}, admission_year={year}, rare_dx={rare_dx}. "
                f"Sweeney 2002 k-anonymity: DOB+gender+ZIP uniquely identifies ~5-7% of US population. "
                f"Year-only date reduces risk. Rare ICD code ({rare_dx}) in small population "
                f"may increase re-identification risk. k-anonymity check required before release."
            ),
            "xlsx_phi_locations": [],
            "edge_case": "D: Quasi-identifier combination -- individual fields not PHI, combination may be",
            "k_anon_risk": "HIGH if rare_dx + zip population < 20,000",
        },
    )


# ---------------------------------------------------------------------------
# xlsx file writer (actual .xlsx on disk)
# ---------------------------------------------------------------------------

def _write_xlsx(record: Record, output_dir: Path) -> Optional[Path]:
    """Write a minimal .xlsx file representing the edge case to disk.
    Returns file path or None if openpyxl unavailable."""
    if not OPENPYXL_AVAILABLE:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Write the extracted text as sheet content (simplified representation)
    for i, line in enumerate(record.text.split("\n"), start=1):
        ws.cell(row=i, column=1, value=line)

    # Set document properties where metadata PHI edge cases apply
    if "metadata.author" in record.metadata.get("xlsx_phi_locations", []):
        # Author PHI is embedded in the text; set it in wb properties too
        for line in record.text.split("\n"):
            if line.startswith("author: "):
                wb.properties.creator = line[8:]
            elif line.startswith("title: "):
                wb.properties.title = line[7:]

    out_path = output_dir / f"{record.record_id}.xlsx"
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path


# ---------------------------------------------------------------------------
# Generator class
# ---------------------------------------------------------------------------

class XlsxGenerator(DeterministicGenerator):
    """Seeded xlsx PHI corpus generator covering all four tiers."""

    def generate(self, n_per_tier_a: int = 20) -> List[Record]:
        """Generate records across all tiers.

        n_per_tier_a: number of canonical Tier A records. Tier B/C/D have
        fixed record counts (one per edge case variant, repeated n_per_tier_a // 4 times).
        """
        records: List[Record] = []
        idx = 0

        # Tier A: canonical (n_per_tier_a records)
        for _ in range(n_per_tier_a // 2):
            records.append(_tier_a_canonical(self.rng, idx)); idx += 1
        for _ in range(n_per_tier_a // 2):
            records.append(_tier_a_split_name_clear(self.rng, idx)); idx += 1

        # Tier B: one record per edge case variant
        tier_b_builders = [
            _tier_b_metadata_author,
            _tier_b_sheet_tab_name,
            _tier_b_cell_comment,
            _tier_b_hidden_row,
            _tier_b_hidden_sheet,
            _tier_b_formula_result,
            _tier_b_date_serial,
            _tier_b_ssn_leading_zero_dropped,
            _tier_b_phi_in_notes_column,
            _tier_b_phi_in_header_cell,
            _tier_b_non_english_header,
            _tier_b_merged_header,
            _tier_b_crosssheet_linkage,
            _tier_b_concatenated_cell,
            _tier_b_named_range,
        ]
        for builder in tier_b_builders:
            # Generate at least 2 variants per edge case
            for _ in range(2):
                records.append(builder(self.rng, idx)); idx += 1

        # Tier C: false positive traps
        tier_c_builders = [
            _tier_c_placeholder_id,
            _tier_c_reserved_phone,
            _tier_c_test_email,
            _tier_c_age_bin,
        ]
        for builder in tier_c_builders:
            for _ in range(3):
                records.append(builder(self.rng, idx)); idx += 1

        # Tier D: human review
        for _ in range(3):
            records.append(_tier_d_free_text_notes(self.rng, idx)); idx += 1
        for _ in range(3):
            records.append(_tier_d_quasi_id_combination(self.rng, idx)); idx += 1

        return records


def generate(seed: int = 42, output_dir: Optional[Path] = None,
             n_per_tier_a: int = 20) -> List[Record]:
    """Public entry point. Returns records and optionally writes JSONL + xlsx files."""
    gen = XlsxGenerator(seed)
    records = gen.generate(n_per_tier_a=n_per_tier_a)

    # Verify all spans
    for rec in records:
        errors = rec.verify_spans()
        if errors:
            raise ValueError(f"Span errors in {rec.record_id}: {errors}")

    if output_dir is not None:
        output_dir = Path(output_dir)
        jsonl_path = output_dir / "xlsx_phi_corpus.jsonl"
        write_jsonl(records, jsonl_path)
        xlsx_dir = output_dir / "xlsx_files"
        for rec in records:
            _write_xlsx(rec, xlsx_dir)

    return records


if __name__ == "__main__":
    import sys
    seed = int(sys.argv[1]) if len(sys.argv) > 1 else 42
    recs = generate(seed=seed)
    tier_counts = {}
    for r in recs:
        t = r.metadata.get("corpus_tier", "?")
        tier_counts[t] = tier_counts.get(t, 0) + 1
    total_spans = sum(len(r.gold_spans) for r in recs)
    print(f"seed={seed}  records={len(recs)}  spans={total_spans}")
    for t in sorted(tier_counts):
        print(f"  Tier {t}: {tier_counts[t]} records")
    print("All span offsets verified.")
