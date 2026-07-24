"""Local, support-only AI boundary for deriving an intake study name.

**What.** Given a verified :class:`~phi_engine.pipeline.intake_preflight.IntakePreflight`
result, decide the study name for an intake run: the user-supplied name, an
AI-inferred name derived *only* from support-component evidence (``forms``,
``dictionary_mapping`` -- never ``datasets`` or ``_unclassified``),
or a random fallback when neither is available.

**Why.** Dataset content must never reach an LLM, even a local one. This
module is the *only* place in the standalone pipeline that is allowed to
open a source file for AI evidence, and it does so exclusively through
:func:`~phi_engine.pipeline.verified_source.open_verified_source` with a
required, re-verified ``source_component`` and identity -- the structural
proof that a dataset artifact (or anything preflight left ``_unclassified``)
can never become AI evidence, by construction rather than by convention.
A post-preflight ``datasets/`` hardlink is an additional, independent TOCTOU
threat this identity check alone cannot see (link count and bytes/inode do
not change): every naming candidate is re-checked against a fresh
descriptor-relative scan of ``datasets/`` immediately before its content is
parsed, and again immediately before any evidence is dispatched to the
local model, discarding all collected evidence and dispatching nothing the
moment such an alias is found. Identity alone still cannot see a same-
inode, same-size dataset content swap whose mtime is restored to its
preflight-recorded value, nor a lexically-distinct, byte-identical dataset
COPY landing in a support component -- neither shares an inode with
anything identity comparison would flag. A second, independent layer
closes both gaps: immediately before any support evidence is parsed, and
again immediately before every individual local-model dispatch, EVERY
preflight-known ``datasets/`` candidate is freshly re-opened and re-
hashed against its preflight-recorded SHA-256; any drift aborts the whole
attempt, and any support candidate whose identity or SHA-256 already
matches a freshly verified dataset is silently excluded from admission,
causing zero parser calls, not merely zero dispatch. The same all-or-
nothing rule applies to any OTHER symlink/hardlink/identity/hash safety
finding, whether preflight
already recorded it (a support file it quarantined to ``_unclassified`` for
``cross-component-hardlink``/``source-symlink-not-allowed`` before naming
ever ran -- checked first, before any candidate is even opened) or naming
discovers it itself while validating an admitted candidate (a descriptor
identity/hash mismatch, or a naming-time symlink rejection): the entire
naming attempt aborts with zero further reads and zero dispatch, not just
exclusion of the one affected candidate, while the specific fixed
review/error record for that finding is preserved.

**How.** Every admitted candidate is validated in a complete first pass --
descriptor open, expected identity, a fresh ``datasets/``-alias rescan,
and a bounded SHA-256 hash, each through
:func:`~phi_engine.pipeline.verified_source.open_verified_source` -- before
ANY candidate's content is parsed in the second pass; a later candidate's
safety failure is never discoverable only after an earlier, otherwise-
clean candidate was already parsed. The second pass performs exactly ONE
bounded read of the verified descriptor into an immutable local
``data: bytes`` buffer, computing its SHA-256 in the same pass
(:func:`_read_and_hash_fd_bounded`) and comparing it against
``candidate.sha256`` before any parsing begins -- never a second,
separate hash-then-rewind-then-reread of the same descriptor, for any
format. Evidence is then extracted from that one buffer with fixed,
ordered readers (PDF via ``pdfplumber``, isolated in a spawned worker
process bound by hard address-space, CPU-time, a single monotonic
wall-clock deadline, and result-byte limits so a small, highly
compressible content stream cannot decompress into far more memory than
its on-disk size implies; CSV via ``TextIOWrapper``/``csv.reader`` over
``io.BytesIO(data)``; ``.xlsx`` via ``openpyxl`` over
``io.BytesIO(data)``; ``.xls`` via the isolated
:func:`phi_engine.pipeline.xls_isolation.extract_xls_naming` worker
boundary, the ONLY module outside ``xls_isolation.py``/``_xls_worker.py``
allowed to parse legacy BIFF bytes) -- dispatched purely by suffix, never
by a stat-time guess. Every reader/workbook object is closed before this
pass returns; the descriptor's own post-read identity check (performed
by ``open_verified_source`` on context exit) always covers the exact,
complete read that produced ``data``. A single candidate's bounded
``data: bytes`` is transiently held for its own hash+parse and released
before the next candidate is read -- never more than one candidate's
buffer alive at once, and never all candidates' buffers simultaneously.
Every reader/parser failure -- including openpyxl's lazy worksheet
iteration, any PDF worker limit/termination/pathology, and any
``xls_isolation`` worker/isolation error -- collapses to the fixed
``support-evidence-limit`` code; a raw descriptor ``OSError`` collapses
to ``source-unreadable``; never a raw exception. Evidence is built
incrementally against each 8,192-UTF-8-byte cap *while parsing* (forms and
dictionary/mapping tracked independently; combined rebuilt from the same
already-bounded fragments), so retained state never scales with the total
number or size of support candidates -- only with what can ever fit the
canonical payload. ``.xlsx`` ZIP/archive bounds are enforced through the
one shared, directory-bounded ``zipfile.ZipFile`` primitive
(:func:`phi_engine.pipeline.support_files.open_bounded_zipfile`) also used
by :mod:`phi_engine.pipeline.intake_preflight`, never a second divergent
validator. CSV field allocation is bounded by a process-wide, lock-
serialized ``csv.field_size_limit`` in addition to an explicit UTF-8
byte-length check on every parsed field (not just the retained ones) before
any column slicing, so a field beyond column 20 cannot bypass the check by
never being kept. Every canonical evidence payload is gated through
``phi_gate_check`` before local dispatch and every response through
``guard_llm_output`` after, and dispatched only through the loopback-only,
attested, digest-pinned
:class:`~phi_engine.security.model_routing.OfflineLocalLLMClient`
(never ``config.get_llm_client()``). Local client construction/
configuration failures, transport/model failures, and output-validation
failures all collapse to the fixed ``study-name-inspection-failed`` error.
No source filename, path, header, sheet name, cell, value, raw exception,
prompt, or model response ever escapes this module.

Registry scanning and generated-tree reuse/promotion (matching a freshly
allocated ``study-<hex>`` name back to a prior study-less intake of the same
source, or promoting it once a real name is later supplied) are
:mod:`phi_engine.pipeline.intake` reconciliation concerns, not this
module's: they require the ``intake-registry`` lock and the v3 manifest
store, neither of which this module touches. :func:`canonical_source_root`
is the narrow, pure (no filesystem scan, no mutation) helper that
reconciliation layer needs to key its lookup; ``intake_root`` is accepted
here only so a future caller can thread it through unchanged.
"""

from __future__ import annotations

import contextlib
import csv
import hashlib
import io
import json
import math
import multiprocessing
import os
import secrets
import threading
import time
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any, BinaryIO, Callable, Iterator, Literal, Mapping

import openpyxl

from phi_engine.audit.review_paths import safe_review_slug
from phi_engine.config import config
from phi_engine.pipeline import _pdf_extract_worker, intake_preflight, support_files, xls_isolation
from phi_engine.pipeline.intake_preflight import IntakeCandidate, IntakePreflight
from phi_engine.pipeline.verified_source import VerifiedSourceError, open_verified_source
from phi_engine.security.llm_tool_guard import LLMToolOutputBlocked, guard_llm_output
from phi_engine.security.model_routing import (
    LocalModelUnavailableError,
    ModelResponseError,
    OfflineLocalLLMClient,
    new_offline_local_client,
)
from phi_engine.security.phi_gate import PHIEgressBlockedError, phi_gate_check
from phi_engine.utils import pipeline_lock

__all__ = [
    "StudyNameSource",
    "StudyResolution",
    "resolve_intake_study",
    "canonical_source_root",
]

StudyNameSource = Literal["user", "ai", "generated"]

# --- fixed evidence bounds (plan step 2) ------------------------------------------------
_MAX_DOCUMENT_BYTES = 16 * 1024 * 1024
_MAX_PDF_PAGES = 2
_MAX_WORKBOOK_SHEETS = 4
_MAX_ROWS = 20
_MAX_COLS = 20
_MAX_FRAGMENT_CODEPOINTS = 256
_MAX_CSV_FIELD_BYTES = 32 * 1024
_MAX_EVIDENCE_BYTES = 8192
_MAX_OUTPUT_TOKENS = 128
_MAX_MODEL_RESPONSE_BYTES = 4096
_HASH_CHUNK_SIZE = 1 << 20

# --- PDF worker isolation bounds (bounded, standard-library-only) ----------------------
#
# pdfplumber/pdfminer decompress PDF content streams fully before this
# module's own _MAX_FRAGMENT_CODEPOINTS check ever sees the result, so a
# small, highly compressible stream can otherwise expand into far more
# memory than _MAX_DOCUMENT_BYTES (the on-disk cap) would suggest. The
# actual pdfplumber.open()/extract_text() call therefore never runs in
# this process: it runs in a spawned child bound by a hard address-space
# ceiling (RLIMIT_AS), a hard CPU-time ceiling (RLIMIT_CPU) applied inside
# the child before pdfplumber ever touches the bytes, a wall-clock ceiling
# enforced here via poll()/join() timeouts, and a bounded result read
# (recv_bytes(maxlength=...)) so an oversized/malformed reply cannot be
# received either. Any of these firing -- OOM kill, CPU-time kill, hang,
# oversized reply, or a caught pdfplumber/pdfminer exception inside the
# child -- collapses to the same fixed _EvidenceLimitError.
_PDF_WORKER_MAX_ADDRESS_BYTES = 256 * 1024 * 1024
_PDF_WORKER_MAX_CPU_SECONDS = 5
_PDF_WORKER_MAX_WALL_SECONDS = 10
_PDF_WORKER_MAX_RESULT_BYTES = 65536

# VerifiedSourceError reasons that mean "we could not trust this read at
# all" -- bucketed as value-free errors, mirroring intake_preflight.py's
# own convention so the two modules never disagree about which fixed
# reasons are retryable-by-a-human review items versus hard errors.
_ERROR_REASONS = frozenset({"source-unreadable", "source-target-outside-root"})

_NAMING_COMPONENTS = frozenset({"dictionary_mapping", "forms"})
_ROOT_PATH = ""  # fixed string path for whole-source-root review/error records


@dataclass(frozen=True)
class StudyResolution:
    name: str
    source: StudyNameSource
    review_items: tuple[dict[str, Any], ...]
    errors: tuple[dict[str, Any], ...]


class _EvidenceLimitError(Exception):
    """Private control-flow signal: a candidate's content could not be
    safely bounded for naming evidence (oversized, expanded, or
    pathological input, including zip-bomb, malformed-archive, and any
    other parser/load/lazy-iteration/close failure)."""


class _InspectionFailed(Exception):
    """Private control-flow signal: the local model's response failed
    strict size/parse/key/type/range validation."""


def canonical_source_root(source: Path) -> str:
    """Canonical absolute source-root string, pure path resolution only.

    No filesystem scanning, no manifest reads, no mutation. Step 3's
    registry-scan/promotion reconciliation is expected to key its
    generated-manifest lookup on this exact value alongside
    ``StudyResolution.source == "generated"``.
    """
    return str(source.resolve())


def _generate_study_name() -> str:
    return f"study-{secrets.token_hex(4)}"


# --- descriptor-relative dataset-hardlink guard ---------------------------------------------
#
# FileIdentity (device/inode/size/mtime_ns) alone cannot see a hardlink
# created *after* preflight computed it: a new dataset dirent pointing at
# the same inode as an already-admitted support candidate changes neither
# that candidate's stat fields nor its content/hash. Closing this TOCTOU
# window requires an independent, fresh, descriptor-relative scan of the
# *current* datasets/ tree, compared by (device, inode) against whatever
# support candidate is about to be read.


def _current_dataset_identities(source: Path) -> frozenset[tuple[int, int]] | None:
    """Fresh scan of every regular file's ``(device, inode)`` currently
    reachable under ``<source>/datasets/``, delegated entirely to the
    shared, preflight-proven
    :func:`phi_engine.pipeline.intake_preflight._scan_component_identities`
    primitive (built on the same ``_walk_tree`` engine, NOFOLLOW
    discipline, and per-directory identity recheck
    :func:`~phi_engine.pipeline.intake_preflight.inspect_intake_source`
    itself uses) -- never a second, divergent traversal. Computed fresh
    immediately before parsing each naming candidate and again immediately
    before every individual model dispatch; never cached across those
    checkpoints.

    Returns ``None`` -- never a silently-empty set -- whenever the scan is
    inconclusive for any reason (unpinnable root, absent/symlinked/
    unopenable ``datasets/``, or any traversal-time failure, including a
    final/intermediate path swap). Callers MUST treat ``None`` identically
    to a confirmed hardlink alias: fail closed, discard all evidence,
    dispatch nothing.
    """
    identities, _reason = intake_preflight._scan_component_identities(source, "datasets")
    return identities


def _hardlink_race_detected(source: Path, identities: frozenset[tuple[int, int]] | set[tuple[int, int]]) -> bool:
    if not identities:
        return False
    current = _current_dataset_identities(source)
    if current is None:
        return True
    return bool(identities & current)


# --- preflight-time source-trust safety findings ----------------------------------------
#
# Preflight already quarantines a support file it catches aliased into
# datasets/ (cross-component-hardlink) or sitting behind a symlink
# (source-symlink-not-allowed) to _unclassified, recording a fixed review
# item for it and never handing it to naming as a candidate at all; a
# source it could not even trust enough to classify (source-unreadable,
# source-target-outside-root) is recorded as a preflight ERROR instead,
# using the identical verified-source vocabulary this module's own
# _ERROR_REASONS bucket already uses. Both buckets are the SAME safety
# catalog -- one complete, fixed set of verified-source trust codes --
# and naming must abort before opening ANY candidate or constructing a
# local model client the moment either bucket contains one, not just the
# review-item subset: an unreadable/out-of-root source elsewhere in the
# tree is exactly as untrustworthy a signal about the whole source as a
# caught symlink or hardlink is. Naming adds nothing new here, it just
# refuses to read or dispatch anything.

_SOURCE_TRUST_SAFETY_REASONS = frozenset({"source-symlink-not-allowed", "cross-component-hardlink"} | _ERROR_REASONS)


def _preflight_has_safety_finding(preflight: IntakePreflight) -> bool:
    return any(item.get("reason") in _SOURCE_TRUST_SAFETY_REASONS for item in preflight.review_items) or any(
        item.get("reason") in _SOURCE_TRUST_SAFETY_REASONS for item in preflight.errors
    )


# --- entry point -------------------------------------------------------------------------


def _resolve_intake_study(
    source: Path,
    preflight: IntakePreflight,
    *,
    explicit_study: str | None,
    support_confirmed_no_phi: bool,
    intake_root: Path,
    generate_study_name: Callable[[], str] = _generate_study_name,
) -> StudyResolution:
    del intake_root  # reserved for step 3's registry-scan/promotion wiring; unused here

    if explicit_study is not None:
        # No-write validation contract: raises ValueError on a malformed
        # name. Deliberately not caught -- an invalid --study is a caller
        # bug, not a data-driven outcome with a fixed reason code.
        pipeline_lock.lock_path_for(explicit_study)
        return StudyResolution(name=explicit_study, source="user", review_items=(), errors=())

    if not support_confirmed_no_phi:
        # Zero naming-content extraction, zero model calls. The default
        # false/unknown consent state performs no reads beyond what
        # preflight already did.
        return StudyResolution(
            name=generate_study_name(),
            source="generated",
            review_items=({"path": _ROOT_PATH, "reason": "support-phi-status-required", "blocking": True},),
            errors=(),
        )

    if _preflight_has_safety_finding(preflight):
        return StudyResolution(
            name=generate_study_name(),
            source="generated",
            review_items=(),
            errors=(),
        )

    admitted = sorted(
        (
            candidate
            for candidate in preflight.candidates
            if candidate.component in _NAMING_COMPONENTS and candidate.component == candidate.source_component
        ),
        key=lambda c: c.relative_path,
    )

    # Item 4: before parsing ANY support evidence, freshly verify every
    # preflight-known datasets/ candidate's current identity AND SHA-256
    # against its preflight record. Any failure aborts the whole naming
    # attempt (source-unreadable, zero further reads, zero dispatch).
    # Candidates that are themselves a lexical alias or an independent
    # byte-identical copy of a verified dataset are silently excluded
    # from admission -- they cause zero parser calls, never even
    # reaching _validate_candidate/_extract_candidate -- and are never
    # modified or reported a second time (preflight's own phase-2
    # cross-component quarantine already reports what it can see).
    dataset_ok, verified_dataset_identities, verified_dataset_hashes, dataset_error = _verify_dataset_snapshot(
        source, preflight
    )
    if not dataset_ok:
        return StudyResolution(
            name=generate_study_name(),
            source="generated",
            review_items=(),
            errors=(dataset_error,) if dataset_error is not None else (),
        )

    admitted = [
        candidate
        for candidate in admitted
        if (candidate.identity.device, candidate.identity.inode) not in verified_dataset_identities
        and candidate.sha256 not in verified_dataset_hashes
    ]

    forms_docs, dict_docs, review_items, errors, verified_identities, verified_candidates, aborted = _collect_evidence(
        source, admitted
    )

    if aborted:
        return StudyResolution(
            name=generate_study_name(),
            source="generated",
            review_items=tuple(review_items),
            errors=tuple(errors),
        )

    forms_payload = _forms_payload_dict(forms_docs)
    dict_payload = _dict_payload_dict(dict_docs)
    forms_json = _canonical_json(forms_payload) if forms_payload["documents"] else None
    dict_json = _canonical_json(dict_payload) if dict_payload["documents"] else None

    chosen: str | None = None
    phi_blocked = False
    hardlink_race = False
    client: OfflineLocalLLMClient | None = None

    def get_client() -> OfflineLocalLLMClient:
        nonlocal client
        if client is None:
            client = new_offline_local_client()
        return client

    def dispatch_guarded(evidence_json: str) -> str | None:
        """Recheck two INDEPENDENT, deliberately redundant guards
        immediately before THIS specific dispatch (not once for the
        whole resolution) -- a client whose own first call creates a
        dict-to-datasets hardlink must never see a second call. First,
        the pre-existing identity-only ``_hardlink_race_detected`` scan
        (unchanged). Second, item 5's fresh repeat of the item-4 dataset
        descriptor scan/hash PLUS a genuinely live re-hash of every
        retained support candidate's CURRENT bytes
        (:func:`_live_support_hashes`) -- comparing a fresh dataset
        SHA-256 set against these candidates' STATIC preflight-recorded
        hashes would be a no-op (identical by construction since
        admission, so it could never observe a race happening strictly
        after admission); only a live re-read of the support side closes
        that window. Every dataset candidate must still match its
        preflight record, AND no freshly recomputed dataset SHA-256 may
        now intersect any retained support candidate's LIVE current
        SHA-256 (a copy-based race introduced between evidence
        collection and this specific dispatch, which an identity-only
        check alone cannot see). A scan-failure (``not dataset_ok`` or
        ``live_support_hashes is None``, as opposed to a SHA
        intersection) has no distinct review code of its own here --
        deliberately folded into the same fixed cross-component-hardlink
        vocabulary as every other post-collection race this module
        tracks, rather than exposing a new code per failure mode."""
        nonlocal hardlink_race
        if _hardlink_race_detected(source, verified_identities):
            hardlink_race = True
            return None
        dataset_ok, _fresh_identities, fresh_hashes, _dataset_error = _verify_dataset_snapshot(source, preflight)
        live_support_hashes = _live_support_hashes(source, verified_candidates)
        if not dataset_ok or live_support_hashes is None or (fresh_hashes & live_support_hashes):
            hardlink_race = True
            return None
        return _dispatch(get_client, evidence_json, errors)

    forms_name: str | None = None
    dict_name: str | None = None

    try:
        if forms_json is not None and not hardlink_race:
            forms_name = dispatch_guarded(forms_json)
        if dict_json is not None and not hardlink_race and not phi_blocked:
            dict_name = dispatch_guarded(dict_json)
    except PHIEgressBlockedError:
        phi_blocked = True
        forms_name = dict_name = None

    if not phi_blocked and not hardlink_race:
        if forms_name is not None and dict_name is not None:
            if forms_name.casefold() == dict_name.casefold():
                chosen = forms_name  # preserve the forms spelling
            else:
                review_items.append(
                    {
                        "path": _ROOT_PATH,
                        "reason": "study-name-conflict",
                        "blocking": True,
                        "candidates": {"forms": forms_name, "dictionary_mapping": dict_name},
                    }
                )
        elif forms_name is not None:
            chosen = forms_name
        elif dict_name is not None:
            chosen = dict_name
        elif forms_docs or dict_docs:
            forms_fragments = _forms_fragments_from_docs(forms_docs)
            dict_fragments = _dict_fragments_from_docs(dict_docs)
            combined_payload = _grow_combined(forms_fragments, dict_fragments, _MAX_EVIDENCE_BYTES)
            combined_json = (
                _canonical_json(combined_payload)
                if (combined_payload["forms"] or combined_payload["dictionary_mapping"])
                else None
            )
            if combined_json is not None:
                try:
                    chosen = dispatch_guarded(combined_json)
                except PHIEgressBlockedError:
                    phi_blocked = True

    if hardlink_race:
        review_items.append({"path": _ROOT_PATH, "reason": "cross-component-hardlink", "blocking": True})
        chosen = None
    if phi_blocked:
        review_items.append({"path": _ROOT_PATH, "reason": "possible-phi-requires-study", "blocking": True})
        chosen = None

    if chosen is not None:
        return StudyResolution(name=chosen, source="ai", review_items=tuple(review_items), errors=tuple(errors))

    return StudyResolution(
        name=generate_study_name(),
        source="generated",
        review_items=tuple(review_items),
        errors=tuple(errors),
    )


def resolve_intake_study(
    source: Path,
    preflight: IntakePreflight,
    *,
    explicit_study: str | None,
    support_confirmed_no_phi: bool,
    intake_root: Path,
) -> StudyResolution:
    return _resolve_intake_study(
        source,
        preflight,
        explicit_study=explicit_study,
        support_confirmed_no_phi=support_confirmed_no_phi,
        intake_root=intake_root,
    )


# --- local dispatch ------------------------------------------------------------------------

_PROMPT_PREFIX = (
    "You infer a short filesystem-safe study folder name from de-identified "
    "structural evidence only (no patient data). Respond with exactly one "
    'JSON object and nothing else: {"study_name": <string-or-null>, '
    '"confidence": <number 0 to 1>}. Use null when no clear name is present. '
    "Evidence:\n"
)


def _build_prompt(evidence_json: str) -> str:
    return _PROMPT_PREFIX + evidence_json


def _dispatch(
    get_client: Callable[[], OfflineLocalLLMClient], evidence_json: str, errors: list[dict[str, Any]]
) -> str | None:
    gate = phi_gate_check(evidence_json)
    if gate.blocked:
        raise PHIEgressBlockedError("naming evidence blocked by phi_gate_check")

    prompt = _build_prompt(evidence_json)
    try:
        client = get_client()
        raw = client.complete_bounded(
            prompt, max_output_tokens=_MAX_OUTPUT_TOKENS, max_response_bytes=_MAX_MODEL_RESPONSE_BYTES
        )
        guard_llm_output(raw)
        study_name, confidence = _parse_model_output(raw)
    except (
        LocalModelUnavailableError,
        ModelResponseError,
        LLMToolOutputBlocked,
        _InspectionFailed,
        config.LocalLLMConfigurationError,
    ):
        errors.append({"path": None, "reason": "study-name-inspection-failed"})
        return None

    if study_name is None or confidence < config.PHI_CONFIDENCE_THRESHOLD:
        return None
    return _normalize_and_validate(study_name)


def _parse_model_output(raw: str) -> tuple[str | None, float]:
    if not isinstance(raw, str) or not raw:
        raise _InspectionFailed()
    try:
        value = json.loads(raw, object_pairs_hook=_reject_duplicate_keys)
    except (ValueError, TypeError, RecursionError):
        raise _InspectionFailed() from None
    if not isinstance(value, dict) or set(value) != {"study_name", "confidence"}:
        raise _InspectionFailed()

    study_name = value["study_name"]
    if study_name is not None and not isinstance(study_name, str):
        raise _InspectionFailed()

    confidence = value["confidence"]
    if isinstance(confidence, bool) or not isinstance(confidence, (int, float)):
        raise _InspectionFailed()
    confidence = float(confidence)
    if not math.isfinite(confidence) or not (0.0 <= confidence <= 1.0):
        raise _InspectionFailed()

    return study_name, confidence


def _reject_duplicate_keys(pairs: list[tuple[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in pairs:
        if key in result:
            raise ValueError("duplicate key")
        result[key] = value
    return result


def _normalize_and_validate(raw: str) -> str | None:
    slug = safe_review_slug(raw)[:128]
    try:
        pipeline_lock.lock_path_for(slug)
    except ValueError:
        return None
    return slug


# --- shared, process-wide CSV field-size ceiling --------------------------------------------
#
# csv.field_size_limit() is process-global mutable state. Serializing its
# mutation under a lock and always restoring the previous value keeps this
# module's bound from leaking into (or being clobbered by) any other CSV
# parsing happening concurrently elsewhere in the process.

_CSV_FIELD_LIMIT_LOCK = threading.Lock()


@contextlib.contextmanager
def _bounded_csv_field_limit() -> Iterator[None]:
    with _CSV_FIELD_LIMIT_LOCK:
        previous = csv.field_size_limit()
        csv.field_size_limit(_MAX_CSV_FIELD_BYTES)
        try:
            yield
        finally:
            csv.field_size_limit(previous)


# --- candidate evidence extraction ----------------------------------------------------------


def _hash_fd_bounded(fd: int, max_bytes: int) -> str:
    """Chunked SHA-256 of the current descriptor's remaining content,
    never retaining more than one bounded chunk in memory at a time (no
    full-document byte copy). Raises :class:`_EvidenceLimitError` if the
    actual bytes read exceed ``max_bytes``, independent of the recorded
    stat size."""
    digest = hashlib.sha256()
    total = 0
    while True:
        chunk = os.read(fd, _HASH_CHUNK_SIZE)
        if not chunk:
            break
        total += len(chunk)
        if total > max_bytes:
            raise _EvidenceLimitError()
        digest.update(chunk)
    return digest.hexdigest()


def _read_and_hash_fd_bounded(fd: int, max_bytes: int) -> tuple[bytes, str]:
    """Single bounded read of the current descriptor's remaining content --
    the ONE read pass 2 ever performs, closing the former hash-then-
    rewind-then-reread window. Accumulates bytes into a local ``bytearray``
    while simultaneously updating a running SHA-256 digest over each chunk
    read, mirroring :func:`_hash_fd_bounded`'s own chunked style but
    RETAINING the bytes (up to ``max_bytes``) for the caller's single parse
    instead of discarding them. Raises :class:`_EvidenceLimitError` the
    instant total bytes read exceeds ``max_bytes``, exactly as
    :func:`_hash_fd_bounded` already does."""
    digest = hashlib.sha256()
    buf = bytearray()
    while True:
        chunk = os.read(fd, _HASH_CHUNK_SIZE)
        if not chunk:
            break
        buf += chunk
        if len(buf) > max_bytes:
            raise _EvidenceLimitError()
        digest.update(chunk)
    return bytes(buf), digest.hexdigest()


# --- fresh dataset descriptor/hash snapshot (plan step 3, items 4-5) ------------------------
#
# The pre-existing _current_dataset_identities/_hardlink_race_detected
# pair is an identity-only (device, inode) guard: it catches a NEW
# datasets/ dirent hardlinked to an already-admitted support candidate's
# inode, but it cannot see a lexically-distinct, byte-identical COPY of
# a dataset (no shared inode at all) landing in a support component, nor
# can it independently confirm that preflight's own recorded dataset
# bytes are still exactly what preflight saw. _verify_dataset_snapshot
# closes both gaps with a fresh, independent, descriptor-verified
# re-hash of every preflight-known dataset candidate, run immediately
# before any support evidence is ever parsed and again immediately
# before every individual local-model dispatch -- deliberately redundant
# with (never a replacement for) the identity-only guard above.
#
# Cost note: every datasets/ candidate is fully re-hashed once at item 4
# and again inside EVERY dispatch_guarded call (up to 3x for forms/
# dict/combined), so up to 4 full re-hashes of the whole datasets/ tree
# can happen per naming resolution -- a deliberate, correct security
# trade-off (never a shortcut on this boundary), not an accidental
# performance regression a future investigation should mistake for one.


def _hash_fd_streaming(fd: int) -> str:
    """Chunked SHA-256 of the current descriptor's remaining content with
    no evidence-sized bound -- used only to verify a ``datasets/``
    candidate's current bytes against its preflight-recorded ``sha256``,
    never to retain or parse dataset content. Unlike
    :func:`_hash_fd_bounded`/:func:`_read_and_hash_fd_bounded`, this
    applies no ``_MAX_DOCUMENT_BYTES``-style cap: dataset files are
    legitimately far larger than any support document this module ever
    parses, and none of their bytes are ever retained here regardless."""
    digest = hashlib.sha256()
    while True:
        chunk = os.read(fd, _HASH_CHUNK_SIZE)
        if not chunk:
            break
        digest.update(chunk)
    return digest.hexdigest()


def _verify_dataset_snapshot(
    source: Path, preflight: IntakePreflight
) -> tuple[bool, frozenset[tuple[int, int]], frozenset[str], dict[str, Any] | None]:
    """Fresh, independent descriptor-open plus streaming-hash of EVERY
    ``preflight.candidates`` entry whose ``source_component ==
    "datasets"`` -- filtered by ``source_component``, never by the
    (possibly reclassified) logical ``component``, so a dataset preflight
    itself downgraded to ``_unclassified`` (for example a rejected
    ``dataset-xls-multiple-sheets`` candidate) is still included: its
    bytes must never leak into naming evidence via a support candidate
    that happens to share its hash. Each candidate is opened through the
    same :func:`open_verified_source` primitive used everywhere else in
    this module, requiring BOTH the current descriptor identity
    (``open_verified_source``'s own ``expected_identity`` check) AND a
    freshly recomputed SHA-256 to equal the candidate's preflight-
    recorded ``sha256``.

    Returns ``(ok, identities, hashes, error_item)``. ``ok`` is ``False``
    the instant any dataset candidate cannot be opened/verified or its
    hash has drifted -- the caller MUST abort the WHOLE naming attempt
    with the same all-or-nothing contract as every other safety failure
    in this module: zero further reads, zero dispatch, reporting the
    fixed ``source-unreadable`` ``error_item``. On success,
    ``identities``/``hashes`` are this fresh scan's complete
    ``(device, inode)``/``sha256`` sets, used by the caller to silently
    exclude any dictionary_mapping/forms candidate that is a lexical
    alias or an independent byte-identical copy of a verified dataset --
    never a second review/error report; preflight's own phase-2
    cross-component quarantine already reports what it can see from its
    own vantage point, this is naming's own independent, silent,
    defense-in-depth exclusion layer."""
    identities: set[tuple[int, int]] = set()
    hashes: set[str] = set()
    for candidate in preflight.candidates:
        if candidate.source_component != "datasets":
            continue
        try:
            with open_verified_source(
                source,
                candidate.relative_path,
                required_source_component="datasets",
                expected_identity=candidate.identity,
            ) as fd:
                digest = _hash_fd_streaming(fd)
        except (VerifiedSourceError, OSError):
            return False, frozenset(), frozenset(), {"path": candidate.relative_path, "reason": "source-unreadable"}
        if digest != candidate.sha256:
            return False, frozenset(), frozenset(), {"path": candidate.relative_path, "reason": "source-unreadable"}
        identities.add((candidate.identity.device, candidate.identity.inode))
        hashes.add(candidate.sha256)
    return True, frozenset(identities), frozenset(hashes), None


def _live_support_hashes(source: Path, candidates: tuple[IntakeCandidate, ...]) -> frozenset[str] | None:
    """Re-hash the CURRENT on-disk bytes of every pass-1-admitted support
    candidate, fresh, immediately before a dispatch. Unlike a comparison
    against ``candidate.sha256`` (the preflight-recorded value, static
    since admission -- identical whether read at item 4 or item 5, so it
    can never observe a race that happens strictly after admission),
    this is a genuinely live read: a candidate whose bytes are swapped
    for a dataset's bytes AFTER pass-1 validation, with identity
    (size/mtime) restored to defeat ``open_verified_source``'s own
    check alone, is still caught here. Returns ``None`` -- never a
    partial/best-effort set -- the instant any candidate cannot be
    reopened/rehashed, exactly like every other verified-source failure
    in this module: fail closed."""
    hashes: set[str] = set()
    for candidate in candidates:
        try:
            with open_verified_source(
                source,
                candidate.relative_path,
                required_source_component=candidate.source_component,
                expected_identity=candidate.identity,
            ) as fd:
                hashes.add(_hash_fd_bounded(fd, _MAX_DOCUMENT_BYTES))
        except (VerifiedSourceError, OSError, _EvidenceLimitError):
            return None
    return frozenset(hashes)


def _validate_candidate(
    source: Path, candidate: IntakeCandidate
) -> tuple[bool, dict[str, Any] | None, dict[str, Any] | None]:
    """Pass 1 of 2 (validation): open the verified descriptor, freshly
    rescan ``datasets/`` for a hardlink alias, and hash-verify content --
    the complete safety checkpoint for one candidate, run for EVERY
    admitted candidate before ANY candidate's content is parsed (see
    :func:`_collect_evidence`). Never opens by raw pathname; the caller's
    own :class:`~phi_engine.pipeline.intake_preflight.IntakeCandidate`
    (its already-computed identity/sha256) is the only thing carried
    forward -- there is no live descriptor to "keep" across passes, so
    pass 2 re-verifies through the same sanctioned primitive immediately
    before its one parser call instead.

    Returns ``(ok, review_item, error_item)``. ``ok`` is ``False`` for a
    dataset-alias hit, a descriptor identity/hash mismatch, or a
    naming-time symlink rejection -- the caller must discard everything
    collected so far and abort the WHOLE naming attempt, not just this
    candidate. An oversized candidate (``support-evidence-limit``) is the
    one failure that leaves ``ok`` ``True``: its descriptor identity is
    still trusted and tracked by the caller for the pre-dispatch
    hardlink guard, only its content is never parsed.
    """
    try:
        with open_verified_source(
            source,
            candidate.relative_path,
            required_source_component=candidate.source_component,
            expected_identity=candidate.identity,
        ) as fd:
            info = os.fstat(fd)
            current_datasets = _current_dataset_identities(source)
            if current_datasets is None or (info.st_dev, info.st_ino) in current_datasets:
                return False, None, None

            digest = _hash_fd_bounded(fd, _MAX_DOCUMENT_BYTES)
            if digest != candidate.sha256:
                raise VerifiedSourceError("source-unreadable")
    except VerifiedSourceError as exc:
        if exc.reason in _ERROR_REASONS:
            return False, None, {"path": candidate.relative_path, "reason": exc.reason}
        return False, {"path": candidate.relative_path, "reason": exc.reason, "blocking": True}, None
    except OSError:
        # Descriptor-level read failure distinct from what
        # open_verified_source's own identity/symlink checks normalize.
        return False, None, {"path": candidate.relative_path, "reason": "source-unreadable"}
    except _EvidenceLimitError:
        return True, {"path": candidate.relative_path, "reason": "support-evidence-limit", "blocking": True}, None

    return True, None, None


def _extract_candidate(
    source: Path, candidate: IntakeCandidate
) -> tuple[Any, bool, dict[str, Any] | None, dict[str, Any] | None]:
    """Pass 2 of 2 (extraction): re-open the ALREADY-validated candidate
    through the same sanctioned :func:`open_verified_source` primitive --
    re-checking its expected identity and freshly rescanning
    ``datasets/`` again immediately before this specific parser call, so
    a hardlink created in the window between pass 1 and this call is
    still caught. A same-inode, same-size mutation whose mtime has been
    restored to the identity pass 1 already validated defeats
    ``open_verified_source``'s identity check alone, so this pass also
    performs exactly ONE bounded read of the freshly opened descriptor
    into an immutable local ``data: bytes`` while simultaneously
    computing its SHA-256 (:func:`_read_and_hash_fd_bounded`) -- never a
    second, separate hash-then-rewind-then-reread of the same
    descriptor -- and compares the finished digest against
    ``candidate.sha256`` -- exactly like pass 1 -- before parsing.
    Never reopens by raw pathname.

    Returns ``(fragments, abort, review_item, error_item)`` with the same
    ``abort``/record contract as :func:`_validate_candidate`'s ``ok``.
    ``fragments`` is ``list[str]`` (pages) for a forms candidate, or
    ``list[tuple[sheet_index, rows]]`` for a dictionary/mapping
    candidate; ``None``/``[]`` when nothing was retained.

    The verified-descriptor context stays open for the whole read+hash;
    only the resulting ``data: bytes`` (never the live descriptor) is
    handed to a parser, so the descriptor's own post-read identity check
    (performed by ``open_verified_source`` on context exit) always
    covers the exact, complete read.
    """
    is_forms = candidate.component == "forms"
    suffix = PurePosixPath(candidate.relative_path).suffix.lower()

    try:
        with open_verified_source(
            source,
            candidate.relative_path,
            required_source_component=candidate.source_component,
            expected_identity=candidate.identity,
        ) as fd:
            info = os.fstat(fd)
            current_datasets = _current_dataset_identities(source)
            if current_datasets is None or (info.st_dev, info.st_ino) in current_datasets:
                return None, True, None, None

            data, digest = _read_and_hash_fd_bounded(fd, _MAX_DOCUMENT_BYTES)
            if digest != candidate.sha256:
                raise VerifiedSourceError("source-unreadable")

            if is_forms:
                fragments: Any = _extract_pdf_pages(data)
            elif suffix == ".csv":
                fragments = [(1, _extract_csv_rows(io.BytesIO(data)))]
            elif suffix == ".xlsx":
                fragments = _extract_xlsx_sheets(io.BytesIO(data))
            elif suffix == ".xls":
                try:
                    fragments = xls_isolation.extract_xls_naming(data, candidate.sha256)
                except (xls_isolation.XlsIsolationError, xls_isolation.XlsWorkerError):
                    raise _EvidenceLimitError() from None
            else:
                # Structurally impossible given
                # intake_preflight._COMPONENT_SUFFIXES -- defensive only.
                raise _EvidenceLimitError()
    except VerifiedSourceError as exc:
        if exc.reason in _ERROR_REASONS:
            return None, True, None, {"path": candidate.relative_path, "reason": exc.reason}
        return None, True, {"path": candidate.relative_path, "reason": exc.reason, "blocking": True}, None
    except OSError:
        # Descriptor-level read/dup/fdopen failure distinct from what
        # open_verified_source's own identity/symlink checks normalize.
        return None, True, None, {"path": candidate.relative_path, "reason": "source-unreadable"}
    except _EvidenceLimitError:
        return None, False, {"path": candidate.relative_path, "reason": "support-evidence-limit", "blocking": True}, None

    return fragments, False, None, None


def _collect_evidence(
    source: Path, admitted: list[IntakeCandidate]
) -> tuple[
    dict[int, list[str]],
    dict[int, dict[int, list[list[str]]]],
    list[dict[str, Any]],
    list[dict[str, Any]],
    frozenset[tuple[int, int]],
    tuple[IntakeCandidate, ...],
    bool,
]:
    """Two-pass evidence collection. Pass 1 validates EVERY admitted
    candidate (descriptor open, expected identity, fresh dataset-alias
    scan, bounded hash, post-read identity via
    :func:`_validate_candidate`) before pass 2 ever parses ANY of them --
    a later candidate's safety failure is no longer discoverable only
    after an earlier, otherwise-clean candidate was already parsed. A
    single symlink/hardlink/identity/hash safety failure on any
    candidate, in either pass, immediately discards everything collected
    so far (including from earlier, otherwise-clean candidates) and
    aborts the whole naming attempt with zero further reads and zero
    dispatch; an oversized (``support-evidence-limit``) candidate does
    not abort.

    Pass 2 then parses only the pass-1-validated candidates, honoring
    the per-component evidence-byte budget -- once a component's budget
    closes, further candidates of that component are skipped entirely
    (their pass-1 validation already ran) -- and revalidates (fresh
    descriptor open + expected identity + dataset-alias rescan)
    immediately before each individual parser call.

    The identity set returned to the caller is the COMPLETE pass-1
    admitted set -- oversized, budget-closed (non-contributing), and
    parser-rejected candidates included, not just candidates that
    contributed a retained evidence fragment -- so the caller's
    pre-dispatch hardlink guard cannot be defeated by hardlinking a
    candidate that never produced retained evidence. The candidate
    objects returned alongside it are that SAME complete pass-1-admitted
    set (not just fragment-contributing ones, for the identical reason),
    letting the caller's pre-dispatch dataset-SHA cross-check
    (:func:`_live_support_hashes`) re-open and re-hash each one's LIVE
    current bytes immediately before a dispatch -- a genuinely fresh
    read, never a comparison against these candidates' own static
    preflight-recorded ``sha256`` (which cannot, by construction, differ
    from what admission already saw and so could never observe a race
    happening strictly after admission).
    """
    review_items: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    verified: list[IntakeCandidate] = []
    admitted_identities: set[tuple[int, int]] = set()
    admitted_candidates: list[IntakeCandidate] = []

    for candidate in admitted:
        ok, review_item, error_item = _validate_candidate(source, candidate)
        if review_item is not None:
            review_items.append(review_item)
        if error_item is not None:
            errors.append(error_item)
        if not ok:
            # No candidate-specific record (a bare dataset-alias scan
            # hit) -- fall back to one fixed root-level note so the
            # caller still has something to report.
            if review_item is None and error_item is None:
                review_items.append({"path": _ROOT_PATH, "reason": "cross-component-hardlink", "blocking": True})
            return {}, {}, review_items, errors, frozenset(), (), True
        admitted_identities.add((candidate.identity.device, candidate.identity.inode))
        admitted_candidates.append(candidate)
        if review_item is None:
            # An oversized (support-evidence-limit) candidate already got
            # its one fixed review record here in pass 1 -- its identity
            # is tracked above, but pass 2 must not attempt to read it
            # again (which would just rediscover the same oversize and
            # append a second, duplicate record).
            verified.append(candidate)

    forms_docs: dict[int, list[str]] = {}
    dict_docs: dict[int, dict[int, list[list[str]]]] = {}
    form_index = 0
    dict_index = 0
    forms_budget_open = True
    dict_budget_open = True

    for candidate in verified:
        is_forms = candidate.component == "forms"
        budget_open = forms_budget_open if is_forms else dict_budget_open
        if not budget_open:
            continue

        fragments, abort, review_item, error_item = _extract_candidate(source, candidate)
        if review_item is not None:
            review_items.append(review_item)
        if error_item is not None:
            errors.append(error_item)
        if abort:
            if review_item is None and error_item is None:
                review_items.append({"path": _ROOT_PATH, "reason": "cross-component-hardlink", "blocking": True})
            return {}, {}, review_items, errors, frozenset(), (), True
        if not fragments:
            continue

        if is_forms:
            form_index += 1
            idx = form_index
            for page in fragments:
                trial = {key: list(value) for key, value in forms_docs.items()}
                trial.setdefault(idx, []).append(page)
                if _encoded_len(_forms_payload_dict(trial)) > _MAX_EVIDENCE_BYTES:
                    forms_budget_open = False
                    break
                forms_docs = trial
        else:
            dict_index += 1
            idx = dict_index
            stop = False
            for sheet_index, rows in fragments:
                if stop:
                    break
                for row in rows:
                    trial = {
                        key: {sk: list(sv) for sk, sv in value.items()} for key, value in dict_docs.items()
                    }
                    sheets = dict(trial.get(idx, {}))
                    sheets[sheet_index] = [*sheets.get(sheet_index, []), row]
                    trial[idx] = sheets
                    if _encoded_len(_dict_payload_dict(trial)) > _MAX_EVIDENCE_BYTES:
                        dict_budget_open = False
                        stop = True
                        break
                    dict_docs = trial

    return forms_docs, dict_docs, review_items, errors, frozenset(admitted_identities), tuple(admitted_candidates), False


_PDF_WORKER_CONTEXT = multiprocessing.get_context("spawn")

# Reap grace is deliberately SEPARATE from _PDF_WORKER_MAX_WALL_SECONDS:
# the wall bound governs how long an ordinary poll()/join() may wait for
# a cooperating worker; this grace governs only how long terminate()/
# kill() are given to actually land once the wall deadline has already
# expired and the child is being forcibly reaped. Applying the wall
# bound twice (once to poll, again to join) would silently double the
# effective hang tolerance -- this module uses ONE monotonic deadline
# shared by poll() and join(), never two independent full-duration waits.
_PDF_WORKER_REAP_GRACE_SECONDS = 2.0


def _reap_pdf_worker(process: Any) -> None:
    """Positively terminate, escalating to kill, a still-alive worker.
    Never called unless ``process.start()`` already succeeded. Each step
    is independently exception-guarded -- a raw exception from
    ``is_alive``/``terminate``/``join`` must never skip the kill
    escalation that follows it, or a child could be left alive simply
    because ``terminate()`` itself raised. Best-effort: this function
    itself never raises."""
    try:
        if process.is_alive():
            process.terminate()
    except Exception:
        pass
    try:
        if process.is_alive():
            process.join(_PDF_WORKER_REAP_GRACE_SECONDS)
    except Exception:
        pass
    try:
        if process.is_alive():
            process.kill()
    except Exception:
        pass
    try:
        if process.is_alive():
            process.join(_PDF_WORKER_REAP_GRACE_SECONDS)
    except Exception:
        pass


def _run_pdf_worker(process: Any, parent_conn: Any, child_conn: Any) -> bytes | None:
    """Runs the spawned worker under ONE monotonic wall-clock deadline
    shared by the initial ``poll()`` and the post-signal ``join()`` --
    the configured wall bound is never applied twice before termination.
    Positively reaps a started child on every path (never leaves it
    alive), closes both pipe endpoints and the ``Process`` handle
    regardless of outcome, and normalizes every lifecycle exception
    (``start``/``poll``/``recv_bytes``/``join``/``terminate``/``kill``)
    to a bare ``None`` result rather than letting a raw library
    exception (whose message could echo source content, or simply
    library-internal detail) escape this boundary -- the caller raises
    the one fixed ``_EvidenceLimitError`` for every such outcome, with no
    exception chaining back to whatever was actually raised here.
    """
    deadline = time.monotonic() + _PDF_WORKER_MAX_WALL_SECONDS
    started = False
    raw: bytes | None = None
    exitcode: int | None = None
    try:
        process.start()
        started = True
        try:
            child_conn.close()  # only the child writes; the parent's copy must not linger
        except OSError:
            pass
        remaining = deadline - time.monotonic()
        if remaining > 0 and parent_conn.poll(remaining):
            raw = parent_conn.recv_bytes(maxlength=_PDF_WORKER_MAX_RESULT_BYTES)
        process.join(max(0.0, deadline - time.monotonic()))
    except Exception:
        raw = None
    finally:
        try:
            if started:
                _reap_pdf_worker(process)
        except Exception:
            raw = None
        if started:
            # Exitcode must be read BEFORE process.close() -- a closed
            # Process handle raises on every further attribute access,
            # exitcode included.
            try:
                exitcode = process.exitcode
            except Exception:
                exitcode = None
        for conn in (parent_conn, child_conn):
            try:
                conn.close()
            except Exception:
                pass
        try:
            process.close()
        except Exception:
            pass

    if not started or raw is None or exitcode != 0:
        return None
    return raw


def _decode_worker_reply(raw: bytes) -> list[str]:
    """Strict schema validation of the worker's bounded, non-executable
    UTF-8 (ASCII-subset) JSON reply: an exact two-key
    ``{"status", "pages"}`` object, literal ``"ok"`` status, a ``pages``
    list bounded by ``_MAX_PDF_PAGES``, and each page a non-boolean
    ``str`` within ``_MAX_FRAGMENT_CODEPOINTS``. Any other shape, type,
    key, or cardinality collapses to the fixed ``_EvidenceLimitError`` --
    never a raw parse/type exception, and never ``pickle`` or any other
    format capable of executing code while being decoded."""
    try:
        payload = json.loads(raw.decode("ascii"))
    except (UnicodeDecodeError, json.JSONDecodeError):
        raise _EvidenceLimitError() from None
    if not isinstance(payload, dict) or set(payload) != {"status", "pages"}:
        raise _EvidenceLimitError()
    status = payload["status"]
    texts = payload["pages"]
    if status != "ok" or not isinstance(texts, list) or len(texts) > _MAX_PDF_PAGES:
        raise _EvidenceLimitError()
    for text in texts:
        if not isinstance(text, str) or isinstance(text, bool) or len(text) > _MAX_FRAGMENT_CODEPOINTS:
            raise _EvidenceLimitError()
    return texts


def _extract_pdf_pages(data: bytes) -> list[str]:
    """Extract text from at most ``_MAX_PDF_PAGES`` pages. ``data`` is the
    caller's already-hashed, already-bounded (``_MAX_DOCUMENT_BYTES``)
    buffer from :func:`_extract_candidate`'s single read -- this function
    never reads a stream itself. The actual ``pdfplumber``/``pdfminer``
    parse -- the only step that can decompress a hostile PDF content
    stream into far more memory than its on-disk size implies -- runs
    isolated in a spawned child bound by hard address-space
    (``RLIMIT_AS``), CPU-time (``RLIMIT_CPU``), a single monotonic
    wall-clock deadline (:func:`_run_pdf_worker`), and result-byte
    (bounded ``recv_bytes``) limits. The spawned child runs
    :func:`phi_engine.pipeline._pdf_extract_worker.run` -- a private
    module that imports nothing from ``phi_engine`` -- rather than a
    function defined in this module, because ``intake_naming.py``'s own
    top-level imports (``config``, the security/model-routing
    chokepoints, ``intake_preflight``) pull in a multi-gigabyte virtual-
    address-space baseline before a single PDF byte is ever touched,
    which would make the address-space bound meaningless for a spawned
    process that had already imported this module first. Any worker
    limit, termination, or pathology -- a failed resource-limit
    application, an out-of-memory kill, a CPU-time kill, a hang past the
    wall-clock deadline, or an oversized/malformed reply -- collapses to
    the fixed ``_EvidenceLimitError``, never a raw exception. The reply
    itself crosses the process boundary as bounded, non-executable JSON,
    never ``pickle``.
    """
    parent_conn, child_conn = _PDF_WORKER_CONTEXT.Pipe(duplex=False)
    process = _PDF_WORKER_CONTEXT.Process(
        target=_pdf_extract_worker.run,
        args=(data, _MAX_PDF_PAGES, _PDF_WORKER_MAX_ADDRESS_BYTES, _PDF_WORKER_MAX_CPU_SECONDS, child_conn),
        daemon=True,
    )
    raw = _run_pdf_worker(process, parent_conn, child_conn)
    if raw is None:
        raise _EvidenceLimitError()
    return _decode_worker_reply(raw)


def _extract_csv_rows(stream: BinaryIO) -> list[list[str]]:
    with _bounded_csv_field_limit():
        try:
            text_stream = io.TextIOWrapper(stream, encoding="utf-8-sig", errors="strict", newline="")
            try:
                reader = csv.reader(text_stream)
                rows: list[list[str]] = []
                for _row_index, row in zip(range(_MAX_ROWS), reader):
                    # Validate every parsed field's UTF-8 byte size before
                    # any column slicing, so a field past column 20 cannot
                    # bypass the bound by never being retained.
                    for cell in row:
                        if len(cell.encode("utf-8")) > _MAX_CSV_FIELD_BYTES:
                            raise _EvidenceLimitError()
                    trimmed = row[:_MAX_COLS]
                    for cell in trimmed:
                        if len(cell) > _MAX_FRAGMENT_CODEPOINTS:
                            raise _EvidenceLimitError()
                    rows.append(trimmed)
            finally:
                text_stream.close()
        except _EvidenceLimitError:
            raise
        except (UnicodeDecodeError, csv.Error):
            raise _EvidenceLimitError() from None
    return rows


def _extract_xlsx_sheets(stream: BinaryIO) -> list[tuple[int, list[list[str]]]]:
    limits = support_files.DEFAULT_LIMITS
    try:
        try:
            # Pass 1: shared, directory-bounded ZipFile validation --
            # member count, per-member size, aggregate expansion, and
            # decompression ratio -- all against DEFAULT_LIMITS, via the
            # exact same primitive intake_preflight.py uses. Its central-
            # directory read is itself capped at max_zip_directory_bytes,
            # closing the allocation gap a bare zipfile.ZipFile(...) would
            # leave open.
            with support_files.open_bounded_zipfile(stream, limits) as (_zf, _guarded):
                pass
            # Pass 2: openpyxl performs its own internal ZipFile
            # construction regardless of what we already validated, so we
            # rewind and hand it the SAME stream wrapped in a fresh bound
            # (now the generous max_source_bytes -- the archive's
            # structure is already known-safe from pass 1) rather than an
            # unguarded raw stream.
            stream.seek(0)
            bounded_for_openpyxl = support_files._BoundedReader(stream, limits["max_source_bytes"])
            workbook = openpyxl.load_workbook(bounded_for_openpyxl, read_only=True, data_only=True, keep_links=False)
            try:
                sheets: list[tuple[int, list[list[str]]]] = []
                for sheet_index, worksheet in enumerate(workbook.worksheets[:_MAX_WORKBOOK_SHEETS], start=1):
                    rows: list[list[str]] = []
                    # Lazy XML parsing happens here, not at load_workbook()
                    # -- deliberately inside this same try/except so any
                    # failure during iteration is normalized identically.
                    for row in worksheet.iter_rows(
                        min_row=1, max_row=_MAX_ROWS, min_col=1, max_col=_MAX_COLS, values_only=True
                    ):
                        cells: list[str] = []
                        for value in row:
                            text = "" if value is None else str(value)
                            if len(text) > _MAX_FRAGMENT_CODEPOINTS:
                                raise _EvidenceLimitError()
                            cells.append(text)
                        rows.append(cells)
                    sheets.append((sheet_index, rows))
                return sheets
            finally:
                workbook.close()
        finally:
            stream.close()
    except _EvidenceLimitError:
        raise
    except Exception:
        # Catches support_files.BoundedZipMemberError, zipfile.BadZipFile,
        # openpyxl load/close failures, and any lazy-worksheet-iteration
        # failure (malformed cell references, corrupt shared strings,
        # etc.) uniformly -- never a raw exception.
        raise _EvidenceLimitError() from None


# --- canonical evidence JSON, incrementally bounded to 8192 bytes -------------------------


def _canonical_json(payload: Mapping[str, Any]) -> str:
    return json.dumps(payload, sort_keys=True, separators=(",", ":"), ensure_ascii=False)


def _encoded_len(payload: Mapping[str, Any]) -> int:
    return len(_canonical_json(payload).encode("utf-8"))


def _forms_fragments_from_docs(docs: dict[int, list[str]]) -> list[tuple[int, str]]:
    return [(index, page) for index in sorted(docs) for page in docs[index]]


def _dict_fragments_from_docs(
    docs: dict[int, dict[int, list[list[str]]]]
) -> list[tuple[int, int, list[str]]]:
    result: list[tuple[int, int, list[str]]] = []
    for index in sorted(docs):
        sheets = docs[index]
        for sheet_index in sorted(sheets):
            for row in sheets[sheet_index]:
                result.append((index, sheet_index, row))
    return result


def _forms_payload_dict(docs: dict[int, list[str]]) -> dict[str, Any]:
    return {"component": "forms", "documents": [{"index": i, "pages": docs[i]} for i in sorted(docs)]}


def _dict_payload_dict(docs: dict[int, dict[int, list[list[str]]]]) -> dict[str, Any]:
    documents = []
    for i in sorted(docs):
        sheets = docs[i]
        documents.append(
            {"index": i, "sheets": [{"index": s, "rows": sheets[s]} for s in sorted(sheets)]}
        )
    return {"component": "dictionary_mapping", "documents": documents}


def _combined_payload_dict(
    forms_docs: dict[int, list[str]], dict_docs: dict[int, dict[int, list[list[str]]]]
) -> dict[str, Any]:
    return {
        "component": "combined",
        "forms": _forms_payload_dict(forms_docs)["documents"],
        "dictionary_mapping": _dict_payload_dict(dict_docs)["documents"],
    }


def _grow_combined(
    forms_fragments: list[tuple[int, str]], dict_fragments: list[tuple[int, int, list[str]]], budget: int
) -> dict[str, Any]:
    """Rebuild the combined payload from the SAME already-bounded
    forms/dictionary_mapping fragments under one shared budget, appending
    forms fragments first then dictionary_mapping fragments, stopping at
    the first fragment that would exceed the cap (a single, deterministic,
    monotonic truncation boundary)."""
    forms_docs: dict[int, list[str]] = {}
    dict_docs: dict[int, dict[int, list[list[str]]]] = {}
    stopped = False

    for index, page in forms_fragments:
        if stopped:
            break
        trial = {key: list(value) for key, value in forms_docs.items()}
        trial.setdefault(index, []).append(page)
        if _encoded_len(_combined_payload_dict(trial, dict_docs)) > budget:
            stopped = True
            break
        forms_docs = trial

    for index, sheet_index, row in dict_fragments:
        if stopped:
            break
        trial = {key: {sk: list(sv) for sk, sv in value.items()} for key, value in dict_docs.items()}
        sheets = dict(trial.get(index, {}))
        sheets[sheet_index] = [*sheets.get(sheet_index, []), row]
        trial[index] = sheets
        if _encoded_len(_combined_payload_dict(forms_docs, trial)) > budget:
            stopped = True
            break
        dict_docs = trial

    return _combined_payload_dict(forms_docs, dict_docs)
